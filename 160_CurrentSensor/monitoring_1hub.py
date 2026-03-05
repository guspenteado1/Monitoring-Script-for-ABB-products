# For this script to work, it's necessary to run the following command on the Windows terminal:
# pip install streamlit pymodbus openpyxl

# After installing the necessary packages, on the terminal you need to navigate to the folder in which the script is located and run the following line:
# streamlit run monitoring_1hub.py

# The only difference between monitoring_1hub.py and monitoring_2hub.py are the lines 37 and 337

import os
import re
import time
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from struct import pack, unpack
from copy import copy
import zipfile

import pandas as pd
import streamlit as st
from pymodbus.client import ModbusTcpClient
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
import openpyxl.chart.axis

# ---------------------- CONFIGURATION SECTION ----------------------

# This is the only section that needs editing if you want to change the variables, connection and sensors

# Modbus connection parameters
# IP = '192.168.0.217'                                                                          # Lab's WiFi network (slowest)
IP = '10.39.37.112'                                                                           # Ethernet connection (fastest)
# IP = '192.168.2.1'                                                                            # SCU200's WiFi network
PORT = 502                                                                                    # Depends on the port configured on the Web UI
SLAVE_IDS = list(range(1, 33)) + list(range(33, 65)) + list(range(97, 129))                   # IDs of all Modbus slave devices
POLL_INTERVAL = 3                                                                             # Time between polls in seconds
EXCEL_FILE_FINAL = "modbus_poll_log.xlsx"
EXCEL_FILE_BACKUP = "backup_poll_data.xlsx"

# List of variables to read from each slave, according to the Modbus Map
# The first value is the address of the variable, and the second value is the size of the variable
# To select which variables you don't want to monitor, just comment the corrisponding line (the fewer variables, the faster is the reading)
VARIABLES = [
#   (0, 3, "swVersion"),
#   (3, 2, "hwVersion"),
#   (5, 4, "serialNumber"),
    (9, 1, "currentTrms"),
    (10, 1, "currentAc"),
    (11, 1, "currentDc"),
#   (22, 1, "deviceStatus"),
    (100, 2, "activePowerTotal"),
    (102, 2, "activeEnergyTotal")
]

# Multiplier factors for specific variables
SCALE_FACTORS = {
    "currentTrms": 0.01,
    "currentAc": 0.01,
    "currentDc": 0.01,
    "activePowerTotal": 1,
    "activeEnergyTotal": 1
}

# ---------------------- DECODING FUNCTIONS ----------------------

# Remove non-printable characters from string or list
def clean_value(value):
    if isinstance(value, str):
        cleaned = re.sub(r'[^\x20-\x7E]', '', value)
        return cleaned
    elif isinstance(value, list):
        return [clean_value(v) for v in value]
    else:
        return value

# Decode software/hardware version from registers
def decode_version(registers):                                      # Each register contains version info in high nibble (4 bits)
    return '.'.join(str((reg >> 12) & 0xF) for reg in registers)    # Shift right 12 bits and mask to get the version number part

# Decode serial number from 4 registers (64 bits)
def decode_serial_number(registers):
    try:
        b = b''.join(pack('>H', r) for r in registers)              # Pack 4 registers (16 bits each) into bytes in big-endian order
        serial_num = unpack('>Q', b)[0]                             # Unpack as a 64-bit unsigned integer (Q)
        return str(serial_num)
    except Exception as e:
        return None

# Decode 32-bit signed integer
def decode_32bit_signed(registers):
    try:
        b = pack('>HH', registers[0], registers[1])                 # Pack two 16-bit registers into bytes
        return unpack('>i', b)[0]                                   # Unpack as signed 32-bit int (i)
    except Exception: 
        return None

# Decode 32-bit unsigned integer
def decode_32bit_unsigned(registers):
    try:
        b = pack('>HH', registers[0], registers[1])
        return unpack('>I', b)[0]
    except Exception:
        return None

# Decode 16-bit signed integer
def decode_16bit_signed(register):
    if register >= 0x8000:
        return register - 0x10000
    else:
        return register

# Convert device status to 8-bit binary string (the meaning of each byte is on the Modbus Map)
def decode_device_status(value):
    try:
        return format(value, '08b')
    except:
        return value

# Decode 32-bit little endian integer from two 16-bit registers
def decode_32bit_little_endian(regs):                               # Combines two 16-bit registers into a single 32-bit integer, considering the low word first (little endian)
    if len(regs) < 2:
        return None
    return (regs[1] << 16) | regs[0]

# Join list into a string
def flatten_value(value):
    if isinstance(value, list):
        return ','.join(str(v) for v in value)
    return value

# ---------------------- SENSOR POLLING FUNCTION ----------------------

# Read all defined Modbus variables from a specific slave
def read_slave_with_client(slave_id, thread_id, client):
    values = []
    thread_prefix = f"[Thread {thread_id}] "

    for addr, size, name in VARIABLES:
        try:
            resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)           # Read the specified number of registers starting at the given address
            if resp.isError() or not hasattr(resp, 'registers'):
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | No response or error")
                values.append(None)
                continue

            regs = resp.registers
            print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Data: {regs}")

            if name in ("swVersion", "hwVersion"):                                                   # Decode the value based on its name/typ
                decoded = decode_version(regs)
                values.append(decoded)

            elif name == "serialNumber":
                decoded = decode_serial_number(regs)
                values.append(decoded)

            elif name in ("activePowerTotal", "activeEnergyTotal"):
                raw = decode_32bit_little_endian(regs)
                values.append(raw * SCALE_FACTORS[name])

            elif name == "deviceStatus":
                values.append(decode_device_status(regs[0]))

            elif name == "currentDc":
                signed_val = decode_16bit_signed(regs[0]) if regs else None
                values.append(None if signed_val is None else signed_val * SCALE_FACTORS[name])

            elif name in SCALE_FACTORS:
                values.append(None if not regs else regs[0] * SCALE_FACTORS[name])

            else:
                values.append(flatten_value(regs))

        except Exception as e:
            print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
            values.append(None)

    return values

# ---------------------- SAVE TO EXCEL ----------------------

# The data is registered on the internal memory and then written in an Excel file
def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
    wb = Workbook()
    wb.remove(wb.active)
    for slave_id in SLAVE_IDS:
        ws = wb.create_sheet(title=f"Slave_{slave_id}")
        headers = ['timestamp'] + [var[2] for var in VARIABLES]
        ws.append(headers)

    for poll_entry in all_poll_data:
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']
        for slave_id, values in data.items():
            ws = wb[f"Slave_{slave_id}"]
            safe_values = [clean_value(v) for v in values]
            ws.append([timestamp] + safe_values)

    wb.save(filename)

# Variables to be plotted on the graphs on excel
VARIABLES_TO_PLOT = ["currentTrms", "currentAc", "currentDc", "activePowerTotal", "activeEnergyTotal"]

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
    units = {
        "currentTrms": "(A)",
        "currentAc": "(A)",
        "currentDc": "(A)",
        "activePowerTotal": "(W)",
        "activeEnergyTotal": "(kWh)"
    }

    chart = ScatterChart()
    chart.title = f"{var_name} over time - ID {slave_id}"
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = f"{var_name} {units.get(var_name, '')}".strip()
    chart.legend = None

    chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    title_with_unit = f"{var_name} {units.get(var_name, '')}".strip()
    series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.smooth = True
    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'H'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)

# Add charts to each Excel sheet using the custom create_scatter_chart function
def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
    wb = load_workbook(filename)

    for slave_id in SLAVE_IDS:
        sheet_name = f"Slave_{slave_id}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        num_rows = ws.max_row

        chart_index = 0
        timestamp_col = 1

        for var in VARIABLES_TO_PLOT:
            if var not in headers:
                continue

            var_col = headers.index(var) + 1
            create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
            chart_index += 1

    wb.save(filename)

# ---------------------- MULTITHREADING POLLING FOR EACH CMS BUS PORT ----------------------

# Multithreaded polling of all slaves
def poll_slaves_multithread(stop_event, all_poll_data):                                         # Each thread polls a range of slave IDs
    def poll_slave_range(start_id, end_id, all_data, lock, thread_id):
        client = ModbusTcpClient(IP, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect")
            return

        def task(slave_id):
            values = read_slave_with_client(slave_id, thread_id, client)
            with lock:
                all_data[slave_id] = values

        with ThreadPoolExecutor(max_workers=32) as executor:
            executor.map(task, range(start_id, end_id + 1))

        client.close()

    # Main polling loop
    while not stop_event.is_set():
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()                                                                 # Lock to synchronize access to all_data dictionary
        threads = []

        # Divide polling among 3 threads, one for each flat cable/CMS BUS port
        ranges = [(1, 32), (33, 64), (97, 128)]

        for i, (start, end) in enumerate(ranges):
            t = threading.Thread(target=poll_slave_range, args=(start, end, all_data, lock, i+1))
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        all_poll_data.append({'timestamp': timestamp, 'data': all_data})

        # Wait until next polling interval
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        print(f"\n Completed polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

def poll_slaves_multithread_once(stop_event, all_poll_data):
    if stop_event.is_set():
        return

    start_time = time.time()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    all_data = {}
    lock = threading.Lock()
    threads = []

    ranges = [(1, 32), (33, 64), (97, 128)]

    def poll_slave_range(start_id, end_id, all_data, lock, thread_id):
        client = ModbusTcpClient(IP, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect")
            return

        def task(slave_id):
            values = read_slave_with_client(slave_id, thread_id, client)
            with lock:
                all_data[slave_id] = values

        with ThreadPoolExecutor(max_workers=32) as executor:
            executor.map(task, range(start_id, end_id + 1))

        client.close()

    for i, (start, end) in enumerate(ranges):
        t = threading.Thread(target=poll_slave_range, args=(start, end, all_data, lock, i+1))
        threads.append(t)
        t.start()

    for t in threads:
        t.join()

    all_poll_data.append({'timestamp': timestamp, 'data': all_data})

    elapsed = time.time() - start_time
    sleep_time = max(0, POLL_INTERVAL - elapsed)
    print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
    time.sleep(sleep_time)

def periodic_polling_loop(stop_event, all_poll_data):
    while not stop_event.is_set():
        start_time = time.time()
        one_minute = 60
        cycle_duration = 30 * 60

        while time.time() - start_time < one_minute and not stop_event.is_set():
            poll_slaves_multithread_once(stop_event, all_poll_data)

        if stop_event.is_set():
            break

        try:
            print(f"Saving backup with {len(all_poll_data)} readings...")
            save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_BACKUP)
        except Exception as e:
            print(f"Error on backup: {e}")

        print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
        for _ in range(int(cycle_duration)):
            if stop_event.is_set():
                break
            time.sleep(1)

# ---------------------- SPLIT EXCEL FUNCTIONS ----------------------

def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
    if not os.path.exists(path_file_excel):
        st.error(f"Error: File '{path_file_excel}' not found.")
        return None

    try:
        if not os.path.exists(name_output_folder):
            os.makedirs(name_output_folder)
        else:
            for file_name in os.listdir(name_output_folder):
                path_old_file = os.path.join(name_output_folder, file_name)
                if os.path.isfile(path_old_file):
                    os.remove(path_old_file)
            st.info(f"Cleaning folder '{name_output_folder}'.")

        source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in source_workbook.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue

            sheet = source_workbook[sheet_name]

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            for row in sheet.iter_rows(values_only=True):
                new_ws.append(row)

            num_rows = new_ws.max_row
            num_cols = new_ws.max_column

            if num_rows < 2 or num_cols < 2:
                st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                continue

            timestamp_col = 1

            for idx, var_name in enumerate(VARIABLES_TO_PLOT):
                if var_name in [cell.value for cell in new_ws[1]]:
                    var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                    create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
            output_filepath = os.path.join(name_output_folder, output_filename)

            new_wb.save(output_filepath)
            generated_files.append(output_filepath)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
        complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

        with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All files were zipped into '{complete_zip_path}'")
        return complete_zip_path

    except Exception as e:
        st.error(f"Error splitting and generating charts: {e}")
        return None

def copy_sheet_without_charts(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col_dim in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_dim] = copy(source_sheet.column_dimensions[col_dim])
    for row_dim in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row_dim] = copy(source_sheet.row_dimensions[row_dim])

def extract_slave_id(sheet_name):
    match = re.search(r'\d+', sheet_name)
    return int(match.group(0)) if match else None

def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
    if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
        st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
        return None

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    else:
        for file_name in os.listdir(output_folder):
            path_to_remove = os.path.join(output_folder, file_name)
            if os.path.isfile(path_to_remove):
                os.remove(path_to_remove)
        st.info(f"Folder '{output_folder}' cleaned.")

    wb_modbus = openpyxl.load_workbook(path_modbus_log)
    wb_backup = openpyxl.load_workbook(path_backup_data)
    
    generated_files = []

    columns_for_df = ["timestamp"] + [var[2] for var in VARIABLES]

    for i, slave_id in enumerate(selected_slave_ids):
        sheet_name = f"Slave_{slave_id}"
        if status_text:
            status_text.text(f"Processing Slave_{slave_id}...")
        if sheet_name not in wb_modbus.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
            continue
        if sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
            continue

        modbus_sheet = wb_modbus[sheet_name]
        backup_sheet = wb_backup[sheet_name]

        modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
        backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

        modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
        backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

        merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
        merged_data.dropna(subset=["timestamp"], inplace=True)
        merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
        merged_data.sort_values("timestamp", inplace=True)
        merged_data.reset_index(drop=True, inplace=True)

        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        new_sheet = new_wb.create_sheet(title=sheet_name)

        headers = [cell.value for cell in modbus_sheet[1]]
        new_sheet.append(headers)

        for r_idx, row in merged_data.iterrows():
            new_sheet.append(list(row.values))

        num_rows = new_sheet.max_row
        current_slave_id = extract_slave_id(sheet_name)

        for chart_index, var_name in enumerate(VARIABLES_TO_PLOT):
            if var_name in headers:
                col_idx = headers.index(var_name) + 1
                create_scatter_chart(
                    new_sheet,
                    timestamp_col=1,
                    var_col=col_idx,
                    num_rows=num_rows,
                    var_name=var_name,
                    slave_id=current_slave_id,
                    chart_index=chart_index
                )
            else:
                st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

        safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
        new_wb.save(path_output)
        generated_files.append(path_output)
        
        if progress_bar and total_slaves > 0:
            progress_percent = int(((i + 1) / total_slaves) * 100)
            progress_bar.progress(progress_percent)

    if not generated_files:
        st.warning("No individual files were generated for the selected slaves.")
        return None

    zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_to_add in generated_files:
            zipf.write(file_to_add, os.path.basename(file_to_add))

    st.success(f"All merged individual files have been zipped into: {zip_path}")
    return zip_path

def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
    if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
        st.error("Excel files (final or backup) not found for comparison.")
        return

    wb_final = load_workbook(excel_file_final, data_only=True)
    wb_backup = load_workbook(excel_file_backup, data_only=True)

    needs_merge = False
    for slave_id in selected_slave_ids:
        sheet_name = f"Slave_{slave_id}"
        if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
            continue

        ws_final = wb_final[sheet_name]
        ws_backup = wb_backup[sheet_name]

        timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
        timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

        if not timestamps_backup.issubset(timestamps_final):
            needs_merge = True
            break
    
    return needs_merge

# ---------------------- CREATION OF THE DASHBOARD ----------------------

# Streamlit web interface
st.title("96 Current Sensors Setup")

# Initialize session state variables
if "polling" not in st.session_state:
    st.session_state.polling = False
    st.session_state.stop_event = threading.Event()
    st.session_state.thread = None
    st.session_state.all_poll_data = []

# Start monitoring button
if st.button("Start Monitoring", disabled=st.session_state.polling):
    st.session_state.all_poll_data = []
    st.session_state.stop_event.clear()
    st.session_state.thread = threading.Thread(target=periodic_polling_loop, args=(st.session_state.stop_event, st.session_state.all_poll_data),
    daemon=True
)

    st.session_state.thread.start()
    st.session_state.polling = True
    st.success("Monitoring started!")

# Stop monitoring button
if st.button("Stop Monitoring"):
    if st.session_state.polling:
        st.session_state.stop_event.set()
        st.session_state.thread.join()
        st.session_state.polling = False

        with st.spinner("Saving data and generating graphs..."):
            save_all_data_to_excel(st.session_state.all_poll_data)
            add_charts_to_excel()

        st.success("Monitoring stopped and graphs successfully created")
    else:
        st.warning("Monitoring is not running.")

## Split / Merge Excel Files by Slave
if not os.path.exists(EXCEL_FILE_FINAL):
    st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
else:
    all_possible_slave_ids = sorted(list(set(SLAVE_IDS))) # Ensures unique and sorted IDs
    
    display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

    if "All" in st.session_state.get('selected_display_options', []):
        selected_display_options = st.multiselect(
            "Choose slaves to separate:",
            display_options,
            default=["All"],
            disabled=True
        )
    else:
        selected_display_options = st.multiselect(
            "Choose slaves to separate:",
            display_options,
            default=[]
        )
    
    st.session_state.selected_display_options = selected_display_options

    selected_slave_ids_numeric = []
    if "All" in selected_display_options:
        selected_slave_ids_numeric = all_possible_slave_ids
    elif selected_display_options:
        for ds_id in selected_display_options:
            match = re.search(r'\d+', ds_id)
            if match:
                selected_slave_ids_numeric.append(int(match.group(0)))

    if st.button("Generate Individual Files"):
        if not selected_slave_ids_numeric:
            st.warning("Please select at least one slave or 'All'.")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            total_slaves_to_process = len(selected_slave_ids_numeric)
            zip_path = None

            if not os.path.exists(EXCEL_FILE_BACKUP):
                st.warning("Backup file not found. Performing only the split of the final Excel.")
                with st.spinner("Splitting Excel and generating charts..."):
                    zip_path = split_and_generate_charts_for_slaves(
                        EXCEL_FILE_FINAL, 
                        selected_slave_ids_numeric, 
                        progress_bar=progress_bar, 
                        status_text=status_text, 
                        total_slaves=total_slaves_to_process
                    )
            else:
                st.info("Comparing timestamps between final Excel and backup...")
                needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                if needs_merge:
                    st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                    with st.spinner("Merging data from backup and final, and generating individual files..."):
                        zip_path = merge_data_and_generate_charts_for_slaves(
                            EXCEL_FILE_FINAL, 
                            EXCEL_FILE_BACKUP, 
                            selected_slave_ids_numeric, 
                            progress_bar=progress_bar, 
                            status_text=status_text, 
                            total_slaves=total_slaves_to_process
                        )
                else:
                    st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                    with st.spinner("Splitting Excel and generating charts..."):
                        zip_path = split_and_generate_charts_for_slaves(
                            EXCEL_FILE_FINAL, 
                            selected_slave_ids_numeric, 
                            progress_bar=progress_bar, 
                            status_text=status_text, 
                            total_slaves=total_slaves_to_process
                        )
            
            if zip_path:
                status_text.text("Processing complete!")
                st.success("Individual files generated and available for download.")

## Example of Latest Reading (First 5 Slaves)
if os.path.exists(EXCEL_FILE_FINAL):
    wb = Workbook()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_FILE_FINAL, read_only=True)
    except Exception:
        pass
    st.subheader("Example of Latest Reading (First 5 Slaves)")
    displayed_slaves = 0
    for sid in sorted(list(SLAVE_IDS)):
        if displayed_slaves >= 5:
            break
        if f"Slave_{sid}" not in wb.sheetnames:
            continue
        ws = wb[f"Slave_{sid}"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            st.markdown(f"**Slave {sid}**")
            df = {rows[0][i]: rows[-1][i] for i in range(len(rows[0]))}
            st.write(df)
            displayed_slaves += 1