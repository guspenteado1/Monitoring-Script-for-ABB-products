import os
import re
import time
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from struct import pack, unpack
from copy import copy
import zipfile

import pandas as pd
import streamlit as st
from pymodbus.client import ModbusTcpClient
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl.chart.axis

# ---------------------- CONFIGURATION SECTION ----------------------

PORT = 502
SENSOR_IPS = [f"192.168.1.{i}" for i in list(range(2, 11))+ list(range(14, 21))]   # IPs of all Modbus slave devices
POLL_INTERVAL = 60
EXCEL_FILE_FINAL = "modbus_poll_log.xlsx"
EXCEL_FILE_BACKUP = "backup_poll_data.xlsx"

DEVICE_TYPE = {
    "M4M": {
        "variables": [
            (20480, 4, "activeImportEnergyTotal", 3),
            (20484, 4, "activeExportEnergyTotal", 3),
            (20488, 4, "activeNetEnergyTotal", 3),
            (20492, 4, "reactiveImportEnergyTotal", 3),
            (20496, 4, "reactiveExportEnergyTotal", 3),
            (20500, 4, "reactiveNetEnergyTotal", 3),
            (20504, 4, "apparentImportEnergyTotal", 3),
            (20508, 4, "apparentExportEnergyTotal", 3),
            (20512, 4, "apparentNetEnergyTotal", 3),
            # (20516, 4, "CO2conversionFactor", 3),
            # (20532, 4, "currencyConversionFactor", 3),
            # (20848, 4, "activeImpEnergyTotTariff1", 3),
            # (20852, 4, "activeImpEnergyTotTariff2", 3),
            # (20856, 4, "activeImpEnergyTotTariff3", 3),
            # (20860, 4, "activeImpEnergyTotTariff4", 3),
            # (20864, 4, "activeImpEnergyTotTariff5", 3),
            # (20868, 4, "activeImpEnergyTotTariff6", 3),
            # (20880, 4, "activeExpEnergyTotTariff1", 3),
            # (20884, 4, "activeExpEnergyTotTariff2", 3),
            # (20888, 4, "activeExpEnergyTotTariff3", 3),
            # (20892, 4, "activeImpEnergyTotTariff4", 3),
            # (20896, 4, "activeImpEnergyTotTariff5", 3),
            # (20900, 4, "activeImpEnergyTotTariff6", 3),
            # (20912, 4, "reactiveImpEnergyTotTariff1", 3),
            # (20916, 4, "reactiveImpEnergyTotTariff2", 3),
            # (20920, 4, "reactiveImpEnergyTotTariff3", 3),
            # (20924, 4, "reactiveImpEnergyTotTariff4", 3),
            # (20928, 4, "reactiveImpEnergyTotTariff5", 3),
            # (20932, 4, "reactiveImpEnergyTotTariff6", 3),
            # (20944, 4, "reactiveExpEnergyTotTariff1", 3),
            # (20948, 4, "reactiveExpEnergyTotTariff2", 3),
            # (20952, 4, "reactiveExpEnergyTotTariff3", 3),
            # (20956, 4, "reactiveExpEnergyTotTariff4", 3),
            # (20960, 4, "reactiveExpEnergyTotTariff5", 3),
            # (20964, 4, "reactiveExpEnergyTotTariff6", 3),
            (21600, 4, "activeImpEnergyL1", 3),
            (21604, 4, "activeImpEnergyL2", 3),
            (21608, 4, "activeImpEnergyL3", 3),
            (21612, 4, "activeExpEnergyL1", 3),
            (21616, 4, "activeExpEnergyL2", 3),
            (21620, 4, "activeExpEnergyL3", 3),
            (21624, 4, "activeNetEnergyL1", 3),
            (21628, 4, "activeNetEnergyL2", 3),
            (21632, 4, "activeNetEnergyL3", 3),
            (21636, 4, "reactiveImpEnergyL1", 3),
            (21640, 4, "reactiveImpEnergyL2", 3),
            (21644, 4, "reactiveImpEnergyL3", 3),
            (21648, 4, "reactiveExpEnergyL1", 3),
            (21652, 4, "reactiveExpEnergyL2", 3),
            (21656, 4, "reactiveExpEnergyL3", 3),
            (21660, 4, "reactiveNetEnergyL1", 3),
            (21664, 4, "reactiveNetEnergyL2", 3),
            (21668, 4, "reactiveNetEnergyL3", 3),
            (21672, 4, "apparentImpEnergyL1", 3),
            (21676, 4, "apparentImpEnergyL2", 3),
            (21680, 4, "apparentImpEnergyL3", 3),
            (21684, 4, "apparentExpEnergyL1", 3),
            (21688, 4, "apparentExpEnergyL2", 3),
            (21692, 4, "apparentExpEnergyL3", 3),
            (21696, 4, "apparentNetEnergyL1", 3),
            (21700, 4, "apparentNetEnergyL2", 3),
            (21704, 4, "apparentNetEnergyL3", 3),
            # (23296, 2, "threePhaseSystemVoltage", 3),
            # (23298, 2, "phaseVoltageL1", 3),
            # (23300, 2, "phaseVoltageL2", 3),
            # (23302, 2, "phaseVoltageL3", 3),
            # (23304, 2, "lineVoltageL1L2", 3),
            # (23306, 2, "lineVoltageL3L2", 3),
            # (23308, 2, "lineVoltageL1L3", 3),
            # (23310, 2, "threePhaseSystemCurrent", 3),
            (23312, 2, "currentL1", 3),
            (23314, 2, "currentL2", 3),
            (23316, 2, "currentL3", 3),
            (23318, 2, "currentN", 3),
            (23322, 2, "activePowerTotal", 3),
            (23324, 2, "activePowerL1", 3),
            (23326, 2, "activePowerL2", 3),
            (23328, 2, "activePowerL3", 3),
            (23330, 2, "reactivePowerTotal", 3),
            (23332, 2, "reactivePowerL1", 3),
            (23334, 2, "reactivePowerL2", 3),
            (23336, 2, "reactivePowerL3", 3),
            (23338, 2, "apparentPowerTotal", 3),
            (23340, 2, "apparentPowerL1", 3),
            (23342, 2, "apparentPowerL2", 3),
            (23344, 2, "apparentPowerL3", 3)
            # (23346, 1, "frequency", 3),
            # (23347, 1, "phaseAnglePowerTotal", 3),
            # (23348, 1, "phaseAnglePowerL1", 3),
            # (23349, 1, "phaseAnglePowerL2", 3),
            # (23350, 1, "phaseAnglePowerL3", 3),
            # (23351, 1, "phaseAngleVoltageL1", 3),
            # (23352, 1, "phaseAngleVoltageL2", 3),
            # (23353, 1, "phaseAngleVoltageL3", 3),
            # (23357, 1, "phaseAngleCurrentL1", 3),
            # (23358, 1, "phaseAngleCurrentL2", 3),
            # (23359, 1, "phaseAngleCurrentL3", 3),
            # (23360, 1, "powerFactorTotal", 3),
            # (23361, 1, "powerFactorL1", 3),
            # (23362, 1, "powerFactorL2", 3),
            # (23363, 1, "powerFactorL3", 3),
            # (23364, 1, "currentQuadrantTotal", 3),
            # (23365, 1, "currentQuadrantL1", 3),
            # (23366, 1, "currentQuadrantL2", 3),
            # (23367, 1, "currentQuadrantL3", 3),
            # (23368, 1, 'cosphiTotal', 3),
            # (23369, 1, 'cosphiL1', 3),
            # (23370, 1, 'cosphiL2', 3),
            # (23371, 1, 'cosphiL3', 3),
            # (23508, 2, "avgCurrentL1", 3),
            # (23510, 2, "avgCurrentL2", 3),
            # (23512, 2, "avgCurrentL3", 3),
            # (23514, 2, "avgCurrentNeutral", 3),
            # (23516, 2, "avgVoltageL1", 3),
            # (23518, 2, "avgVoltageL2", 3),
            # (23520, 2, "avgVoltageL3", 3),
            # (23522, 2, "avgVoltageL1L2", 3),
            # (23524, 2, "avgVoltageL2L3", 3),
            # (23526, 2, "avgVoltageL1L3", 3),
            # (23528, 2, "avgActivePowerTotal", 3),
            # (23530, 2, "avgActivePowerL1", 3),
            # (23532, 2, "avgActivePowerL2", 3),
            # (23534, 2, "avgActivePowerL3", 3),
            # (23536, 2, "avgReactivePowerTotal", 3),
            # (23538, 2, "avgReactivePowerL1", 3),
            # (23540, 2, "avgReactivePowerL2", 3),
            # (23542, 2, "avgReactivePowerL3", 3),
            # (23544, 2, "avgApparentPowerTotal", 3),
            # (23546, 2, "avgApparentPowerL1", 3),
            # (23548, 2, "avgApparentPowerL2", 3),
            # (23550, 2, "avgApparentPowerL3", 3),
            # (23568, 2, "maxCurrentL1", 3),
            # (23570, 2, "maxCurrentL2", 3),
            # (23572, 2, "maxCurrentL3", 3),
            # (23574, 2, "maxCurrentNeutral", 3),
            # (23576, 2, "maxVoltageL1", 3),
            # (23578, 2, "maxVoltageL2", 3),
            # (23580, 2, "maxVoltageL3", 3),
            # (23582, 2, "maxVoltageL1L2", 3),
            # (23584, 2, "maxVoltageL2L3", 3),
            # (23586, 2, "maxVoltageL1L3", 3),
            # (23588, 2, "maxActivePowerTotal", 3),
            # (23590, 2, "maxActivePowerL1", 3),
            # (23592, 2, "maxActivePowerL2", 3),
            # (23594, 2, "maxActivePowerL3", 3),
            # (23596, 2, "maxReactivePowerTotal", 3),
            # (23598, 2, "maxReactivePowerL1", 3),
            # (23600, 2, "maxReactivePowerL2", 3),
            # (23602, 2, "maxReactivePowerL3", 3),
            # (23604, 2, "maxApparentPowerTotal", 3),
            # (23606, 2, "maxApparentPowerL1", 3),
            # (23608, 2, "maxApparentPowerL2", 3),
            # (23610, 2, "maxApparentPowerL3", 3),
            # (23688, 2, "minCurrentL1", 3),
            # (23690, 2, "minCurrentL2", 3),
            # (23692, 2, "minCurrentL3", 3),
            # (23694, 2, "minCurrentNeutral", 3),
            # (23696, 2, "minVoltageL1", 3),
            # (23698, 2, "minVoltageL2", 3),
            # (23700, 2, "minVoltageL3", 3),
            # (23702, 2, "minVoltageL1L2", 3),
            # (23704, 2, "minVoltageL2L3", 3),
            # (23706, 2, "minVoltageL1L3", 3),
            # (23708, 2, "minActivePowerTotal", 3),
            # (23710, 2, "minActivePowerL1", 3),
            # (23712, 2, "minActivePowerL2", 3),
            # (23714, 2, "minActivePowerL3", 3),
            # (23716, 2, "minReactivePowerTotal", 3),
            # (23718, 2, "minReactivePowerL1", 3),
            # (23720, 2, "minReactivePowerL2", 3),
            # (23722, 2, "minReactivePowerL3", 3),
            # (23724, 2, "minApparentPowerTotal", 3),
            # (23726, 2, "minApparentPowerL1", 3),
            # (23728, 2, "minApparentPowerL2", 3),
            # (23730, 2, "minApparentPowerL3", 3),
            # (35072, 5, "serialNumber", 3),
            # (35080, 8, "firmwareVersion", 3),
            # (36064, 2, "CO2conversionFactor", 3),
            # (36066, 2, "currencyConversionFactor", 3),
            # (51994, 2, "activePowerTotalScaler", 3),
            # (51996, 2, "activePowerL1Scaler", 3),
            # (51998, 2, "activePowerL2Scaler", 3),
            # (52000, 2, "activePowerL3Scaler", 3),
            # (52002, 2, "reactivePowerTotalScaler", 3),
            # (52004, 2, "reactivePowerL1Scaler", 3),
            # (52006, 2, "reactivePowerL2Scaler", 3),
            # (52008, 2, "reactivePowerL3Scaler", 3),
            # (52010, 2, "apparentPowerTotalScaler", 3),
            # (52012, 2, "apparentPowerL1Scaler", 3),
            # (52014, 2, "apparentPowerL2Scaler", 3),
            # (52016, 2, "apparentPowerL3Scaler", 3),
            # (52200, 2, "avgActivePowerTotalScaler", 3),
            # (52202, 2, "avgActivePowerL1Scaler", 3),
            # (52204, 2, "avgActivePowerL2Scaler", 3),
            # (52206, 2, "avgActivePowerL3Scaler", 3),
            # (52208, 2, "avgReactivePowerTotalScaler", 3),
            # (52210, 2, "avgReactivePowerL1Scaler", 3),
            # (52212, 2, "avgReactivePowerL2Scaler", 3),
            # (52214, 2, "avgReactivePowerL3Scaler", 3),
            # (52216, 2, "avgApparentPowerTotalScaler", 3),
            # (52218, 2, "avgApparentPowerL1Scaler", 3),
            # (52220, 2, "avgApparentPowerL2Scaler", 3),
            # (52222, 2, "avgApparentPowerL3Scaler", 3),
            # (52260, 2, "maxActivePowerTotalScaler", 3),
            # (52262, 2, "maxActivePowerL1Scaler", 3),
            # (52264, 2, "maxActivePowerL2Scaler", 3),
            # (52266, 2, "maxActivePowerL3Scaler", 3),
            # (52268, 2, "maxReactivePowerTotalScaler", 3),
            # (52270, 2, "maxReactivePowerL1Scaler", 3),
            # (52272, 2, "maxReactivePowerL2Scaler", 3),
            # (52274, 2, "maxReactivePowerL3Scaler", 3),
            # (52276, 2, "maxApparentPowerTotalScaler", 3),
            # (52278, 2, "maxApparentPowerL1Scaler", 3),
            # (52280, 2, "maxApparentPowerL2Scaler", 3),
            # (52282, 2, "maxApparentPowerL3Scaler", 3),
            # (52380, 2, "minActivePowerTotalScaler", 3),
            # (52382, 2, "minActivePowerL1Scaler", 3),
            # (52384, 2, "minActivePowerL2Scaler", 3),
            # (52386, 2, "minActivePowerL3Scaler", 3),
            # (52388, 2, "minReactivePowerTotalScaler", 3),
            # (52390, 2, "minReactivePowerL1Scaler", 3),
            # (52392, 2, "minReactivePowerL2Scaler", 3),
            # (52394, 2, "minReactivePowerL3Scaler", 3),
            # (52396, 2, "minApparentPowerTotalScaler", 3),
            # (52398, 2, "minApparentPowerL1Scaler", 3),
            # (52400, 2, "minApparentPowerL2Scaler", 3),
            # (52402, 2, "minApparentPowerL3Scaler", 3)
        ],
        
        "scale_factors": {
            "activeImportEnergyTotal": 0.01,
            "activeExportEnergyTotal": 0.01,
            "activeNetEnergyTotal": 0.01,
            "reactiveImportEnergyTotal": 0.01,
            "reactiveExportEnergyTotal": 0.01,
            "reactiveNetEnergyTotal": 0.01,
            "apparentImportEnergyTotal": 0.01,
            "apparentExportEnergyTotal": 0.01,
            "apparentNetEnergyTotal": 0.01,
            # "CO2conversionFactor": 0.001,
            # "currencyConversionFactor": 0.001,
            # "activeImpEnergyTotTariff1": 0.01,
            # "activeImpEnergyTotTariff2": 0.01,
            # "activeImpEnergyTotTariff3": 0.01,
            # "activeImpEnergyTotTariff4": 0.01,
            # "activeImpEnergyTotTariff5": 0.01,
            # "activeImpEnergyTotTariff6": 0.01,
            # "activeExpEnergyTotTariff1": 0.01,
            # "activeExpEnergyTotTariff2": 0.01,
            # "activeExpEnergyTotTariff3": 0.01,
            # "activeImpEnergyTotTariff4": 0.01,
            # "activeImpEnergyTotTariff5": 0.01,
            # "activeImpEnergyTotTariff6": 0.01,
            # "reactiveImpEnergyTotTariff1": 0.01,
            # "reactiveImpEnergyTotTariff2": 0.01,
            # "reactiveImpEnergyTotTariff3": 0.01,
            # "reactiveImpEnergyTotTariff4": 0.01,
            # "reactiveImpEnergyTotTariff5": 0.01,
            # "reactiveImpEnergyTotTariff6": 0.01,
            # "reactiveExpEnergyTotTariff1": 0.01,
            # "reactiveExpEnergyTotTariff2": 0.01,
            # "reactiveExpEnergyTotTariff3": 0.01,
            # "reactiveExpEnergyTotTariff4": 0.01,
            # "reactiveExpEnergyTotTariff5": 0.01,
            # "reactiveExpEnergyTotTariff6": 0.01,
            "activeImpEnergyL1": 0.01,
            "activeImpEnergyL2": 0.01,
            "activeImpEnergyL3": 0.01,
            "activeExpEnergyL1": 0.01,
            "activeExpEnergyL2": 0.01,
            "activeExpEnergyL3": 0.01,
            "activeNetEnergyL1": 0.01,
            "activeNetEnergyL2": 0.01,
            "activeNetEnergyL3": 0.01,
            "reactiveImpEnergyL1": 0.01,
            "reactiveImpEnergyL2": 0.01,
            "reactiveImpEnergyL3": 0.01,
            "reactiveExpEnergyL1": 0.01,
            "reactiveExpEnergyL2": 0.01,
            "reactiveExpEnergyL3": 0.01,
            "reactiveNetEnergyL1": 0.01,
            "reactiveNetEnergyL2": 0.01,
            "reactiveNetEnergyL3": 0.01,
            "apparentImpEnergyL1": 0.01,
            "apparentImpEnergyL2": 0.01,
            "apparentImpEnergyL3": 0.01,
            "apparentExpEnergyL1": 0.01,
            "apparentExpEnergyL2": 0.01,
            "apparentExpEnergyL3": 0.01,
            "apparentNetEnergyL1": 0.01,
            "apparentNetEnergyL2": 0.01,
            "apparentNetEnergyL3": 0.01,
            # "threePhaseSystemVoltage": 0.1,
            # "phaseVoltageL1": 0.1,
            # "phaseVoltageL2": 0.1,
            # "phaseVoltageL3": 0.1,
            # "lineVoltageL1L2": 0.1,
            # "lineVoltageL3L2": 0.1,
            # "lineVoltageL1L3": 0.1,
            # "threePhaseSystemCurrent": 0.01,
            "currentL1": 0.01,
            "currentL2": 0.01,
            "currentL3": 0.01,
            "currentN": 0.01,
            "activePowerTotal": 0.01,
            "activePowerL1": 0.01,
            "activePowerL2": 0.01,
            "activePowerL3": 0.01,
            "reactivePowerTotal": 0.01,
            "reactivePowerL1": 0.01,
            "reactivePowerL2": 0.01,
            "reactivePowerL3": 0.01,
            "apparentPowerTotal": 0.01,
            "apparentPowerL1": 0.01,
            "apparentPowerL2": 0.01,
            "apparentPowerL3": 0.01
            # "frequency": 0.01,
            # "phaseAnglePowerTotal": 0.1,
            # "phaseAnglePowerL1": 0.1,
            # "phaseAnglePowerL2": 0.1,
            # "phaseAnglePowerL3": 0.1,
            # "phaseAngleVoltageL1": 0.1,
            # "phaseAngleVoltageL2": 0.1,
            # "phaseAngleVoltageL3": 0.1,
            # "phaseAngleCurrentL1": 0.1,
            # "phaseAngleCurrentL2": 0.1,
            # "phaseAngleCurrentL3": 0.1,
            # "powerFactorTotal": 0.001,
            # "powerFactorL1": 0.001,
            # "powerFactorL2": 0.001,
            # "powerFactorL3": 0.001,
            # "currentQuadrantTotal": 1,
            # "currentQuadrantL1": 1,
            # "currentQuadrantL2": 1,
            # "currentQuadrantL3": 1,
            # "cosphiTotal": 0.001,
            # "cosphiL1": 0.001,
            # "cosphiL2": 0.001,
            # "cosphiL3": 0.001,
            # "avgCurrentL1": 0.01,
            # "avgCurrentL2": 0.01,
            # "avgCurrentL3": 0.01,
            # "avgCurrentNeutral": 0.01,
            # "avgVoltageL1": 0.1,
            # "avgVoltageL2": 0.1,
            # "avgVoltageL3": 0.1,
            # "avgVoltageL1L2": 0.1,
            # "avgVoltageL2L3": 0.1,
            # "avgVoltageL1L3": 0.1,
            # "avgActivePowerTotal": 0.01,
            # "avgActivePowerL1": 0.01,
            # "avgActivePowerL2": 0.01,
            # "avgActivePowerL3": 0.01,
            # "avgReactivePowerTotal": 0.01,
            # "avgReactivePowerL1": 0.01,
            # "avgReactivePowerL2": 0.01,
            # "avgReactivePowerL3": 0.01,
            # "avgApparentPowerTotal": 0.01,
            # "avgApparentPowerL1": 0.01,
            # "avgApparentPowerL2": 0.01,
            # "avgApparentPowerL3": 0.01,
            # "maxCurrentL1": 0.01,
            # "maxCurrentL2": 0.01,
            # "maxCurrentL3": 0.01,
            # "maxCurrentNeutral": 0.01,
            # "maxVoltageL1": 0.1,
            # "maxVoltageL2": 0.1,
            # "maxVoltageL3": 0.1,
            # "maxVoltageL1L2": 0.1,
            # "maxVoltageL2L3": 0.1,
            # "maxVoltageL1L3": 0.1,
            # "maxActivePowerTotal": 0.01,
            # "maxActivePowerL1": 0.01,
            # "maxActivePowerL2": 0.01,
            # "maxActivePowerL3": 0.01,
            # "maxReactivePowerTotal": 0.01,
            # "maxReactivePowerL1": 0.01,
            # "maxReactivePowerL2": 0.01,
            # "maxReactivePowerL3": 0.01,
            # "maxApparentPowerTotal": 0.01,
            # "maxApparentPowerL1": 0.01,
            # "maxApparentPowerL2": 0.01,
            # "maxApparentPowerL3": 0.01,
            # "minCurrentL1": 0.01,
            # "minCurrentL2": 0.01,
            # "minCurrentL3": 0.01,
            # "minCurrentNeutral": 0.01,
            # "minVoltageL1": 0.1,
            # "minVoltageL2": 0.1,
            # "minVoltageL3": 0.1,
            # "minVoltageL1L2": 0.1,
            # "minVoltageL2L3": 0.1,
            # "minVoltageL1L3": 0.1,
            # "minActivePowerTotal": 0.01,
            # "minActivePowerL1": 0.01,
            # "minActivePowerL2": 0.01,
            # "minActivePowerL3": 0.01,
            # "minReactivePowerTotal": 0.01,
            # "minReactivePowerL1": 0.01,
            # "minReactivePowerL2": 0.01,
            # "minReactivePowerL3": 0.01,
            # "minApparentPowerTotal": 0.01,
            # "minApparentPowerL1": 0.01,
            # "minApparentPowerL2": 0.01,
            # "minApparentPowerL3": 0.01,
            # "CO2conversionFactor": 0.001,
            # "currencyConversionFactor": 0.001,
            # "activePowerTotalScaler": 1,
            # "activePowerL1Scaler": 1,
            # "activePowerL2Scaler": 1,
            # "activePowerL3Scaler": 1,
            # "reactivePowerTotalScaler": 1,
            # "reactivePowerL1Scaler": 1,
            # "reactivePowerL2Scaler": 1,
            # "reactivePowerL3Scaler": 1,
            # "apparentPowerTotalScaler": 1,
            # "apparentPowerL1Scaler": 1,
            # "apparentPowerL2Scaler": 1,
            # "apparentPowerL3Scaler": 1,
            # "avgActivePowerTotalScaler": 1,
            # "avgActivePowerL1Scaler": 1,
            # "avgActivePowerL2Scaler": 1,
            # "avgActivePowerL3Scaler": 1,
            # "avgReactivePowerTotalScaler": 1,
            # "avgReactivePowerL1Scaler": 1,
            # "avgReactivePowerL2Scaler": 1,
            # "avgReactivePowerL3Scaler": 1,
            # "avgApparentPowerTotalScaler": 1,
            # "avgApparentPowerL1Scaler": 1,
            # "avgApparentPowerL2Scaler": 1,
            # "avgApparentPowerL3Scaler": 1,
            # "maxActivePowerTotalScaler": 1,
            # "maxActivePowerL1Scaler": 1,
            # "maxActivePowerL2Scaler": 1,
            # "maxActivePowerL3Scaler": 1,
            # "maxReactivePowerTotalScaler": 1,
            # "maxReactivePowerL1Scaler": 1,
            # "maxReactivePowerL2Scaler": 1,
            # "maxReactivePowerL3Scaler": 1,
            # "maxApparentPowerTotalScaler": 1,
            # "maxApparentPowerL1Scaler": 1,
            # "maxApparentPowerL2Scaler": 1,
            # "maxApparentPowerL3Scaler": 1,
            # "minActivePowerTotalScaler": 1,
            # "minActivePowerL1Scaler": 1,
            # "minActivePowerL2Scaler": 1,
            # "minActivePowerL3Scaler": 1,
            # "minReactivePowerTotalScaler": 1,
            # "minReactivePowerL1Scaler": 1,
            # "minReactivePowerL2Scaler": 1,
            # "minReactivePowerL3Scaler": 1,
            # "minApparentPowerTotalScaler": 1,
            # "minApparentPowerL1Scaler": 1,
            # "minApparentPowerL2Scaler": 1,
            # "minApparentPowerL3Scaler": 1,
        },
        
        "decoding_map": {
            "activeImportEnergyTotal": "decode_64bit_unsigned",
            "activeExportEnergyTotal": "decode_64bit_unsigned",
            "activeNetEnergyTotal": "decode_64bit_signed",
            "reactiveImportEnergyTotal": "decode_64bit_unsigned",
            "reactiveExportEnergyTotal": "decode_64bit_unsigned",
            "reactiveNetEnergyTotal": "decode_64bit_signed",
            "apparentImportEnergyTotal": "decode_64bit_unsigned",
            "apparentExportEnergyTotal": "decode_64bit_unsigned",
            "apparentNetEnergyTotal": "decode_64bit_signed",
            # "CO2conversionFactor": "decode_64bit_unsigned",
            # "currencyConversionFactor": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff1": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff2": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff3": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff4": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff5": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff6": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff1": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff2": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff3": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff4": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff5": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff6": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff1": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff2": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff3": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff4": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff5": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff6": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff1": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff2": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff3": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff4": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff5": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff6": "decode_64bit_unsigned",
            "activeImpEnergyL1": "decode_64bit_unsigned",
            "activeImpEnergyL2": "decode_64bit_unsigned",
            "activeImpEnergyL3": "decode_64bit_unsigned",
            "activeExpEnergyL1": "decode_64bit_unsigned",
            "activeExpEnergyL2": "decode_64bit_unsigned",
            "activeExpEnergyL3": "decode_64bit_unsigned",
            "activeNetEnergyL1": "decode_64bit_signed",
            "activeNetEnergyL2": "decode_64bit_signed",
            "activeNetEnergyL3": "decode_64bit_signed",
            "reactiveImpEnergyL1": "decode_64bit_unsigned",
            "reactiveImpEnergyL2": "decode_64bit_unsigned",
            "reactiveImpEnergyL3": "decode_64bit_unsigned",
            "reactiveExpEnergyL1": "decode_64bit_unsigned",
            "reactiveExpEnergyL2": "decode_64bit_unsigned",
            "reactiveExpEnergyL3": "decode_64bit_unsigned",
            "reactiveNetEnergyL1": "decode_64bit_signed",
            "reactiveNetEnergyL2": "decode_64bit_signed",
            "reactiveNetEnergyL3": "decode_64bit_signed",
            "apparentImpEnergyL1": "decode_64bit_unsigned",
            "apparentImpEnergyL2": "decode_64bit_unsigned",
            "apparentImpEnergyL3": "decode_64bit_unsigned",
            "apparentExpEnergyL1": "decode_64bit_unsigned",
            "apparentExpEnergyL2": "decode_64bit_unsigned",
            "apparentExpEnergyL3": "decode_64bit_unsigned",
            "apparentNetEnergyL1": "decode_64bit_signed",
            "apparentNetEnergyL2": "decode_64bit_signed",
            "apparentNetEnergyL3": "decode_64bit_signed",
            # "threePhaseSystemVoltage": "decode_32bit_unsigned",
            # "phaseVoltageL1": "decode_32bit_unsigned",
            # "phaseVoltageL2": "decode_32bit_unsigned",
            # "phaseVoltageL3": "decode_32bit_unsigned",
            # "lineVoltageL1L2": "decode_32bit_unsigned",
            # "lineVoltageL3L2": "decode_32bit_unsigned",
            # "lineVoltageL1L3": "decode_32bit_unsigned",
            # "threePhaseSystemCurrent": "decode_32bit_unsigned",
            "currentL1": "decode_32bit_unsigned",
            "currentL2": "decode_32bit_unsigned",
            "currentL3": "decode_32bit_unsigned",
            "currentN": "decode_32bit_unsigned",
            "activePowerTotal": "decode_32bit_signed",
            "activePowerL1": "decode_32bit_signed",
            "activePowerL2": "decode_32bit_signed",
            "activePowerL3": "decode_32bit_signed",
            "reactivePowerTotal": "decode_32bit_signed",
            "reactivePowerL1": "decode_32bit_signed",
            "reactivePowerL2": "decode_32bit_signed",
            "reactivePowerL3": "decode_32bit_signed",
            "apparentPowerTotal": "decode_32bit_signed",
            "apparentPowerL1": "decode_32bit_signed",
            "apparentPowerL2": "decode_32bit_signed",
            "apparentPowerL3": "decode_32bit_signed"
            # "frequency": "decode_16bit_unsigned",
            # "phaseAnglePowerTotal": "decode_16bit_signed",
            # "phaseAnglePowerL1": "decode_16bit_signed",
            # "phaseAnglePowerL2": "decode_16bit_signed",
            # "phaseAnglePowerL3": "decode_16bit_signed",
            # "phaseAngleVoltageL1": "decode_16bit_signed",
            # "phaseAngleVoltageL2": "decode_16bit_signed",
            # "phaseAngleVoltageL3": "decode_16bit_signed",
            # "phaseAngleCurrentL1": "decode_16bit_signed",
            # "phaseAngleCurrentL2": "decode_16bit_signed",
            # "phaseAngleCurrentL3": "decode_16bit_signed",
            # "powerFactorTotal": "decode_16bit_signed",
            # "powerFactorL1": "decode_16bit_signed",
            # "powerFactorL2": "decode_16bit_signed",
            # "powerFactorL3": "decode_16bit_signed",
            # "currentQuadrantTotal": "decode_16bit_unsigned",
            # "currentQuadrantL1": "decode_16bit_unsigned",
            # "currentQuadrantL2": "decode_16bit_unsigned",
            # "currentQuadrantL3": "decode_16bit_unsigned",
            # "cosphiTotal": "decode_16bit_signed",
            # "cosphiL1": "decode_16bit_signed",
            # "cosphiL2": "decode_16bit_signed",
            # "cosphiL3": "decode_16bit_signed",
            # "avgCurrentL1": "decode_32bit_unsigned",
            # "avgCurrentL2": "decode_32bit_unsigned",
            # "avgCurrentL3": "decode_32bit_unsigned",
            # "avgCurrentNeutral": "decode_32bit_unsigned",
            # "avgVoltageL1": "decode_32bit_unsigned",
            # "avgVoltageL2": "decode_32bit_unsigned",
            # "avgVoltageL3": "decode_32bit_unsigned",
            # "avgVoltageL1L2": "decode_32bit_unsigned",
            # "avgVoltageL2L3": "decode_32bit_unsigned",
            # "avgVoltageL1L3": "decode_32bit_unsigned",
            # "avgActivePowerTotal": "decode_32bit_signed",
            # "avgActivePowerL1": "decode_32bit_signed",
            # "avgActivePowerL2": "decode_32bit_signed",
            # "avgActivePowerL3": "decode_32bit_signed",
            # "avgReactivePowerTotal": "decode_32bit_signed",
            # "avgReactivePowerL1": "decode_32bit_signed",
            # "avgReactivePowerL2": "decode_32bit_signed",
            # "avgReactivePowerL3": "decode_32bit_signed",
            # "avgApparentPowerTotal": "decode_32bit_signed",
            # "avgApparentPowerL1": "decode_32bit_signed",
            # "avgApparentPowerL2": "decode_32bit_signed",
            # "avgApparentPowerL3": "decode_32bit_signed",
            # "maxCurrentL1": "decode_32bit_unsigned",
            # "maxCurrentL2": "decode_32bit_unsigned",
            # "maxCurrentL3": "decode_32bit_unsigned",
            # "maxCurrentNeutral": "decode_32bit_unsigned",
            # "maxVoltageL1": "decode_32bit_unsigned",
            # "maxVoltageL2": "decode_32bit_unsigned",
            # "maxVoltageL3": "decode_32bit_unsigned",
            # "maxVoltageL1L2": "decode_32bit_unsigned",
            # "maxVoltageL2L3": "decode_32bit_unsigned",
            # "maxVoltageL1L3": "decode_32bit_unsigned",
            # "maxActivePowerTotal": "decode_32bit_signed",
            # "maxActivePowerL1": "decode_32bit_signed",
            # "maxActivePowerL2": "decode_32bit_signed",
            # "maxActivePowerL3": "decode_32bit_signed",
            # "maxReactivePowerTotal": "decode_32bit_signed",
            # "maxReactivePowerL1": "decode_32bit_signed",
            # "maxReactivePowerL2": "decode_32bit_signed",
            # "maxReactivePowerL3": "decode_32bit_signed",
            # "maxApparentPowerTotal": "decode_32bit_signed",
            # "maxApparentPowerL1": "decode_32bit_signed",
            # "maxApparentPowerL2": "decode_32bit_signed",
            # "maxApparentPowerL3": "decode_32bit_signed",
            # "minCurrentL1": "decode_32bit_unsigned",
            # "minCurrentL2": "decode_32bit_unsigned",
            # "minCurrentL3": "decode_32bit_unsigned",
            # "minCurrentNeutral": "decode_32bit_unsigned",
            # "minVoltageL1": "decode_32bit_unsigned",
            # "minVoltageL2": "decode_32bit_unsigned",
            # "minVoltageL3": "decode_32bit_unsigned",
            # "minVoltageL1L2": "decode_32bit_unsigned",
            # "minVoltageL2L3": "decode_32bit_unsigned",
            # "minVoltageL1L3": "decode_32bit_unsigned",
            # "minActivePowerTotal": "decode_32bit_signed",
            # "minActivePowerL1": "decode_32bit_signed",
            # "minActivePowerL2": "decode_32bit_signed",
            # "minActivePowerL3": "decode_32bit_signed",
            # "minReactivePowerTotal": "decode_32bit_signed",
            # "minReactivePowerL1": "decode_32bit_signed",
            # "minReactivePowerL2": "decode_32bit_signed",
            # "minReactivePowerL3": "decode_32bit_signed",
            # "minApparentPowerTotal": "decode_32bit_signed",
            # "minApparentPowerL1": "decode_32bit_signed",
            # "minApparentPowerL2": "decode_32bit_signed",
            # "minApparentPowerL3": "decode_32bit_signed",
            # "serialNumber": "decode_serial_number",
            # "firmwareVersion": "decode_version",
            # "activePowerTotalScaler": "decode_32bit_signed",
            # "activePowerL1Scaler": "decode_32bit_signed",
            # "activePowerL2Scaler": "decode_32bit_signed",
            # "activePowerL3Scaler": "decode_32bit_signed",
            # "reactivePowerTotalScaler": "decode_32bit_signed",
            # "reactivePowerL1Scaler": "decode_32bit_signed",
            # "reactivePowerL2Scaler": "decode_32bit_signed",
            # "reactivePowerL3Scaler": "decode_32bit_signed",
            # "apparentPowerTotalScaler": "decode_32bit_signed",
            # "apparentPowerL1Scaler": "decode_32bit_signed",
            # "apparentPowerL2Scaler": "decode_32bit_signed",
            # "apparentPowerL3Scaler": "decode_32bit_signed",
            # "avgActivePowerTotalScaler": "decode_32bit_signed",
            # "avgActivePowerL1Scaler": "decode_32bit_signed",
            # "avgActivePowerL2Scaler": "decode_32bit_signed",
            # "avgActivePowerL3Scaler": "decode_32bit_signed",
            # "avgReactivePowerTotalScaler": "decode_32bit_signed",
            # "avgReactivePowerL1Scaler": "decode_32bit_signed",
            # "avgReactivePowerL2Scaler": "decode_32bit_signed",
            # "avgReactivePowerL3Scaler": "decode_32bit_signed",
            # "avgApparentPowerTotalScaler": "decode_32bit_signed",
            # "avgApparentPowerL1Scaler": "decode_32bit_signed",
            # "avgApparentPowerL2Scaler": "decode_32bit_signed",
            # "avgApparentPowerL3Scaler": "decode_32bit_signed",
            # "maxActivePowerTotalScaler": "decode_32bit_signed",
            # "maxActivePowerL1Scaler": "decode_32bit_signed",
            # "maxActivePowerL2Scaler": "decode_32bit_signed",
            # "maxActivePowerL3Scaler": "decode_32bit_signed",
            # "maxReactivePowerTotalScaler": "decode_32bit_signed",
            # "maxReactivePowerL1Scaler": "decode_32bit_signed",
            # "maxReactivePowerL2Scaler": "decode_32bit_signed",
            # "maxReactivePowerL3Scaler": "decode_32bit_signed",
            # "maxApparentPowerTotalScaler": "decode_32bit_signed",
            # "maxApparentPowerL1Scaler": "decode_32bit_signed",
            # "maxApparentPowerL2Scaler": "decode_32bit_signed",
            # "maxApparentPowerL3Scaler": "decode_32bit_signed",
            # "minActivePowerTotalScaler": "decode_32bit_signed",
            # "minActivePowerL1Scaler": "decode_32bit_signed",
            # "minActivePowerL2Scaler": "decode_32bit_signed",
            # "minActivePowerL3Scaler": "decode_32bit_signed",
            # "minReactivePowerTotalScaler": "decode_32bit_signed",
            # "minReactivePowerL1Scaler": "decode_32bit_signed",
            # "minReactivePowerL2Scaler": "decode_32bit_signed",
            # "minReactivePowerL3Scaler": "decode_32bit_signed",
            # "minApparentPowerTotalScaler": "decode_32bit_signed",
            # "minApparentPowerL1Scaler": "decode_32bit_signed",
            # "minApparentPowerL2Scaler": "decode_32bit_signed",
            # "minApparentPowerL3Scaler": "decode_32bit_signed",
        },
        
        "plot_variables": [
            "activeImportEnergyTotal",
            "activeExportEnergyTotal",
            "activeNetEnergyTotal",
            "reactiveImportEnergyTotal",
            "reactiveExportEnergyTotal",
            "reactiveNetEnergyTotal",
            "apparentImportEnergyTotal",
            "apparentExportEnergyTotal",
            "apparentNetEnergyTotal",
            "activeImpEnergyL1",
            "activeImpEnergyL2",
            "activeImpEnergyL3",
            "activeExpEnergyL1",
            "activeExpEnergyL2",
            "activeExpEnergyL3",
            "activeNetEnergyL1",
            "activeNetEnergyL2",
            "activeNetEnergyL3",
            "reactiveImpEnergyL1",
            "reactiveImpEnergyL2",
            "reactiveImpEnergyL3",
            "reactiveExpEnergyL1",
            "reactiveExpEnergyL2",
            "reactiveExpEnergyL3",
            "reactiveNetEnergyL1",
            "reactiveNetEnergyL2",
            "reactiveNetEnergyL3",
            "apparentImpEnergyL1",
            "apparentImpEnergyL2",
            "apparentImpEnergyL3",
            "apparentExpEnergyL1",
            "apparentExpEnergyL2",
            "apparentExpEnergyL3",
            "apparentNetEnergyL1",
            "apparentNetEnergyL2",
            "apparentNetEnergyL3",
            "currentL1",
            "currentL2",
            "currentL3",
            "currentN",
            "activePowerTotal",
            "activePowerL1",
            "activePowerL2",
            "activePowerL3",
            "reactivePowerTotal",
            "reactivePowerL1",
            "reactivePowerL2",
            "reactivePowerL3",
            "apparentPowerTotal",
            "apparentPowerL1",
            "apparentPowerL2",
            "apparentPowerL3"
        ],
        
        "variable_units": {
            "activeImportEnergyTotal": "(kWh)",
            "activeExportEnergyTotal": "(kWh)",
            "activeNetEnergyTotal": "(kWh)",
            "reactiveImportEnergyTotal": "(kVarh)",
            "reactiveExportEnergyTotal": "(kVarh)",
            "reactiveNetEnergyTotal": "(kVarh)",
            "apparentImportEnergyTotal": "(kVAh)",
            "apparentExportEnergyTotal": "(kVAh)",
            "apparentNetEnergyTotal": "(kVAh)",
            "activeImpEnergyL1": "(kWh)",
            "activeImpEnergyL2": "(kWh)",
            "activeImpEnergyL3": "(kWh)",
            "activeExpEnergyL1": "(kWh)",
            "activeExpEnergyL2": "(kWh)",
            "activeExpEnergyL3": "(kWh)",
            "activeNetEnergyL1": "(kWh)",
            "activeNetEnergyL2": "(kWh)",
            "activeNetEnergyL3": "(kWh)",
            "reactiveImpEnergyL1": "(kVarh)",
            "reactiveImpEnergyL2": "(kVarh)",
            "reactiveImpEnergyL3": "(kVarh)",
            "reactiveExpEnergyL1": "(kVarh)",
            "reactiveExpEnergyL2": "(kVarh)",
            "reactiveExpEnergyL3": "(kVarh)",
            "reactiveNetEnergyL1": "(kVarh)",
            "reactiveNetEnergyL2": "(kVarh)",
            "reactiveNetEnergyL3": "(kVarh)",
            "apparentImpEnergyL1": "(kVAh)",
            "apparentImpEnergyL2": "(kVAh)",
            "apparentImpEnergyL3": "(kVAh)",
            "apparentExpEnergyL1": "(kVAh)",
            "apparentExpEnergyL2": "(kVAh)",
            "apparentExpEnergyL3": "(kVAh)",
            "apparentNetEnergyL1": "(kVAh)",
            "apparentNetEnergyL2": "(kVAh)",
            "apparentNetEnergyL3": "(kVAh)",
            "currentL1": "(A)",
            "currentL2": "(A)",
            "currentL3": "(A)",
            "currentN": "(A)",
            "activePowerTotal": "(W)",
            "activePowerL1": "(W)",
            "activePowerL2": "(W)",
            "activePowerL3": "(W)",
            "reactivePowerTotal": "(var)",
            "reactivePowerL1": "(var)",
            "reactivePowerL2": "(var)",
            "reactivePowerL3": "(var)",
            "apparentPowerTotal": "(VA)",
            "apparentPowerL1": "(VA)",
            "apparentPowerL2": "(VA)",
            "apparentPowerL3": "(VA)"
        }
    }
}

SLAVE_DEVICE_ASSIGNMENTS = [
    {"ips": SENSOR_IPS, "device_type_key": "M4M"}
]

if SLAVE_DEVICE_ASSIGNMENTS:
    first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
    if first_assigned_device_type_key in DEVICE_TYPE:
        VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
    else:
        VARIABLES = []
else:
    VARIABLES = []




# ---------------------- DECODING FUNCTIONS ----------------------

# Remove non-printable characters from string or list
def clean_value(value):
    if isinstance(value, str):
        cleaned = re.sub(r'[^\x20-\x7E]', '', value)
        return cleaned
    elif isinstance(value, list):
        return [clean_value(v) for v in value]
    else:
        return value

# Decode software/hardware version from registers
def decode_version(registers):
    if not registers:
        return None
    return '.'.join(str((reg >> 12) & 0xF) for reg in registers)

# Decode serial number from registers (assuming 64 bits, 4 registers)
def decode_serial_number(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = b''.join(pack('>H', r) for r in registers)
        serial_num = unpack('>Q', b)[0]
        return str(serial_num)
    except Exception as e:
        print(f"Error decoding serial number: {e}")
        return None

# Decode string from registers (2 bytes per register)
def decode_string(registers):
    try:
        if not registers:
            return None
        byte_array = b''.join(pack('>H', r) for r in registers) # Big-endian
        return byte_array.decode('ascii').strip('\0')
    except Exception as e:
        print(f"Error decoding string: {e}")
        return None

# Decode 32-bit signed integer (from 2 registers, big-endian)
def decode_32bit_signed(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>i', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit signed: {e}")
        return None

# Decode 32-bit unsigned integer (from 2 registers, big-endian)
def decode_32bit_unsigned(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>I', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit unsigned: {e}")
        return None

# Decode 64-bit signed integer (from 4 registers, big-endian)
def decode_64bit_signed(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit signed: {e}")
        return None

# Decode 64-bit unsigned integer (from 4 registers, big-endian)
def decode_64bit_unsigned(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>Q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit unsigned: {e}")
        return None

# Decode 16-bit signed integer
def decode_16bit_signed(register):
    try:
        return unpack('>h', pack('>H', register))[0]
    except Exception as e:
        print(f"Error decoding 16-bit signed: {e}")
        return None

# Decode 16-bit unsigned integer
def decode_16bit_unsigned(register):
    try:
        return register
    except Exception as e:
        print(f"Error decoding 16-bit unsigned: {e}")
        return None

# Convert device status to 8-bit binary string (the meaning of each byte is on the Modbus Map)
def decode_device_status(value):
    try:
        return format(int(value), '08b')
    except (TypeError, ValueError) as e:
        print(f"Error decoding device status (value: {value}): {e}")
        return str(value)

# Decode a specific bit from a single register value
def decode_bit(register_value, bit_pos):
    try:
        return (register_value >> bit_pos) & 1
    except Exception as e:
        print(f"Error decoding bit: {e}")
        return None

# Decode a range of bits from a single register value
def decode_bits(register_value, bit_range):
    try:
        start_bit = min(bit_range)
        end_bit = max(bit_range)
        mask = ((1 << (end_bit - start_bit + 1)) - 1) << start_bit
        return (register_value & mask) >> start_bit
    except Exception as e:
        print(f"Error decoding bits: {e}")
        return None

# Decode 32-bit little endian integer from two 16-bit registers
def decode_32bit_little_endian(regs):
    if len(regs) < 2:
        return None
    return (regs[1] << 16) | regs[0]

# Decode 32-bit IEEE 754 float from two 16-bit registers (little-endian order of registers)
def decode_32bit_float(regs):
    try:
        if not regs or len(regs) < 2:
            return None
        b = pack('<HH', regs[0], regs[1])      # Use '<HH' for little-endian word packing
        return unpack('<f', b)[0]              # Use '<f' for little-endian float unpack
    except Exception as e:
        print(f"Error decoding 32-bit float: {e}")
        return None

# Function to read coils
def read_coil(resp):
    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
        return None
    return resp.bits[0]

# Join list into a string
def flatten_value(value):
    if isinstance(value, list):
        return ','.join(str(v) for v in value)
    return value

# Map of decoding function names to actual functions
DECODING_FUNCTIONS_MAP = {
    "decode_version": decode_version,
    "decode_serial_number": decode_serial_number,
    "decode_string": decode_string,
    "decode_32bit_signed": decode_32bit_signed,
    "decode_32bit_unsigned": decode_32bit_unsigned,
    "decode_64bit_signed": decode_64bit_signed,
    "decode_64bit_unsigned": decode_64bit_unsigned,
    "decode_16bit_signed": decode_16bit_signed,
    "decode_16bit_unsigned": decode_16bit_unsigned,
    "decode_device_status": decode_device_status,
    "decode_bit": decode_bit,
    "decode_bits": decode_bits,
    "decode_32bit_little_endian": decode_32bit_little_endian,
    "decode_32bit_float": decode_32bit_float,
    "read_coil": read_coil,
    "flatten_value": flatten_value
}

# Helper function to get device configuration for a given slave IP
def get_device_config_for_ip(ip_address):
    for assignment in SLAVE_DEVICE_ASSIGNMENTS:
        if "ips" in assignment:
            if ip_address in assignment["ips"]:
                device_type_key = assignment["device_type_key"]
                if device_type_key in DEVICE_TYPE:
                    return DEVICE_TYPE[device_type_key]
                else:
                    print(f"Error: Device type key '{device_type_key}' not found in DEVICE_TYPE.")
                    return None
    print(f"Warning: IP {ip_address} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping.")
    return None




# ---------------------- SENSOR POLLING FUNCTION ----------------------

# Read all defined Modbus variables from a specific IP
def read_sensor_with_client(ip_address, thread_id, client):
    device_config = get_device_config_for_ip(ip_address)
    if device_config is None:
        dummy_len = len(DEVICE_TYPE.get("M4M", {}).get("variables", []))
        return [None] * dummy_len

    variables = device_config["variables"]
    scale_factors = device_config["scale_factors"]
    decoding_map = device_config["decoding_map"]

    values = []
    thread_prefix = f"[Thread {thread_id}] "

    for var_info in variables:
        addr, size, name, func_code = var_info[:4]
        options = var_info[4] if len(var_info) > 4 else {}

        current_value = None
        try:
            resp = None            
            if func_code == 1:  # Coils
                resp = client.read_coils(address=addr, count=size)
                if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                    print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Coils read error or empty bits.")
                    values.append(None)
                    continue
                regs_or_bits = resp.bits
            elif func_code == 3:  # Holding Registers
                resp = client.read_holding_registers(address=addr, count=size)
                if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                    print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Registers read error or empty registers.")
                    values.append(None)
                    continue
                regs_or_bits = resp.registers
            else:
                print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Unsupported function code {func_code}.")
                values.append(None)
                continue

            print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

            # Step 1: Decoding
            if name in decoding_map:
                func_name = decoding_map[name]
                decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                if decoder_func:
                    if func_name == "decode_bit":
                        bit_pos = options.get('bit')
                        if bit_pos is not None and regs_or_bits:
                            current_value = decoder_func(regs_or_bits[0], bit_pos)
                        else:
                            print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                    elif func_name == "decode_bits":
                        bit_range = options.get('bits')
                        if bit_range is not None and regs_or_bits:
                            current_value = decoder_func(regs_or_bits[0], bit_range)
                        else:
                            print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                    elif func_name == "read_coil":
                        current_value = decoder_func(resp)
                    elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                        current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                    else:
                        current_value = decoder_func(regs_or_bits)
                else:
                    print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Using raw data.")
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
            else:
                current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

            # Step 2: Scale factor
            if name in scale_factors and isinstance(current_value, (int, float)):
                current_value *= scale_factors[name]

            # Step 3: Normalize value for Excel or CSV
            final_value = flatten_value(current_value) if isinstance(current_value, list) else current_value
            values.append(final_value)

        except Exception as e:
            print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Exception: {e}")
            values.append(None)

    return values




# ---------------------- SAVE TO EXCEL ----------------------

# The data is registered on the internal memory and then written in an Excel file
def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
    wb = Workbook()
    wb.remove(wb.active)

    for ip in SENSOR_IPS:
        device_config = get_device_config_for_ip(ip)
        if device_config is None:
            print(f"Skipping sheet creation for Sensor_{ip} due to missing device configuration.")
            continue

        sheet_name = f"Sensor_{ip.replace('.', '_')}"
        ws = wb.create_sheet(title=sheet_name)
        headers = ['timestamp'] + [var[2] for var in device_config["variables"]]
        ws.append(headers)

    for poll_entry in all_poll_data:
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']
        for ip, values in data.items():
            sheet_name = f"Sensor_{ip.replace('.', '_')}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            safe_values = [clean_value(v) for v in values]
            ws.append([timestamp] + safe_values)

    wb.save(filename)

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, ip_address, chart_index):
    device_config = get_device_config_for_ip(ip_address)
    if device_config is None:
        print(f"Warning: No device configuration found for sensor {ip_address}. Cannot create chart for {var_name}.")
        return

    variable_units = device_config.get("variable_units", {})
    unit = variable_units.get(var_name, "")

    chart = ScatterChart()
    chart.title = f"{var_name} over time - IP {ip_address}"
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = f"{var_name} {unit}".strip()
    chart.legend = None

    chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    title_with_unit = f"{var_name} {unit}".strip()
    series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.smooth = True
    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'BC'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)

# Add charts to each Excel sheet using the custom create_scatter_chart function
def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
    wb = load_workbook(filename)

    for ip in SENSOR_IPS:
        sheet_name = f"Sensor_{ip.replace('.', '_')}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        num_rows = ws.max_row

        device_config = get_device_config_for_ip(ip)
        if device_config is None:
            print(f"Skipping chart generation for sensor {ip} due to missing device configuration.")
            continue

        variables_to_plot = device_config.get("plot_variables", [])
        timestamp_col = 1
        chart_index = 0

        for var in variables_to_plot:
            if var not in headers:
                continue
            var_col = headers.index(var) + 1
            create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, ip, chart_index)
            chart_index += 1

    wb.save(filename)




# ---------------------- MULTITHREADING POLLING FOR EACH CMS BUS PORT ----------------------

# Multithreaded polling of all slaves
def poll_ips_multithread(stop_event, all_poll_data):
    def poll_ip(ip, all_data, lock, thread_id):
        client = ModbusTcpClient(ip, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect to {ip}")
            return

        values = read_sensor_with_client(ip, thread_id, client)
        with lock:
            all_data[ip] = values

        client.close()

    while not stop_event.is_set():
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()

        threads = []
        for i, ip in enumerate(SENSOR_IPS):
            t = threading.Thread(target=poll_ip, args=(ip, all_data, lock, i + 1))
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        all_poll_data.append({'timestamp': timestamp, 'data': all_data})

        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        print(f"\n Completed polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

def poll_ips_multithread_once(stop_event, all_poll_data):
    if stop_event.is_set():
        return

    start_time = time.time()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    all_data = {}
    lock = threading.Lock()

    def poll_ip(ip, thread_id):
        client = ModbusTcpClient(ip, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect to {ip}")
            return
        values = read_sensor_with_client(ip, thread_id, client)
        with lock:
            all_data[ip] = values
        client.close()

    with ThreadPoolExecutor(max_workers=len(SENSOR_IPS)) as executor:
        for i, ip in enumerate(SENSOR_IPS):
            executor.submit(poll_ip, ip, i + 1)

    all_poll_data.append({'timestamp': timestamp, 'data': all_data})

    elapsed = time.time() - start_time
    sleep_time = max(0, POLL_INTERVAL - elapsed)
    print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
    time.sleep(sleep_time)

def periodic_polling_loop(stop_event, all_poll_data):
    while not stop_event.is_set():
        start_time = time.time()
        one_cycle = 600
        cycle_duration = 60 * 60

        while time.time() - start_time < one_cycle and not stop_event.is_set():
            poll_ips_multithread_once(stop_event, all_poll_data)

        if stop_event.is_set():
            break

        try:
            print(f"Saving backup with {len(all_poll_data)} readings...")
            save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_BACKUP)
        except Exception as e:
            print(f"Error on backup: {e}")

        print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
        for _ in range(int(cycle_duration)):
            if stop_event.is_set():
                break
            time.sleep(1)




# ---------------------- SPLIT EXCEL FUNCTIONS ----------------------

def split_and_generate_charts_for_ips(path_file_excel, selected_ips, name_output_folder="single_sensors_with_generated_charts", progress_bar=None, status_text=None, total_sensors=0):
    if not os.path.exists(path_file_excel):
        st.error(f"Error: File '{path_file_excel}' not found.")
        return None

    try:
        if not os.path.exists(name_output_folder):
            os.makedirs(name_output_folder)
        else:
            for file_name in os.listdir(name_output_folder):
                path_old_file = os.path.join(name_output_folder, file_name)
                if os.path.isfile(path_old_file):
                    os.remove(path_old_file)
            st.info(f"Cleaning folder '{name_output_folder}'.")

        source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
        generated_files = []

        for i, ip in enumerate(selected_ips):
            safe_ip = ip.replace('.', '_')
            sheet_name = f"Sensor_{safe_ip}"
            if status_text:
                status_text.text(f"Processing Sensor {ip}...")
            if sheet_name not in source_workbook.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue

            sheet = source_workbook[sheet_name]

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            for row in sheet.iter_rows(values_only=True):
                new_ws.append(row)

            num_rows = new_ws.max_row
            num_cols = new_ws.max_column

            if num_rows < 2 or num_cols < 2:
                st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                continue

            timestamp_col = 1

            device_config = get_device_config_for_ip(ip)
            if device_config is None:
                print(f"Skipping chart generation for Sensor {ip} in split function due to missing device configuration.")
                continue

            variables_to_plot = device_config.get("plot_variables", [])

            for idx, var_name in enumerate(variables_to_plot):
                if var_name in [cell.value for cell in new_ws[1]]:
                    var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                    create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, ip, idx)
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
            output_filepath = os.path.join(name_output_folder, output_filename)

            new_wb.save(output_filepath)
            generated_files.append(output_filepath)

            if progress_bar and total_sensors > 0:
                progress_percent = int(((i + 1) / total_sensors) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected sensors.")
            return None

        name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
        complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

        with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All files were zipped into '{complete_zip_path}'")
        return complete_zip_path

    except Exception as e:
        st.error(f"Error splitting and generating charts: {e}")
        return None

def copy_sheet_without_charts(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col_dim in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_dim] = copy(source_sheet.column_dimensions[col_dim])
    for row_dim in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row_dim] = copy(source_sheet.row_dimensions[row_dim])

def extract_ip_from_sheet_name(sheet_name):
    match = re.search(r'Sensor_(\d+_\d+_\d+_\d+)', sheet_name)
    if match:
        return match.group(1).replace('_', '.')
    return None

def merge_data_and_generate_charts_for_ips(path_modbus_log, path_backup_data, selected_ips, output_folder="merged_sensors", progress_bar=None, status_text=None, total_sensors=0):
    if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
        st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
        return None

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    else:
        for file_name in os.listdir(output_folder):
            path_to_remove = os.path.join(output_folder, file_name)
            if os.path.isfile(path_to_remove):
                os.remove(path_to_remove)
        st.info(f"Folder '{output_folder}' cleaned.")

    wb_modbus = openpyxl.load_workbook(path_modbus_log)
    wb_backup = openpyxl.load_workbook(path_backup_data)

    generated_files = []

    for i, ip in enumerate(selected_ips):
        safe_ip = ip.replace('.', '_')
        sheet_name = f"Sensor_{safe_ip}"
        if status_text:
            status_text.text(f"Processing Sensor {ip}...")
        if sheet_name not in wb_modbus.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
            continue
        if sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
            continue

        device_config = get_device_config_for_ip(ip)
        if device_config is None:
            print(f"Skipping merge/chart generation for Sensor {ip} due to missing device configuration.")
            continue

        columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

        modbus_sheet = wb_modbus[sheet_name]
        backup_sheet = wb_backup[sheet_name]

        modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
        backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

        modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
        backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

        merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
        merged_data.dropna(subset=["timestamp"], inplace=True)
        merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
        merged_data.sort_values("timestamp", inplace=True)
        merged_data.reset_index(drop=True, inplace=True)

        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        new_sheet = new_wb.create_sheet(title=sheet_name)

        headers = [cell.value for cell in modbus_sheet[1]]
        new_sheet.append(headers)

        for _, row in merged_data.iterrows():
            new_sheet.append(list(row.values))

        num_rows = new_sheet.max_row
        variables_to_plot = device_config.get("plot_variables", [])

        for chart_index, var_name in enumerate(variables_to_plot):
            if var_name in headers:
                col_idx = headers.index(var_name) + 1
                create_scatter_chart(
                    new_sheet,
                    timestamp_col=1,
                    var_col=col_idx,
                    num_rows=num_rows,
                    var_name=var_name,
                    ip_address=ip,
                    chart_index=chart_index
                )
            else:
                st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

        safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
        new_wb.save(path_output)
        generated_files.append(path_output)

        if progress_bar and total_sensors > 0:
            progress_percent = int(((i + 1) / total_sensors) * 100)
            progress_bar.progress(progress_percent)

    if not generated_files:
        st.warning("No individual files were generated for the selected sensors.")
        return None

    zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_sensor_files.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_to_add in generated_files:
            zipf.write(file_to_add, os.path.basename(file_to_add))

    st.success(f"All merged individual files have been zipped into: {zip_path}")
    return zip_path

def compare_timestamps_and_decide_action_ips(excel_file_final, excel_file_backup, selected_ips):
    if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
        st.error("Excel files (final or backup) not found for comparison.")
        return False

    wb_final = load_workbook(excel_file_final, data_only=True)
    wb_backup = load_workbook(excel_file_backup, data_only=True)

    needs_merge = False
    for ip in selected_ips:
        sheet_name = f"Sensor_{ip.replace('.', '_')}"
        if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
            continue

        ws_final = wb_final[sheet_name]
        ws_backup = wb_backup[sheet_name]

        timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
        timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

        if not timestamps_backup.issubset(timestamps_final):
            needs_merge = True
            break

    return needs_merge




# ---------------------- CREATION OF THE DASHBOARD ----------------------

# Streamlit web interface
st.title("M4M Modbus Ethernet TCP Setup")

if "polling" not in st.session_state:
    st.session_state.polling = False
    st.session_state.stop_event = threading.Event()
    st.session_state.thread = None
    st.session_state.all_poll_data = []

    st.session_state.selected_slaves_by_type = {
        device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
    }
    st.session_state.realtime_table_placeholders = {}
    st.session_state.last_data_length = 0
    st.session_state.last_ui_update_time = time.time()
    st.session_state.last_table_update_timestamp = "N/A"

st.sidebar.header("Modbus Settings")
st.sidebar.write(f"**Port:** {PORT}")
st.sidebar.write(f"**Poll Interval (seconds):** {POLL_INTERVAL}")
st.sidebar.write(f"**Monitoring Cycle:** 10 min polling / 1 hour pause")
st.sidebar.markdown(f"**Status:** {'RUNNING' if st.session_state.polling else 'STOPPED'}")

st.sidebar.subheader("Active Sensors by Type")
sensor_counts = {}
for assignment in SLAVE_DEVICE_ASSIGNMENTS:
    device_type = assignment["device_type_key"]
    if isinstance(assignment["ips"], range):
        count = len(list(assignment["ips"]))
    elif isinstance(assignment["ips"], list):
        count = len(assignment["ips"])
    sensor_counts[device_type] = sensor_counts.get(device_type, 0) + count

for device_type, count in sensor_counts.items():
    st.sidebar.write(f"- **{device_type}**: {count} active sensors")

if st.button("Start Monitoring", disabled=st.session_state.polling):
    st.session_state.all_poll_data = []
    st.session_state.stop_event.clear()

    if st.session_state.thread and st.session_state.thread.is_alive():
        st.session_state.stop_event.set()
        st.session_state.thread.join(timeout=2)
        print("Existing polling thread stopped before starting a new one.")

    st.session_state.thread = threading.Thread(
        target=periodic_polling_loop,
        args=(st.session_state.stop_event, st.session_state.all_poll_data),
        daemon=True
    )
    st.session_state.thread.start()
    st.session_state.polling = True
    st.session_state.last_data_length = 0
    st.success("Monitoring started!")

if st.button("Stop Monitoring"):
    if st.session_state.polling:
        st.session_state.stop_event.set()
        if st.session_state.thread and st.session_state.thread.is_alive():
            st.session_state.thread.join(timeout=10)
            if st.session_state.thread.is_alive():
                st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
            else:
                st.info("Polling thread terminated.")
        st.session_state.polling = False

        with st.spinner("Saving data to Excel..."):
            save_all_data_to_excel(st.session_state.all_poll_data, filename=EXCEL_FILE_FINAL)

            try:
                add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                st.success("Monitoring stopped, data and charts saved to Excel.")
            except Exception as e:
                st.error(f"Error adding charts: {e}")
    else:
        st.warning("Monitoring is not running.")

st.divider()

if st.session_state.polling:
    st.subheader("Realtime data monitoring by sensor type")

    st.session_state.realtime_ui_timestamp_placeholder = st.empty()

    assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

    for device_type_key in assigned_device_types_for_display:
        device_config = DEVICE_TYPE.get(device_type_key)
        if not device_config:
            continue

        has_slaves_assigned = any(
            assignment["device_type_key"] == device_type_key
            for assignment in SLAVE_DEVICE_ASSIGNMENTS
        )
        if not has_slaves_assigned:
            continue

        available_slaves_for_type = []
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if assignment["device_type_key"] == device_type_key:
                if isinstance(assignment["ips"], range):
                    available_slaves_for_type.extend(list(assignment["ips"]))
                elif isinstance(assignment["ips"], list):
                    available_slaves_for_type.extend(assignment["ips"])
        available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

        if not available_slaves_for_type:
            continue

        with st.expander(f"Configuration and Monitoring: **{device_type_key}**"):
            st.session_state.selected_slaves_by_type[device_type_key] = st.multiselect(
                f"Select Slave IDs for {device_type_key}:",
                options=available_slaves_for_type,
                default=st.session_state.selected_slaves_by_type.get(device_type_key, []),
                key=f"multiselect_{device_type_key}"
            )

            st.session_state.realtime_table_placeholders[device_type_key] = st.empty()

    def update_realtime_tables_by_type():
        current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

        st.session_state.realtime_ui_timestamp_placeholder.write(
            f"Last widget update at **{current_ui_update_timestamp}**"
        )

        if st.session_state.all_poll_data:
            latest_poll_entry = st.session_state.all_poll_data[-1]
            raw_data = latest_poll_entry['data']

            for device_type_key, placeholder in st.session_state.realtime_table_placeholders.items():
                selected_slaves = st.session_state.selected_slaves_by_type.get(device_type_key, [])
                device_config = DEVICE_TYPE.get(device_type_key)

                placeholder.empty()

                if not selected_slaves or not device_config:
                    placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                    continue

                table_data = []
                variable_names = [var[2] for var in device_config["variables"]]

                for slave_id in sorted(selected_slaves):
                    if slave_id in raw_data:
                        values = raw_data[slave_id]
                        row_dict = {"Slave ID": f"Slave {slave_id}"}
                        for i, var_name in enumerate(variable_names):
                            display_value = values[i] if i < len(values) and values[i] is not None else "N/A"
                            if isinstance(display_value, float):
                                display_value = f"{display_value:.2f}"
                            row_dict[var_name] = display_value
                        table_data.append(row_dict)
                    else:
                        table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                if table_data:
                    df = pd.DataFrame(table_data)
                    df = df.set_index("Slave ID")

                    placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                else:
                    placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
        else:
            for device_type_key, placeholder in st.session_state.realtime_table_placeholders.items():
                placeholder.empty()
                placeholder.info(f"Waiting for first polling data for {device_type_key}...")

    update_realtime_tables_by_type()

    if len(st.session_state.all_poll_data) > st.session_state.last_data_length:
        st.session_state.last_data_length = len(st.session_state.all_poll_data)
        st.session_state.last_ui_update_time = time.time()
        st.rerun()
    else:
        time_since_last_update = time.time() - st.session_state.last_ui_update_time
        if time_since_last_update >= POLL_INTERVAL:
            st.session_state.last_ui_update_time = time.time()
            st.rerun()
        else:
            time.sleep(max(0, POLL_INTERVAL - time_since_last_update))
            st.rerun()

# Split / Merge Excel Files by IP
if not os.path.exists(EXCEL_FILE_FINAL):
    st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
else:
    all_possible_ips = sorted(SENSOR_IPS)
    display_options = ["All"] + [f"Sensor_{ip}" for ip in all_possible_ips]

    if "All" in st.session_state.get('selected_display_options', []):
        selected_display_options = st.multiselect(
            "Choose sensors to separate:",
            display_options,
            default=["All"],
            disabled=True
        )
    else:
        selected_display_options = st.multiselect(
            "Choose sensors to separate:",
            display_options,
            default=[]
        )

    st.session_state.selected_display_options = selected_display_options

    selected_ips = []
    if "All" in selected_display_options:
        selected_ips = all_possible_ips
    elif selected_display_options:
        for ds in selected_display_options:
            match = re.search(r'(\d{1,3}_\d{1,3}_\d{1,3}_\d{1,3})', ds)
            if match:
                selected_ips.append(match.group(1).replace('_', '.'))

    if st.button("Generate Individual Files"):
        if not selected_ips:
            st.warning("Please select at least one sensor or 'All'.")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()
            total_sensors_to_process = len(selected_ips)
            zip_path = None

            if not os.path.exists(EXCEL_FILE_BACKUP):
                st.warning("Backup file not found. Performing only the split of the final Excel.")
                with st.spinner("Splitting Excel and generating charts..."):
                    zip_path = split_and_generate_charts_for_ips(
                        EXCEL_FILE_FINAL,
                        selected_ips,
                        progress_bar=progress_bar,
                        status_text=status_text,
                        total_sensors=total_sensors_to_process
                    )
            else:
                st.info("Comparing timestamps between final Excel and backup...")
                needs_merge = compare_timestamps_and_decide_action_ips(
                    EXCEL_FILE_FINAL,
                    EXCEL_FILE_BACKUP,
                    selected_ips
                )

                if needs_merge:
                    st.warning("Discrepancies found! Merging and splitting Excel for selected sensors.")
                    with st.spinner("Merging data from backup and final, and generating individual files..."):
                        zip_path = merge_data_and_generate_charts_for_ips(
                            EXCEL_FILE_FINAL,
                            EXCEL_FILE_BACKUP,
                            selected_ips,
                            progress_bar=progress_bar,
                            status_text=status_text,
                            total_sensors=total_sensors_to_process
                        )
                else:
                    st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                    with st.spinner("Splitting Excel and generating charts..."):
                        zip_path = split_and_generate_charts_for_ips(
                            EXCEL_FILE_FINAL,
                            selected_ips,
                            progress_bar=progress_bar,
                            status_text=status_text,
                            total_sensors=total_sensors_to_process
                        )

            if zip_path:
                status_text.text("Processing complete!")
                st.success("Individual files generated and available for download.")

# Example of Latest Reading (First 5 Sensors)
if os.path.exists(EXCEL_FILE_FINAL):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_FILE_FINAL, read_only=True)
    except Exception:
        wb = None

    if wb:
        st.subheader("Example of Latest Reading (First 5 Sensors)")
        displayed = 0

        for ip in sorted(SENSOR_IPS):
            if displayed >= 5:
                break
            sheet_name = f"Sensor_{ip.replace('.', '_')}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                st.markdown(f"**Sensor {ip}**")
                df = {rows[0][i]: rows[-1][i] for i in range(len(rows[0]))}
                st.write(df)
                displayed += 1
