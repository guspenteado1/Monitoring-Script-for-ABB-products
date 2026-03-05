import os
import re
import time
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from struct import pack, unpack
from copy import copy
import zipfile

import pandas as pd
import streamlit as st
from pymodbus.client import ModbusTcpClient
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl.chart.axis

# ---------------------- CONFIGURATION SECTION ----------------------

IP = '192.168.1.100'
PORT = 502
SLAVE_IDS = [3, 10, 12]
POLL_INTERVAL = 60 
EXCEL_FILE_FINAL = "modbus_poll_log.xlsx"
EXCEL_FILE_BACKUP = "backup_poll_data.xlsx"

DEVICE_TYPE = {
    "CurrentSensor": {
        "variables": [
            #(0, 3, "swVersion", 3),                # Address, Size, Name, Function Code (1 for Coils)
            #(3, 2, "hwVersion", 3),
            #(5, 4, "serialNumber", 3),
            (9, 1, "currentTrms", 3),
            (10, 1, "currentAc", 3),
            (11, 1, "currentDc", 3),
            #(22, 1, "deviceStatus", 3),
            (100, 2, "activePowerTotal", 3),
            (102, 2, "activeEnergyTotal", 3)
        ],
        
        "scale_factors": {
            "currentTrms": 0.01,
            "currentAc": 0.01,
            "currentDc": 0.01,
            "activePowerTotal": 1,
            "activeEnergyTotal": 1
        },
        
        "decoding_map": {
            #"swVersion": "decode_version",
            #"hwVersion": "decode_version",
            #"serialNumber": "decode_serial_number",
            "activePowerTotal": "decode_32bit_little_endian",
            "activeEnergyTotal": "decode_32bit_little_endian",
            #"deviceStatus": "decode_device_status",
            "currentDc": "decode_16bit_signed"
        },
        
        "plot_variables": [
            "currentTrms", 
            "currentAc", 
            "currentDc", 
            "activePowerTotal", 
            "activeEnergyTotal"],
        
        "variable_units": {
            "currentTrms": "(A)",
            "currentAc": "(A)",
            "currentDc": "(A)",
            "activePowerTotal": "(W)",
            "activeEnergyTotal": "(kWh)"
        }
    },
    
    "IOModule": {
        "variables": [
            (0, 1, "port1StateCoil", 1),                    # Address, Size, Name, Function Code (1 for Coils)
            (1, 1, "port2StateCoil", 1),
            (2, 1, "port3StateCoil", 1),
            (3, 1, "port4StateCoil", 1),
            (0, 3, "softwareVersion", 3),
            (3, 2, "hardwareVersion", 3),
            (5, 4, "serialNumber", 3),
            (18, 1, "deviceStatus", 3),
            (18, 1, "faultStatus", 3, {'bit': 5}),          # Bit 5 of deviceStatus register
            (18, 1, "inputDisabledStatus", 3, {'bit': 6}),  # Bit 6 of deviceStatus register
            (28, 1, "deviceTypeInteger", 3),
            (28, 1, "deviceType", 3, {'bits': [4, 5]}),     # Bits 4-5 of deviceTypeInteger
            (257, 1, "ioConfiguration", 3),
            (257, 1, "port1Config", 3, {'bits': [0, 1]}),
            (257, 1, "port2Config", 3, {'bits': [4, 5]}),
            (257, 1, "port3Config", 3, {'bits': [6, 7]}),
            (257, 1, "port4Config", 3, {'bits': [12, 13]}),
            (258, 2, "port1", 3),
            (260, 2, "port2", 3),
            (262, 2, "port3", 3),
            (264, 2, "port4", 3),
        ],
        
        "scale_factors": {},
        
        "decoding_map": {
            "port1StateCoil": "read_coil",
            "port2StateCoil": "read_coil",
            "port3StateCoil": "read_coil",
            "port4StateCoil": "read_coil",
            "softwareVersion": "decode_version",
            "hardwareVersion": "decode_version",
            "serialNumber": "decode_serial_number",
            "faultStatus": "decode_bit",
            "inputDisabledStatus": "decode_bit",
            "deviceType": "decode_bits",
            "port1Config": "decode_bits",
            "port2Config": "decode_bits",
            "port3Config": "decode_bits",
            "port4Config": "decode_bits",
            "port1": "decode_32bit_unsigned",        # Assuming unsigned 32-bit for FFFFFFFF
            "port2": "decode_32bit_unsigned",
            "port3": "decode_32bit_unsigned",
            "port4": "decode_32bit_unsigned",
        },
        
        "plot_variables": [],                        # Coils and configurations are not typically plotted
        
        "variable_units": {}
    },

    "INSS-HModule": {
        "variables": [
            (0, 1, "deviceType", 3),
            (1, 2, "swVersion", 3),
            (3, 2, "hwVersion", 3),
            (5, 4, "serialNumber", 3),
            (32, 1, "switchState", 3),
            (32, 1, "switch1", 3, {'bit': 1}),
            (32, 1, "switch2", 3, {'bit': 2}),
            (33, 1, "temperature1", 3),
            (34, 1, "temperature2", 3),
            (35, 1, "temperature3", 3),
            (36, 1, "temperature4", 3),
            (37, 1, "inputVoltage", 3)
        ],
        
        "scale_factors": {
            "temperature1": 0.1,
            "temperature2": 0.1,
            "temperature3": 0.1,
            "temperature4": 0.1,
            "inputVoltage": 0.1
        },
        
        "decoding_map": {
            "swVersion": "decode_version",
            "hwVersion": "decode_version",
            "serialNumber": "decode_serial_number",
            "switch1": "decode_bit",
            "switch2": "decode_bit"
        },
        
        "plot_variables": [
            "temperature1", 
            "temperature2", 
            "temperature3", 
            "temperature4", 
            "inputVoltage"],
        
        "variable_units": {
            "temperature1": "(°C)", 
            "temperature2": "(°C)", 
            "temperature3": "(°C)", 
            "temperature4": "(°C)",
            "inputVoltage": "(V)"
        }
    },
    
    "D11_D13": {
        "variables": [
            (20480, 4, "activeImportEnergyTotal", 3),
            (20484, 4, "activeExportEnergyTotal", 3),
            (20488, 4, "activeNetEnergyTotal", 3),
            (20492, 4, "reactiveImportEnergyTotal", 3),
            (20496, 4, "reactiveExportEnergyTotal", 3),
            (20500, 4, "reactiveNetEnergyTotal", 3),
            (20512, 4, "apparentEnergy", 3),
            # (20516, 4, "equivalentCO2onActiveImpEnergyTot", 3),
            # (20532, 4, "equivalentCurrencyonActiveImpEnergyTot", 3),
            # (20536, 4, "equivalentCurrencyonActiveExpEnergyTot", 3),
            # (20848, 4, "activeImpEnergyTotTariff1", 3),
            # (20852, 4, "activeImpEnergyTotTariff2", 3),
            # (20856, 4, "activeImpEnergyTotTariff3", 3),
            # (20860, 4, "activeImpEnergyTotTariff4", 3),
            # (20880, 4, "activeExpEnergyTotTariff1", 3),
            # (20884, 4, "activeExpEnergyTotTariff2", 3),
            # (20888, 4, "activeExpEnergyTotTariff3", 3),
            # (20892, 4, "activeExpEnergyTotTariff4", 3),
            # (20912, 4, "reactiveImpEnergyTotTariff1", 3),
            # (20916, 4, "reactiveImpEnergyTotTariff2", 3),
            # (20920, 4, "reactiveImpEnergyTotTariff3", 3),
            # (20924, 4, "reactiveImpEnergyTotTariff4", 3),
            # (20944, 4, "reactiveExpEnergyTotTariff1", 3),
            # (20948, 4, "reactiveExpEnergyTotTariff2", 3),
            # (20952, 4, "reactiveExpEnergyTotTariff3", 3),
            # (20956, 4, "reactiveExpEnergyTotTariff4", 3),
            # (20976, 4, "apparentEnergyTotalTariff1", 3),
            # (20980, 4, "apparentEnergyTotalTariff2", 3),
            # (20984, 4, "apparentEnergyTotalTariff3", 3),
            # (20988, 4, "apparentEnergyTotalTariff4", 3),
            (21600, 4, "activeImpEnergyL1", 3),
            (21604, 4, "activeImpEnergyL2", 3),
            (21608, 4, "activeImpEnergyL3", 3),
            (21612, 4, "activeExpEnergyL1", 3),
            (21616, 4, "activeExpEnergyL2", 3),
            (21620, 4, "activeExpEnergyL3", 3),
            (21624, 4, "activeNetEnergyL1", 3),
            (21628, 4, "activeNetEnergyL2", 3),
            (21632, 4, "activeNetEnergyL3", 3),
            (21636, 4, "reactiveImpEnergyL1", 3),
            (21640, 4, "reactiveImpEnergyL2", 3),
            (21644, 4, "reactiveImpEnergyL3", 3),
            (21648, 4, "reactiveExpEnergyL1", 3),
            (21652, 4, "reactiveExpEnergyL2", 3),
            (21656, 4, "reactiveExpEnergyL3", 3),
            (21660, 4, "reactiveNetEnergyL1", 3),
            (21664, 4, "reactiveNetEnergyL2", 3),
            (21668, 4, "reactiveNetEnergyL3", 3),
            (21696, 4, "apparentEnergyL1", 3),
            (21700, 4, "apparentEnergyL2", 3),
            (21704, 4, "apparentEnergyL3", 3),
            (21804, 4, "partialActiveImpEnergyTot", 3),
            (21808, 4, "partialActiveExpEnergyTot", 3),
            (21812, 4, "partialReactiveImpEnergyTot", 3),
            (21816, 4, "partialReactiveExpEnergyTot", 3),
            (23296, 2, "voltageL1N", 3),
            (23298, 2, "voltageL2N", 3),
            (23300, 2, "voltageL3N", 3),
            (23302, 2, "voltageL1L2", 3),
            (23304, 2, "voltageL2L3", 3),
            (23306, 2, "voltageL3L1", 3),
            (23308, 2, "currentL1", 3),
            (23310, 2, "currentL2", 3),
            (23312, 2, "currentL3", 3),
            (23314, 2, "currentIN", 3),
            (23316, 2, "activePowerTotal", 3),
            (23318, 2, "activePowerL1", 3),
            (23320, 2, "activePowerL2", 3), 
            (23322, 2, "activePowerL3", 3), 
            (23324, 2, "reactivePowerTotal", 3),
            (23326, 2, "reactivePowerL1", 3), 
            (23328, 2, "reactivePowerL2", 3), 
            (23330, 2, "reactivePowerL3", 3), 
            (23332, 2, "apparentPowerTotal", 3),
            (23334, 2, "apparentPowerL1", 3), 
            (23336, 2, "apparentPowerL2", 3), 
            (23338, 2, "apparentPowerL3", 3)
            # (23340, 1, "frequency", 3), 
            # (23354, 1, "powerFactorTotal", 3), 
            # (23355, 1, "powerFactorL1", 3), 
            # (23356, 1, "powerFactorL2", 3), 
            # (23357, 1, "powerFactorL3", 3), 
            # (23358, 1, "quadrantTotal", 3), 
            # (23359, 1, "quadrantL1", 3), 
            # (23360, 1, "quadrantL2", 3), 
            # (23361, 1, "quadrantL3", 3), 
            # (23368, 1, "cosinePhiTotal", 3), 
            # (23369, 1, "cosinePhiL1", 3), 
            # (23370, 1, "cosinePhiL2", 3), 
            # (23371, 1, "cosinePhiL3", 3), 
            # (35072, 2, "serialNumber", 3), 
            # (35080, 8, "firmwareVersion", 3), 
            # (36064, 2, "CO2conversionFactor", 3), 
            # (36066, 2, "currencyConversionFactorforImpActEnergy", 3), 
            # (36088, 2, "currencyConversionFactorforExpActEnergy", 3)
        ],
        
        "scale_factors": {
            "activeImportEnergyTotal": 0.01,
            "activeExportEnergyTotal": 0.01,
            "activeNetEnergyTotal": 0.01,
            "reactiveImportEnergyTotal": 0.01,
            "reactiveExportEnergyTotal": 0.01,
            "reactiveNetEnergyTotal": 0.01,
            "apparentEnergy": 0.01,
            # "equivalentCO2onActiveImpEnergyTot": 1,
            # "equivalentCurrencyonActiveImpEnergyTot": 0.01,
            # "equivalentCurrencyonActiveExpEnergyTot": 0.01,
            # "activeImpEnergyTotTariff1": 0.01,
            # "activeImpEnergyTotTariff2": 0.01,
            # "activeImpEnergyTotTariff3": 0.01,
            # "activeImpEnergyTotTariff4": 0.01,
            # "activeExpEnergyTotTariff1": 0.01,
            # "activeExpEnergyTotTariff2": 0.01,
            # "activeExpEnergyTotTariff3": 0.01,
            # "activeExpEnergyTotTariff4": 0.01,
            # "reactiveImpEnergyTotTariff1": 0.01,
            # "reactiveImpEnergyTotTariff2": 0.01,
            # "reactiveImpEnergyTotTariff3": 0.01,
            # "reactiveImpEnergyTotTariff4": 0.01,
            # "reactiveExpEnergyTotTariff1": 0.01,
            # "reactiveExpEnergyTotTariff2": 0.01,
            # "reactiveExpEnergyTotTariff3": 0.01,
            # "reactiveExpEnergyTotTariff4": 0.01,
            # "apparentEnergyTotalTariff1": 0.01,
            # "apparentEnergyTotalTariff2": 0.01,
            # "apparentEnergyTotalTariff3": 0.01,
            # "apparentEnergyTotalTariff4": 0.01,
            "activeImpEnergyL1": 0.01,
            "activeImpEnergyL2": 0.01,
            "activeImpEnergyL3": 0.01,
            "activeExpEnergyL1": 0.01,
            "activeExpEnergyL2": 0.01,
            "activeExpEnergyL3": 0.01,
            "activeNetEnergyL1": 0.01,
            "activeNetEnergyL2": 0.01,
            "activeNetEnergyL3": 0.01,
            "reactiveImpEnergyL1": 0.01,
            "reactiveImpEnergyL2": 0.01,
            "reactiveImpEnergyL3": 0.01,
            "reactiveExpEnergyL1": 0.01,
            "reactiveExpEnergyL2": 0.01,
            "reactiveExpEnergyL3": 0.01,
            "reactiveNetEnergyL1": 0.01,
            "reactiveNetEnergyL2": 0.01,
            "reactiveNetEnergyL3": 0.01,
            "apparentEnergyL1": 0.01,
            "apparentEnergyL2": 0.01,
            "apparentEnergyL3": 0.01,
            "partialActiveImpEnergyTot": 0.01,
            "partialActiveExpEnergyTot": 0.01,
            "partialReactiveImpEnergyTot": 0.01,
            "partialReactiveExpEnergyTot": 0.01,
            "voltageL1N": 0.1,
            "voltageL2N": 0.1,
            "voltageL3N": 0.1,
            "voltageL1L2": 0.1,
            "voltageL2L3": 0.1,
            "voltageL3L1": 0.1,
            "currentL1": 0.01,
            "currentL2": 0.01,
            "currentL3": 0.01,
            "currentIN": 0.01,
            "activePowerTotal": 0.01,
            "activePowerL1": 0.01,
            "activePowerL2": 0.01,
            "activePowerL3": 0.01,
            "reactivePowerTotal": 0.01,
            "reactivePowerL1": 0.01,
            "reactivePowerL2": 0.01,
            "reactivePowerL3": 0.01,
            "apparentPowerTotal": 0.01,
            "apparentPowerL1": 0.01,
            "apparentPowerL2": 0.01,
            "apparentPowerL3": 0.01
            # "frequency": 0.01,
            # "powerFactorTotal": 0.001,
            # "powerFactorL1": 0.001,
            # "powerFactorL2": 0.001,
            # "powerFactorL3": 0.001,
            # "quadrantTotal": 0.001,
            # "quadrantL1": 0.001,
            # "quadrantL2": 0.001,
            # "quadrantL3": 0.001,
            # "cosinePhiTotal": 0.001,
            # "cosinePhiL1": 0.001,
            # "cosinePhiL2": 0.001,
            # "cosinePhiL3": 0.001
        },
        
        "decoding_map": {
            "activeImportEnergyTotal": "decode_64bit_unsigned",
            "activeExportEnergyTotal": "decode_64bit_unsigned",
            "activeNetEnergyTotal": "decode_64bit_signed",
            "reactiveImportEnergyTotal": "decode_64bit_unsigned",
            "reactiveExportEnergyTotal": "decode_64bit_unsigned",
            "reactiveNetEnergyTotal": "decode_64bit_signed",
            "apparentEnergy": "decode_64bit_unsigned",
            # "equivalentCO2onActiveImpEnergyTot": "decode_64bit_unsigned",
            # "equivalentCurrencyonActiveImpEnergyTot": "decode_64bit_unsigned",
            # "equivalentCurrencyonActiveExpEnergyTot": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff1": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff2": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff3": "decode_64bit_unsigned",
            # "activeImpEnergyTotTariff4": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff1": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff2": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff3": "decode_64bit_unsigned",
            # "activeExpEnergyTotTariff4": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff1": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff2": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff3": "decode_64bit_unsigned",
            # "reactiveImpEnergyTotTariff4": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff1": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff2": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff3": "decode_64bit_unsigned",
            # "reactiveExpEnergyTotTariff4": "decode_64bit_unsigned",
            # "apparentEnergyTotalTariff1": "decode_64bit_unsigned",
            # "apparentEnergyTotalTariff2": "decode_64bit_unsigned",
            # "apparentEnergyTotalTariff3": "decode_64bit_unsigned",
            # "apparentEnergyTotalTariff4": "decode_64bit_unsigned",
            "activeImpEnergyL1": "decode_64bit_unsigned",
            "activeImpEnergyL2": "decode_64bit_unsigned",
            "activeImpEnergyL3": "decode_64bit_unsigned",
            "activeExpEnergyL1": "decode_64bit_unsigned",
            "activeExpEnergyL2": "decode_64bit_unsigned",
            "activeExpEnergyL3": "decode_64bit_unsigned",
            "activeNetEnergyL1": "decode_64bit_signed",
            "activeNetEnergyL2": "decode_64bit_signed",
            "activeNetEnergyL3": "decode_64bit_signed",
            "reactiveImpEnergyL1": "decode_64bit_unsigned",
            "reactiveImpEnergyL2": "decode_64bit_unsigned",
            "reactiveImpEnergyL3": "decode_64bit_unsigned",
            "reactiveExpEnergyL1": "decode_64bit_unsigned",
            "reactiveExpEnergyL2": "decode_64bit_unsigned",
            "reactiveExpEnergyL3": "decode_64bit_unsigned",
            "reactiveNetEnergyL1": "decode_64bit_signed",
            "reactiveNetEnergyL2": "decode_64bit_signed",
            "reactiveNetEnergyL3": "decode_64bit_signed",
            "apparentEnergyL1": "decode_64bit_unsigned",
            "apparentEnergyL2": "decode_64bit_unsigned",
            "apparentEnergyL3": "decode_64bit_unsigned",
            "partialActiveImpEnergyTot": "decode_64bit_unsigned",
            "partialActiveExpEnergyTot": "decode_64bit_unsigned",
            "partialReactiveImpEnergyTot": "decode_64bit_unsigned",
            "partialReactiveExpEnergyTot": "decode_64bit_unsigned",
            "voltageL1N": "decode_32bit_unsigned", 
            "voltageL2N": "decode_32bit_unsigned", 
            "voltageL3N": "decode_32bit_unsigned",
            "voltageL1L2": "decode_32bit_unsigned", 
            "voltageL2L3": "decode_32bit_unsigned", 
            "voltageL3L1": "decode_32bit_unsigned",
            "currentL1": "decode_32bit_unsigned", 
            "currentL2": "decode_32bit_unsigned", 
            "currentL3": "decode_32bit_unsigned", 
            "currentIN": "decode_32bit_unsigned",
            "activePowerTotal": "decode_32bit_signed", 
            "activePowerL1": "decode_32bit_signed", 
            "activePowerL2": "decode_32bit_signed", 
            "activePowerL3": "decode_32bit_signed",
            "reactivePowerTotal": "decode_32bit_signed", 
            "reactivePowerL1": "decode_32bit_signed", 
            "reactivePowerL2": "decode_32bit_signed", 
            "reactivePowerL3": "decode_32bit_signed",
            "apparentPowerTotal": "decode_32bit_unsigned", 
            "apparentPowerL1": "decode_32bit_unsigned", 
            "apparentPowerL2": "decode_32bit_unsigned", 
            "apparentPowerL3": "decode_32bit_unsigned"
            # "CO2conversionFactor": "decode_32bit_unsigned",
            # "currencyConversionFactorforImpActEnergy": "decode_32bit_unsigned",
            # "currencyConversionFactorforExpActEnergy": "decode_32bit_unsigned",
            # "frequency": "decode_16bit_unsigned",
            # "powerFactorTotal": "decode_16bit_signed", 
            # "powerFactorL1": "decode_16bit_signed", 
            # "powerFactorL2": "decode_16bit_signed", 
            # "powerFactorL3": "decode_16bit_signed",
            # "cosinePhiTotal": "decode_16bit_signed", 
            # "cosinePhiL1": "decode_16bit_signed", 
            # "cosinePhiL2": "decode_16bit_signed", 
            # "cosinePhiL3": "decode_16bit_signed",
            # "quadrantTotal": "decode_16bit_signed",
            # "quadrantL1": "decode_16bit_unsigned",
            # "quadrantL2": "decode_16bit_unsigned",
            # "quadrantL3": "decode_16bit_usigned",
            # "serialNumber": "decode_serial_number",
            # "firmwareVersion": "decode_version"
        },
        
        "plot_variables": [
            "activeImportEnergyTotal",
            "activeExportEnergyTotal",
            "activeNetEnergyTotal",
            "reactiveImportEnergyTotal",
            "reactiveExportEnergyTotal",
            "reactiveNetEnergyTotal",
            "apparentEnergy",
            "activeImpEnergyL1",
            "activeImpEnergyL2",
            "activeImpEnergyL3",
            "activeExpEnergyL1",
            "activeExpEnergyL2",
            "activeExpEnergyL3",
            "activeNetEnergyL1",
            "activeNetEnergyL2",
            "activeNetEnergyL3",
            "reactiveImpEnergyL1",
            "reactiveImpEnergyL2",
            "reactiveImpEnergyL3",
            "reactiveExpEnergyL1",
            "reactiveExpEnergyL2",
            "reactiveExpEnergyL3",
            "reactiveNetEnergyL1",
            "reactiveNetEnergyL2",
            "reactiveNetEnergyL3",
            "apparentEnergyL1",
            "apparentEnergyL2",
            "apparentEnergyL3",
            "partialActiveImpEnergyTot",
            "partialActiveExpEnergyTot",
            "partialReactiveImpEnergyTot",
            "partialReactiveExpEnergyTot",
            "voltageL1N",
            "voltageL2N",
            "voltageL3N",
            "voltageL1L2",
            "voltageL2L3",
            "voltageL3L1",
            "currentL1",
            "currentL2",
            "currentL3",
            "currentIN",
            "activePowerTotal",
            "activePowerL1",
            "activePowerL2",
            "activePowerL3",
            "reactivePowerTotal",
            "reactivePowerL1",
            "reactivePowerL2",
            "reactivePowerL3",
            "apparentPowerTotal",
            "apparentPowerL1",
            "apparentPowerL2",
            "apparentPowerL3"
        ],
        
        "variable_units": {
            "activeImportEnergyTotal": "(kWh)",
            "activeExportEnergyTotal": "(kWh)",
            "activeNetEnergyTotal": "(kWh)",
            "reactiveImportEnergyTotal": "(kVarh)",
            "reactiveExportEnergyTotal": "(kVarh)",
            "reactiveNetEnergyTotal": "(kVarh)",
            "apparentEnergy": "(kVAhh)",
            "activeImpEnergyL1": "(kWh)",
            "activeImpEnergyL2": "(kWh)",
            "activeImpEnergyL3": "(kWh)",
            "activeExpEnergyL1": "(kWh)",
            "activeExpEnergyL2": "(kWh)",
            "activeExpEnergyL3": "(kWh)",
            "activeNetEnergyL1": "(kWh)",
            "activeNetEnergyL2": "(kWh)",
            "activeNetEnergyL3": "(kWh)",
            "reactiveImpEnergyL1": "(kVarh)",
            "reactiveImpEnergyL2": "(kVarh)",
            "reactiveImpEnergyL3": "(kVarh)",
            "reactiveExpEnergyL1": "(kVarh)",
            "reactiveExpEnergyL2": "(kVarh)",
            "reactiveExpEnergyL3": "(kVarh)",
            "reactiveNetEnergyL1": "(kVarh)",
            "reactiveNetEnergyL2": "(kVarh)",
            "reactiveNetEnergyL3": "(kVarh)",
            "apparentEnergyL1": "(kVAhh)",
            "apparentEnergyL2": "(kVAhh)",
            "apparentEnergyL3": "(kVAhh)",
            "partialActiveImpEnergyTot": "(kWh)",
            "partialActiveExpEnergyTot": "(kWh)",
            "partialReactiveImpEnergyTot": "(kWh)",
            "partialReactiveExpEnergyTot": "(kWh)",
            "voltageL1N": "(V)",
            "voltageL2N": "(V)",
            "voltageL3N": "(V)",
            "voltageL1L2": "(V)",
            "voltageL2L3": "(V)",
            "voltageL3L1": "(V)",
            "currentL1": "(A)",
            "currentL2": "(A)",
            "currentL3": "(A)",
            "currentIN": "(A)",
            "activePowerTotal": "(W)",
            "activePowerL1": "(W)",
            "activePowerL2": "(W)",
            "activePowerL3": "(W)",
            "reactivePowerTotal": "(Var)",
            "reactivePowerL1": "(Var)",
            "reactivePowerL2": "(Var)",
            "reactivePowerL3": "(Var)",
            "apparentPowerTotal": "(VA)",
            "apparentPowerL1": "(VA)",
            "apparentPowerL2": "(VA)",
            "apparentPowerL3": "(VA)"
        }
    },

    "M4M_30_A2": {
        "variables": [
            (40945, 4, "activeImportEnergyTotal", 3),
            (40949, 4, "activeExportEnergyTotal", 3),
            (40953, 4, "activeNetEnergyTotal", 3),
            (40957, 4, "reactiveImportEnergyTotal", 3),
            (40961, 4, "reactiveExportEnergyTotal", 3),
            (40965, 4, "reactiveNetEnergyTotal", 3),
            (40969, 4, "apparentImportEnergyTotal", 3),
            (40973, 4, "apparentExportEnergyTotal", 3),
            (40977, 4, "apparentNetEnergyTotal", 3),
            (40420, 4, "apparentNetEnergyL1", 3),
            (40424, 4, "apparentNetEnergyL2", 3),
            (40428, 4, "apparentNetEnergyL3", 3)
        ],
        
        "scale_factors": {
            "activeImportEnergyTotal": 0.01,
            "activeExportEnergyTotal": 0.01,
            "activeNetEnergyTotal": 0.01,
            "reactiveImportEnergyTotal": 0.01,
            "reactiveExportEnergyTotal": 0.01,
            "reactiveNetEnergyTotal": 0.01,
            "apparentImportEnergyTotal": 0.01,
            "apparentExportEnergyTotal": 0.01,
            "apparentNetEnergyTotal": 0.01,
            "apparentNetEnergyL1": 0.01,
            "apparentNetEnergyL2": 0.01,
            "apparentNetEnergyL3": 0.01
        },
        
        "decoding_map": {
            "activeImportEnergyTotal": "decode_64bit_unsigned",
            "activeExportEnergyTotal": "decode_64bit_unsigned",
            "activeNetEnergyTotal": "decode_64bit_signed",
            "reactiveImportEnergyTotal": "decode_64bit_unsigned",
            "reactiveExportEnergyTotal": "decode_64bit_unsigned",
            "reactiveNetEnergyTotal": "decode_64bit_signed",
            "apparentImportEnergyTotal": "decode_64bit_unsigned",
            "apparentExportEnergyTotal": "decode_64bit_unsigned",
            "apparentNetEnergyTotal": "decode_64bit_signed",
            "apparentNetEnergyL1": "decode_64bit_signed",
            "apparentNetEnergyL2": "decode_64bit_signed",
            "apparentNetEnergyL3": "decode_64bit_signed"
        },
        
        "plot_variables": [
            "activeImportEnergyTotal",
            "activeExportEnergyTotal",
            "activeNetEnergyTotal",
            "reactiveImportEnergyTotal",
            "reactiveExportEnergyTotal",
            "reactiveNetEnergyTotal",
            "apparentImportEnergyTotal",
            "apparentExportEnergyTotal",
            "apparentNetEnergyTotal",
            "apparentNetEnergyL1",
            "apparentNetEnergyL2",
            "apparentNetEnergyL3"
        ],
        
        "variable_units": {
            "activeImportEnergyTotal": "(kWh)",
            "activeExportEnergyTotal": "(kWh)",
            "activeNetEnergyTotal": "(kWh)",
            "reactiveImportEnergyTotal": "(kVarh)",
            "reactiveExportEnergyTotal": "(kVarh)",
            "reactiveNetEnergyTotal": "(kVarh)",
            "apparentImportEnergyTotal": "(kVAh)",
            "apparentExportEnergyTotal": "(kVAh)",
            "apparentNetEnergyTotal": "(kVAh)",
            "apparentNetEnergyL1": "(kVAh)",
            "apparentNetEnergyL2": "(kVAh)",
            "apparentNetEnergyL3": "(kVAh)"
        }
    },

    "M4M_30_F4": {
        "variables": [
            (49725, 4, "activeImpEnergyL1", 3),
            (49729, 4, "activeImpEnergyL2", 3),
            (49733, 4, "activeImpEnergyL3", 3),
            (49737, 4, "activeExpEnergyL1", 3),
            (49741, 4, "activeExpEnergyL2", 3),
            (49745, 4, "activeExpEnergyL3", 3),
            (49749, 4, "activeNetEnergyL1", 3),
            (49753, 4, "activeNetEnergyL2", 3),
            (49757, 4, "activeNetEnergyL3", 3),
            (49761, 4, "reactiveImpEnergyL1", 3),
            (49765, 4, "reactiveImpEnergyL2", 3),
            (49769, 4, "reactiveImpEnergyL3", 3),
            (49773, 4, "reactiveExpEnergyL1", 3),
            (49777, 4, "reactiveExpEnergyL2", 3),
            (49781, 4, "reactiveExpEnergyL3", 3),
            (49785, 4, "reactiveNetEnergyL1", 3),
            (49789, 4, "reactiveNetEnergyL2", 3),
            (49793, 4, "reactiveNetEnergyL3", 3),
            (49797, 4, "apparentImpEnergyL1", 3),
            (49801, 4, "apparentImpEnergyL2", 3),
            (49805, 4, "apparentImpEnergyL3", 3),
            (49809, 4, "apparentExpEnergyL1", 3),
            (49813, 4, "apparentExpEnergyL2", 3),
            (49817, 4, "apparentExpEnergyL3", 3)
        ],
        
        "scale_factors": {
            "activeImpEnergyL1": 0.01,
            "activeImpEnergyL2": 0.01,
            "activeImpEnergyL3": 0.01,
            "activeExpEnergyL1": 0.01,
            "activeExpEnergyL2": 0.01,
            "activeExpEnergyL3": 0.01,
            "activeNetEnergyL1": 0.01,
            "activeNetEnergyL2": 0.01,
            "activeNetEnergyL3": 0.01,
            "reactiveImpEnergyL1": 0.01,
            "reactiveImpEnergyL2": 0.01,
            "reactiveImpEnergyL3": 0.01,
            "reactiveExpEnergyL1": 0.01,
            "reactiveExpEnergyL2": 0.01,
            "reactiveExpEnergyL3": 0.01,
            "reactiveNetEnergyL1": 0.01,
            "reactiveNetEnergyL2": 0.01,
            "reactiveNetEnergyL3": 0.01,
            "apparentImpEnergyL1": 0.01,
            "apparentImpEnergyL2": 0.01,
            "apparentImpEnergyL3": 0.01,
            "apparentExpEnergyL1": 0.01,
            "apparentExpEnergyL2": 0.01,
            "apparentExpEnergyL3": 0.01
        },
        
        "decoding_map": {
            "activeImpEnergyL1": "decode_64bit_unsigned",
            "activeImpEnergyL2": "decode_64bit_unsigned",
            "activeImpEnergyL3": "decode_64bit_unsigned",
            "activeExpEnergyL1": "decode_64bit_unsigned",
            "activeExpEnergyL2": "decode_64bit_unsigned",
            "activeExpEnergyL3": "decode_64bit_unsigned",
            "activeNetEnergyL1": "decode_64bit_signed",
            "activeNetEnergyL2": "decode_64bit_signed",
            "activeNetEnergyL3": "decode_64bit_signed",
            "reactiveImpEnergyL1": "decode_64bit_unsigned",
            "reactiveImpEnergyL2": "decode_64bit_unsigned",
            "reactiveImpEnergyL3": "decode_64bit_unsigned",
            "reactiveExpEnergyL1": "decode_64bit_unsigned",
            "reactiveExpEnergyL2": "decode_64bit_unsigned",
            "reactiveExpEnergyL3": "decode_64bit_unsigned",
            "reactiveNetEnergyL1": "decode_64bit_signed",
            "reactiveNetEnergyL2": "decode_64bit_signed",
            "reactiveNetEnergyL3": "decode_64bit_signed",
            "apparentImpEnergyL1": "decode_64bit_unsigned",
            "apparentImpEnergyL2": "decode_64bit_unsigned",
            "apparentImpEnergyL3": "decode_64bit_unsigned",
            "apparentExpEnergyL1": "decode_64bit_unsigned",
            "apparentExpEnergyL2": "decode_64bit_unsigned",
            "apparentExpEnergyL3": "decode_64bit_unsigned"
        },
        
        "plot_variables": [
            "activeImpEnergyL1",
            "activeImpEnergyL2",
            "activeImpEnergyL3",
            "activeExpEnergyL1",
            "activeExpEnergyL2",
            "activeExpEnergyL3",
            "activeNetEnergyL1",
            "activeNetEnergyL2",
            "activeNetEnergyL3",
            "reactiveImpEnergyL1",
            "reactiveImpEnergyL2",
            "reactiveImpEnergyL3",
            "reactiveExpEnergyL1",
            "reactiveExpEnergyL2",
            "reactiveExpEnergyL3",
            "reactiveNetEnergyL1",
            "reactiveNetEnergyL2",
            "reactiveNetEnergyL3",
            "apparentImpEnergyL1",
            "apparentImpEnergyL2",
            "apparentImpEnergyL3",
            "apparentExpEnergyL1",
            "apparentExpEnergyL2",
            "apparentExpEnergyL3"
        ],
        
        "variable_units": {
            "activeImpEnergyL1": "(kWh)",
            "activeImpEnergyL2": "(kWh)",
            "activeImpEnergyL3": "(kWh)",
            "activeExpEnergyL1": "(kWh)",
            "activeExpEnergyL2": "(kWh)",
            "activeExpEnergyL3": "(kWh)",
            "activeNetEnergyL1": "(kWh)",
            "activeNetEnergyL2": "(kWh)",
            "activeNetEnergyL3": "(kWh)",
            "reactiveImpEnergyL1": "(kVarh)",
            "reactiveImpEnergyL2": "(kVarh)",
            "reactiveImpEnergyL3": "(kVarh)",
            "reactiveExpEnergyL1": "(kVarh)",
            "reactiveExpEnergyL2": "(kVarh)",
            "reactiveExpEnergyL3": "(kVarh)",
            "reactiveNetEnergyL1": "(kVarh)",
            "reactiveNetEnergyL2": "(kVarh)",
            "reactiveNetEnergyL3": "(kVarh)",
            "apparentImpEnergyL1": "(kVAh)",
            "apparentImpEnergyL2": "(kVAh)",
            "apparentImpEnergyL3": "(kVAh)",
            "apparentExpEnergyL1": "(kVAh)",
            "apparentExpEnergyL2": "(kVAh)",
            "apparentExpEnergyL3": "(kVAh)"
        }
    },

    "M4M_20_F5": {
        "variables": [
            (52689, 2, "currentL1", 3),
            (52691, 2, "currentL2", 3),
            (52693, 2, "currentL3", 3),
            (52695, 2, "currentN", 3),
            (52699, 2, "activePowerTotal", 3),
            (52701, 2, "activePowerL1", 3),
            (52703, 2, "activePowerL2", 3),
            (52705, 2, "activePowerL3", 3),
            (52707, 2, "reactivePowerTotal", 3),
            (52709, 2, "reactivePowerL1", 3),
            (52711, 2, "reactivePowerL2", 3),
            (52713, 2, "reactivePowerL3", 3),
            (52715, 2, "apparentPowerTotal", 3),
            (52717, 2, "apparentPowerL1", 3),
            (52719, 2, "apparentPowerL2", 3),
            (52721, 2, "apparentPowerL3", 3)
        ],
        
        "scale_factors": {
            "currentL1": 0.01,
            "currentL2": 0.01,
            "currentL3": 0.01,
            "currentN": 0.01,
            "activePowerTotal": 0.01,
            "activePowerL1": 0.01,
            "activePowerL2": 0.01,
            "activePowerL3": 0.01,
            "reactivePowerTotal": 0.01,
            "reactivePowerL1": 0.01,
            "reactivePowerL2": 0.01,
            "reactivePowerL3": 0.01,
            "apparentPowerTotal": 0.01,
            "apparentPowerL1": 0.01,
            "apparentPowerL2": 0.01,
            "apparentPowerL3": 0.01
        },
        
        "decoding_map": {
            "currentL1": "decode_32bit_unsigned",
            "currentL2": "decode_32bit_unsigned",
            "currentL3": "decode_32bit_unsigned",
            "currentN": "decode_32bit_unsigned",
            "activePowerTotal": "decode_32bit_signed",
            "activePowerL1": "decode_32bit_signed",
            "activePowerL2": "decode_32bit_signed",
            "activePowerL3": "decode_32bit_signed",
            "reactivePowerTotal": "decode_32bit_signed",
            "reactivePowerL1": "decode_32bit_signed",
            "reactivePowerL2": "decode_32bit_signed",
            "reactivePowerL3": "decode_32bit_signed",
            "apparentPowerTotal": "decode_32bit_signed",
            "apparentPowerL1": "decode_32bit_signed",
            "apparentPowerL2": "decode_32bit_signed",
            "apparentPowerL3": "decode_32bit_signed"
        },
        
        "plot_variables": [
            "currentL1",
            "currentL2",
            "currentL3",
            "currentN",
            "activePowerTotal",
            "activePowerL1",
            "activePowerL2",
            "activePowerL3",
            "reactivePowerTotal",
            "reactivePowerL1",
            "reactivePowerL2",
            "reactivePowerL3",
            "apparentPowerTotal",
            "apparentPowerL1",
            "apparentPowerL2",
            "apparentPowerL3"
        ],
        
        "variable_units": {
            "currentL1": "(A)",
            "currentL2": "(A)",
            "currentL3": "(A)",
            "currentN": "(A)",
            "activePowerTotal": "(W)",
            "activePowerL1": "(W)",
            "activePowerL2": "(W)",
            "activePowerL3": "(W)",
            "reactivePowerTotal": "(var)",
            "reactivePowerL1": "(var)",
            "reactivePowerL2": "(var)",
            "reactivePowerL3": "(var)",
            "apparentPowerTotal": "(VA)",
            "apparentPowerL1": "(VA)",
            "apparentPowerL2": "(VA)",
            "apparentPowerL3": "(VA)"
        }
    }
}

SLAVE_DEVICE_ASSIGNMENTS = [
    # {"slaves": range(1, 161), "device_type_key": "CurrentSensor"},
    # {"slaves": range(1, 33), "device_type_key": "CurrentSensor"},
    # {"slaves": range(33, 65), "device_type_key": "IOModule"},
    # {"slaves": range(65, 97), "device_type_key": "INSS-HModule"},
    # {"slaves": range(97, 129), "device_type_key": "D11_D13"},
    {"slaves": [3], "device_type_key": "M4M_30_A2"},
    {"slaves": [10], "device_type_key": "M4M_30_F4"},
    {"slaves": [12], "device_type_key": "M4M_20_F5"}
]

if SLAVE_DEVICE_ASSIGNMENTS:
    first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
    if first_assigned_device_type_key in DEVICE_TYPE:
        VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
    else:
        VARIABLES = []
else:
    VARIABLES = []




# ---------------------- DECODING FUNCTIONS ----------------------

# Remove non-printable characters from string or list
def clean_value(value):
    if isinstance(value, str):
        cleaned = re.sub(r'[^\x20-\x7E]', '', value)
        return cleaned
    elif isinstance(value, list):
        return [clean_value(v) for v in value]
    else:
        return value

# Decode software/hardware version from registers
def decode_version(registers):
    if not registers:
        return None
    return '.'.join(str((reg >> 12) & 0xF) for reg in registers)

# Decode serial number from registers (assuming 64 bits, 4 registers)
def decode_serial_number(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = b''.join(pack('>H', r) for r in registers)
        serial_num = unpack('>Q', b)[0]
        return str(serial_num)
    except Exception as e:
        print(f"Error decoding serial number: {e}")
        return None

# Decode string from registers (2 bytes per register)
def decode_string(registers):
    try:
        if not registers:
            return None
        byte_array = b''.join(pack('>H', r) for r in registers) # Big-endian
        return byte_array.decode('ascii').strip('\0')
    except Exception as e:
        print(f"Error decoding string: {e}")
        return None

# Decode 32-bit signed integer (from 2 registers, big-endian)
def decode_32bit_signed(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>i', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit signed: {e}")
        return None

# Decode 32-bit unsigned integer (from 2 registers, big-endian)
def decode_32bit_unsigned(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>I', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit unsigned: {e}")
        return None

# Decode 64-bit signed integer (from 4 registers, big-endian)
def decode_64bit_signed(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit signed: {e}")
        return None

# Decode 64-bit unsigned integer (from 4 registers, big-endian)
def decode_64bit_unsigned(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>Q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit unsigned: {e}")
        return None

# Decode 16-bit signed integer
def decode_16bit_signed(register):
    try:
        return unpack('>h', pack('>H', register))[0]
    except Exception as e:
        print(f"Error decoding 16-bit signed: {e}")
        return None

# Decode 16-bit unsigned integer
def decode_16bit_unsigned(register):
    try:
        return register
    except Exception as e:
        print(f"Error decoding 16-bit unsigned: {e}")
        return None

# Convert device status to 8-bit binary string (the meaning of each byte is on the Modbus Map)
def decode_device_status(value):
    try:
        return format(int(value), '08b')
    except (TypeError, ValueError) as e:
        print(f"Error decoding device status (value: {value}): {e}")
        return str(value)

# Decode a specific bit from a single register value
def decode_bit(register_value, bit_pos):
    try:
        return (register_value >> bit_pos) & 1
    except Exception as e:
        print(f"Error decoding bit: {e}")
        return None

# Decode a range of bits from a single register value
def decode_bits(register_value, bit_range):
    try:
        start_bit = min(bit_range)
        end_bit = max(bit_range)
        mask = ((1 << (end_bit - start_bit + 1)) - 1) << start_bit
        return (register_value & mask) >> start_bit
    except Exception as e:
        print(f"Error decoding bits: {e}")
        return None

# Decode 32-bit little endian integer from two 16-bit registers
def decode_32bit_little_endian(regs):
    if len(regs) < 2:
        return None
    return (regs[1] << 16) | regs[0]

# Decode 32-bit IEEE 754 float from two 16-bit registers (little-endian order of registers)
def decode_32bit_float(regs):
    try:
        if not regs or len(regs) < 2:
            return None
        b = pack('<HH', regs[0], regs[1])      # Use '<HH' for little-endian word packing
        return unpack('<f', b)[0]              # Use '<f' for little-endian float unpack
    except Exception as e:
        print(f"Error decoding 32-bit float: {e}")
        return None

# Function to read coils
def read_coil(resp):
    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
        return None
    return resp.bits[0]

# Join list into a string
def flatten_value(value):
    if isinstance(value, list):
        return ','.join(str(v) for v in value)
    return value

# Map of decoding function names to actual functions
DECODING_FUNCTIONS_MAP = {
    "decode_version": decode_version,
    "decode_serial_number": decode_serial_number,
    "decode_string": decode_string,
    "decode_32bit_signed": decode_32bit_signed,
    "decode_32bit_unsigned": decode_32bit_unsigned,
    "decode_64bit_signed": decode_64bit_signed,
    "decode_64bit_unsigned": decode_64bit_unsigned,
    "decode_16bit_signed": decode_16bit_signed,
    "decode_16bit_unsigned": decode_16bit_unsigned,
    "decode_device_status": decode_device_status,
    "decode_bit": decode_bit,
    "decode_bits": decode_bits,
    "decode_32bit_little_endian": decode_32bit_little_endian,
    "decode_32bit_float": decode_32bit_float,
    "read_coil": read_coil,
    "flatten_value": flatten_value
}

def get_device_config_for_slave(slave_id):
    for assignment in SLAVE_DEVICE_ASSIGNMENTS:
        if isinstance(assignment["slaves"], range):
            if slave_id in assignment["slaves"]:
                device_type_key = assignment["device_type_key"]
                break
        elif isinstance(assignment["slaves"], list):
            if slave_id in assignment["slaves"]:
                device_type_key = assignment["device_type_key"]
                break
    else:
        print(f"Warning: Slave ID {slave_id} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping this slave.")
        return None

    if device_type_key not in DEVICE_TYPE:
        print(f"Error: Device type key '{device_type_key}' assigned to slave {slave_id} not found in DEVICE_TYPE configuration.")
        return None
        
    return DEVICE_TYPE[device_type_key], device_type_key




# ---------------------- SENSOR POLLING FUNCTION ----------------------

def read_slave_with_client(slave_id, thread_id, client):
    device_config, device_type_key = get_device_config_for_slave(slave_id)
    if device_config is None:
        dummy_len = len(DEVICE_TYPE.get("CurrentSensor", {}).get("variables", []))
        return [None] * dummy_len

    variables = device_config["variables"]
    scale_factors = device_config["scale_factors"]
    decoding_map = device_config["decoding_map"]

    values = []
    thread_prefix = f"[Thread {thread_id}] "

    for var_info in variables:
        addr, size, name, func_code = var_info[:4]
        options = var_info[4] if len(var_info) > 4 else {}

        current_value = None
        try:
            resp = None
            if func_code == 1:
                resp = client.read_coils(address=addr, count=size, slave=slave_id)
                if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Coils read error or empty bits.")
                    values.append(None)
                    continue
                regs_or_bits = resp.bits
            elif func_code == 3:
                resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)
                if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Registers read error or empty registers.")
                    values.append(None)
                    continue
                regs_or_bits = resp.registers
                if func_code == 3 and size > 1:                            # Changes the order of the decodification
                    regs_or_bits = list(reversed(regs_or_bits))
            else:
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Unsupported function code {func_code}.")
                values.append(None)
                continue

            print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

            if name in decoding_map:
                func_name = decoding_map[name]
                decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                if decoder_func:
                    if func_name == "decode_bit":
                        bit_pos = options.get('bit')
                        if bit_pos is not None and regs_or_bits:
                            current_value = decoder_func(regs_or_bits[0], bit_pos)
                        else:
                            print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                    elif func_name == "decode_bits":
                        bit_range = options.get('bits')
                        if bit_range is not None and regs_or_bits:
                            current_value = decoder_func(regs_or_bits[0], bit_range)
                        else:
                            print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                    elif func_name == "read_coil":
                        current_value = decoder_func(resp)
                    elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                        current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                    else:
                        current_value = decoder_func(regs_or_bits)
                else:
                    print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Falling back to raw value processing.")
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
            else:
                current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

            if name in scale_factors and isinstance(current_value, (int, float)):
                current_value *= scale_factors[name]

            if isinstance(current_value, list):
                final_value = flatten_value(current_value)
            else:
                final_value = current_value

            values.append(final_value)

        except Exception as e:
            print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
            values.append(None)

    return values




# ---------------------- SAVE TO EXCEL ----------------------

def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
    wb = Workbook()
    wb.remove(wb.active)
    
    for slave_id in SLAVE_IDS:
        config_result = get_device_config_for_slave(slave_id)

        if config_result is None:
            print(f"Skipping sheet creation for Slave_{slave_id} due to missing config.")
            continue

        device_config, _ = config_result

        if not isinstance(device_config, dict):
            print(f"Skipping sheet creation for Slave_{slave_id} due to invalid device config type.")
            continue

        variables = device_config.get("variables", [])

        headers = ['timestamp']
        for var in variables:
            if isinstance(var, dict):
                headers.append(var.get("name", "unknown_var"))
            elif isinstance(var, (list, tuple)):
                headers.append(var[2] if len(var) > 2 else "unknown_var")
            else:
                headers.append("unknown_var")

        ws = wb.create_sheet(title=f"Slave_{slave_id}")
        ws.append(headers)


    # Preencher os dados nas planilhas
    for poll_entry in all_poll_data:
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']
        
        for slave_id, values in data.items():
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            safe_values = [clean_value(v) for v in values]
            ws.append([timestamp] + safe_values)

    wb.save(filename)

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
    config_result = get_device_config_for_slave(slave_id)
    if config_result is None:
        print(f"Warning: No device configuration found for slave {slave_id}. Cannot create chart for {var_name}.")
        return
    
    device_config, _ = config_result 

    variable_units = device_config.get("variable_units", {})
    unit = variable_units.get(var_name, "")

    chart = ScatterChart()
    chart.title = f"{var_name} over time - ID {slave_id}"
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = f"{var_name} {unit}".strip()
    chart.legend = None

    chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    title_with_unit = f"{var_name} {unit}".strip()
    series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.smooth = True
    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'BC'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)

def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
    wb = load_workbook(filename)

    for slave_id in SLAVE_IDS:
        sheet_name = f"Slave_{slave_id}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        num_rows = ws.max_row
        
        config_result = get_device_config_for_slave(slave_id)
        if config_result is None:
            print(f"Skipping chart generation for Slave_{slave_id} due to missing device configuration.")
            continue
        
        device_config, _ = config_result

        variables_to_plot_for_slave = device_config.get("plot_variables", [])

        chart_index = 0
        timestamp_col = 1

        for var in variables_to_plot_for_slave:
            if var not in headers:
                continue

            var_col = headers.index(var) + 1
            create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
            chart_index += 1

    wb.save(filename)




# ---------------------- MULTITHREADING POLLING FOR EACH CMS BUS PORT ----------------------

def poll_slaves_multithread(stop_event, all_poll_data):
    def poll_slave_ids(slave_ids, all_data, lock, thread_id):
        client = ModbusTcpClient(IP, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect")
            return
 
        def task(slave_id):
            values = read_slave_with_client(slave_id, thread_id, client)
            with lock:
                all_data[slave_id] = values
 
        with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
            executor.map(task, slave_ids)
 
        client.close()
 
    # Main polling loop
    while not stop_event.is_set():
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()
        threads = []
 
        slave_ids = [3, 10, 12]
        t = threading.Thread(target=poll_slave_ids, args=(slave_ids, all_data, lock, 1))
        threads.append(t)
        t.start()
 
        for t in threads:
            t.join()
 
        all_poll_data.append({'timestamp': timestamp, 'data': all_data})
 
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        print(f"\n Completed polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)
 
def poll_slaves_multithread_once(stop_event, all_poll_data):
    if stop_event.is_set():
        return
 
    start_time = time.time()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    all_data = {}
    lock = threading.Lock()
 
    def poll_slave_ids(slave_ids, all_data, lock, thread_id):
        client = ModbusTcpClient(IP, port=PORT)
        if not client.connect():
            print(f"[Thread {thread_id}] Failed to connect")
            return
 
        def task(slave_id):
            values = read_slave_with_client(slave_id, thread_id, client)
            with lock:
                all_data[slave_id] = values
 
        with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
            executor.map(task, slave_ids)
 
        client.close()
 
    slave_ids = [3, 10, 12]
    t = threading.Thread(target=poll_slave_ids, args=(slave_ids, all_data, lock, 1))
    t.start()
    t.join()
 
    all_poll_data.append({'timestamp': timestamp, 'data': all_data})
 
    elapsed = time.time() - start_time
    sleep_time = max(0, POLL_INTERVAL - elapsed)
    print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
    time.sleep(sleep_time)

def periodic_polling_loop(stop_event, all_poll_data):
    while not stop_event.is_set():
        start_time = time.time()
        one_minute = 600
        cycle_duration = 60 * 60

        while time.time() - start_time < one_minute and not stop_event.is_set():
            poll_slaves_multithread_once(stop_event, all_poll_data)

        if stop_event.is_set():
            break

        try:
            print(f"Saving backup with {len(all_poll_data)} readings...")
            save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_BACKUP)
        except Exception as e:
            print(f"Error on backup: {e}")

        print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
        for _ in range(int(cycle_duration)):
            if stop_event.is_set():
                break
            time.sleep(1)




# ---------------------- SPLIT EXCEL FUNCTIONS ----------------------

def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
    if not os.path.exists(path_file_excel):
        st.error(f"Error: File '{path_file_excel}' not found.")
        return None

    try:
        if not os.path.exists(name_output_folder):
            os.makedirs(name_output_folder)
        else:
            for file_name in os.listdir(name_output_folder):
                path_old_file = os.path.join(name_output_folder, file_name)
                if os.path.isfile(path_old_file):
                    os.remove(path_old_file)
            st.info(f"Cleaning folder '{name_output_folder}'.")

        source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in source_workbook.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue

            sheet = source_workbook[sheet_name]

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            for row in sheet.iter_rows(values_only=True):
                new_ws.append(row)

            num_rows = new_ws.max_row
            num_cols = new_ws.max_column

            if num_rows < 2 or num_cols < 2:
                st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                continue

            timestamp_col = 1
            
            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping chart generation for Slave_{slave_id} in split function due to missing device configuration.")
                continue

            variables_to_plot_for_slave = device_config["plot_variables"]

            for idx, var_name in enumerate(variables_to_plot_for_slave): # Use the configured plot variables
                if var_name in [cell.value for cell in new_ws[1]]:
                    var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                    create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
            output_filepath = os.path.join(name_output_folder, output_filename)

            new_wb.save(output_filepath)
            generated_files.append(output_filepath)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
        complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

        with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All files were zipped into '{complete_zip_path}'")
        return complete_zip_path

    except Exception as e:
        st.error(f"Error splitting and generating charts: {e}")
        return None

def copy_sheet_without_charts(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col_dim in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_dim] = copy(source_sheet.column_dimensions[col_dim])
    for row_dim in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row_dim] = copy(source_sheet.row_dimensions[row_dim])

def extract_slave_id(sheet_name):
    match = re.search(r'\d+', sheet_name)
    return int(match.group(0)) if match else None

def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
    if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
        st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
        return None

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    else:
        for file_name in os.listdir(output_folder):
            path_to_remove = os.path.join(output_folder, file_name)
            if os.path.isfile(path_to_remove):
                os.remove(path_to_remove)
        st.info(f"Folder '{output_folder}' cleaned.")

    wb_modbus = openpyxl.load_workbook(path_modbus_log)
    wb_backup = openpyxl.load_workbook(path_backup_data)
    
    generated_files = []

    for i, slave_id in enumerate(selected_slave_ids):
        sheet_name = f"Slave_{slave_id}"
        if status_text:
            status_text.text(f"Processing Slave_{slave_id}...")
        if sheet_name not in wb_modbus.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
            continue
        if sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
            continue

        device_config = get_device_config_for_slave(slave_id)
        if device_config is None:
            print(f"Skipping merge/chart generation for Slave_{slave_id} due to missing device configuration.")
            continue
        
        # Use only name for columns, excluding func_code and options
        columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

        modbus_sheet = wb_modbus[sheet_name]
        backup_sheet = wb_backup[sheet_name]

        modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
        backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

        modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
        backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

        merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
        merged_data.dropna(subset=["timestamp"], inplace=True)
        merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
        merged_data.sort_values("timestamp", inplace=True)
        merged_data.reset_index(drop=True, inplace=True)

        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        new_sheet = new_wb.create_sheet(title=sheet_name)

        headers = [cell.value for cell in modbus_sheet[1]]
        new_sheet.append(headers)

        for r_idx, row in merged_data.iterrows():
            new_sheet.append(list(row.values))

        num_rows = new_sheet.max_row
        current_slave_id = extract_slave_id(sheet_name)
        
        variables_to_plot_for_slave = device_config["plot_variables"]

        for chart_index, var_name in enumerate(variables_to_plot_for_slave): # Use the configured plot variables
            if var_name in headers:
                col_idx = headers.index(var_name) + 1
                create_scatter_chart(
                    new_sheet,
                    timestamp_col=1,
                    var_col=col_idx,
                    num_rows=num_rows,
                    var_name=var_name,
                    slave_id=current_slave_id,
                    chart_index=chart_index
                )
            else:
                st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

        safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
        new_wb.save(path_output)
        generated_files.append(path_output)
        
        if progress_bar and total_slaves > 0:
            progress_percent = int(((i + 1) / total_slaves) * 100)
            progress_bar.progress(progress_percent)

    if not generated_files:
        st.warning("No individual files were generated for the selected slaves.")
        return None

    zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_to_add in generated_files:
            zipf.write(file_to_add, os.path.basename(file_to_add))

    st.success(f"All merged individual files have been zipped into: {zip_path}")
    return zip_path

def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
    if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
        st.error("Excel files (final or backup) not found for comparison.")
        return False # Return False if files are missing

    wb_final = load_workbook(excel_file_final, data_only=True)
    wb_backup = load_workbook(excel_file_backup, data_only=True)

    needs_merge = False
    for slave_id in selected_slave_ids:
        sheet_name = f"Slave_{slave_id}"
        if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
            st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
            continue

        ws_final = wb_final[sheet_name]
        ws_backup = wb_backup[sheet_name]

        timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
        timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

        if not timestamps_backup.issubset(timestamps_final):
            needs_merge = True
            break
    
    return needs_merge




# ---------------------- CREATION OF THE DASHBOARD ----------------------

st.set_page_config(layout="wide")
st.title("Long Term Test Setup for SCU200")

if "polling" not in st.session_state:
    st.session_state.polling = False
    st.session_state.stop_event = threading.Event()
    st.session_state.thread = None
    st.session_state.all_poll_data = []

    st.session_state.selected_slaves_by_type = {
        device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
    }
    st.session_state.realtime_table_placeholders = {}
    st.session_state.last_data_length = 0
    st.session_state.last_ui_update_time = time.time()
    st.session_state.last_table_update_timestamp = "N/A"

st.sidebar.header("Modbus Settings")
st.sidebar.write(f"**IP Address:** {IP}")
st.sidebar.write(f"**Port:** {PORT}")
st.sidebar.write(f"**Poll Interval (seconds):** {POLL_INTERVAL}")
st.sidebar.write(f"**Monitoring Cycle:** 10 min polling / 1 hour pause")
st.sidebar.markdown(f"**Status:** {'RUNNING' if st.session_state.polling else 'STOPPED'}")

st.sidebar.subheader("Active Sensors by Type")
sensor_counts = {}
for assignment in SLAVE_DEVICE_ASSIGNMENTS:
    device_type = assignment["device_type_key"]
    if isinstance(assignment["slaves"], range):
        count = len(list(assignment["slaves"]))
    elif isinstance(assignment["slaves"], list):
        count = len(assignment["slaves"])
    sensor_counts[device_type] = sensor_counts.get(device_type, 0) + count

for device_type, count in sensor_counts.items():
    st.sidebar.write(f"- **{device_type}**: {count} active sensors")

if st.button("Start Monitoring", disabled=st.session_state.polling):
    st.session_state.all_poll_data = []
    st.session_state.stop_event.clear()

    if st.session_state.thread and st.session_state.thread.is_alive():
        st.session_state.stop_event.set()
        st.session_state.thread.join(timeout=2)
        print("Existing polling thread stopped before starting a new one.")

    st.session_state.thread = threading.Thread(
        target=periodic_polling_loop,
        args=(st.session_state.stop_event, st.session_state.all_poll_data),
        daemon=True
    )
    st.session_state.thread.start()
    st.session_state.polling = True
    st.session_state.last_data_length = 0
    st.success("Monitoring started!")

if st.button("Stop Monitoring"):
    if st.session_state.polling:
        st.session_state.stop_event.set()
        if st.session_state.thread and st.session_state.thread.is_alive():
            st.session_state.thread.join(timeout=10)
            if st.session_state.thread.is_alive():
                st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
            else:
                st.info("Polling thread terminated.")
        st.session_state.polling = False

        with st.spinner("Saving data to Excel..."):
            save_all_data_to_excel(st.session_state.all_poll_data, filename=EXCEL_FILE_FINAL)

            try:
                add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                st.success("Monitoring stopped, data and charts saved to Excel.")
            except Exception as e:
                st.error(f"Error adding charts: {e}")
    else:
        st.warning("Monitoring is not running.")

st.divider()

if st.session_state.polling:
    st.subheader("Realtime data monitoring by sensor type")

    st.session_state.realtime_ui_timestamp_placeholder = st.empty()

    assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

    for device_type_key in assigned_device_types_for_display:
        device_config = DEVICE_TYPE.get(device_type_key)
        if not device_config:
            continue

        has_slaves_assigned = any(
            assignment["device_type_key"] == device_type_key
            for assignment in SLAVE_DEVICE_ASSIGNMENTS
        )
        if not has_slaves_assigned:
            continue

        available_slaves_for_type = []
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if assignment["device_type_key"] == device_type_key:
                if isinstance(assignment["slaves"], range):
                    available_slaves_for_type.extend(list(assignment["slaves"]))
                elif isinstance(assignment["slaves"], list):
                    available_slaves_for_type.extend(assignment["slaves"])
        available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

        if not available_slaves_for_type:
            continue

        with st.expander(f"Configuration and Monitoring: **{device_type_key}**"):
            st.session_state.selected_slaves_by_type[device_type_key] = st.multiselect(
                f"Select Slave IDs for {device_type_key}:",
                options=available_slaves_for_type,
                default=st.session_state.selected_slaves_by_type.get(device_type_key, []),
                key=f"multiselect_{device_type_key}"
            )

            st.session_state.realtime_table_placeholders[device_type_key] = st.empty()

    def update_realtime_tables_by_type():
        current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

        st.session_state.realtime_ui_timestamp_placeholder.write(
            f"Last widget update at **{current_ui_update_timestamp}**"
        )

        if st.session_state.all_poll_data:
            latest_poll_entry = st.session_state.all_poll_data[-1]
            raw_data = latest_poll_entry['data']

            for device_type_key, placeholder in st.session_state.realtime_table_placeholders.items():
                selected_slaves = st.session_state.selected_slaves_by_type.get(device_type_key, [])
                device_config = DEVICE_TYPE.get(device_type_key)

                placeholder.empty()

                if not selected_slaves or not device_config:
                    placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                    continue

                table_data = []
                variable_names = [var[2] for var in device_config["variables"]]

                for slave_id in sorted(selected_slaves):
                    if slave_id in raw_data:
                        values = raw_data[slave_id]
                        row_dict = {"Slave ID": f"Slave {slave_id}"}
                        for i, var_name in enumerate(variable_names):
                            display_value = values[i] if i < len(values) and values[i] is not None else "N/A"
                            if isinstance(display_value, float):
                                display_value = f"{display_value:.2f}"
                            row_dict[var_name] = display_value
                        table_data.append(row_dict)
                    else:
                        table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                if table_data:
                    df = pd.DataFrame(table_data)
                    df = df.set_index("Slave ID")

                    placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                else:
                    placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
        else:
            for device_type_key, placeholder in st.session_state.realtime_table_placeholders.items():
                placeholder.empty()
                placeholder.info(f"Waiting for first polling data for {device_type_key}...")

    update_realtime_tables_by_type()

    if len(st.session_state.all_poll_data) > st.session_state.last_data_length:
        st.session_state.last_data_length = len(st.session_state.all_poll_data)
        st.session_state.last_ui_update_time = time.time()
        st.rerun()
    else:
        time_since_last_update = time.time() - st.session_state.last_ui_update_time
        if time_since_last_update >= POLL_INTERVAL:
            st.session_state.last_ui_update_time = time.time()
            st.rerun()
        else:
            time.sleep(max(0, POLL_INTERVAL - time_since_last_update))
            st.rerun()

if not os.path.exists(EXCEL_FILE_FINAL):
    st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
else:
    all_possible_slave_ids = sorted(list(set(SLAVE_IDS)))

    display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

    if "All" in st.session_state.get('selected_display_options', []):
        selected_display_options = st.multiselect(
            "Choose slaves to separate:",
            display_options,
            default=["All"],
            disabled=True
        )
    else:
        selected_display_options = st.multiselect(
            "Choose slaves to separate:",
            display_options,
            default=[]
        )

    st.session_state.selected_display_options = selected_display_options

    selected_slave_ids_numeric = []
    if "All" in selected_display_options:
        selected_slave_ids_numeric = all_possible_slave_ids
    elif selected_display_options:
        for ds_id in selected_display_options:
            match = re.search(r'\d+', ds_id)
            if match:
                selected_slave_ids_numeric.append(int(match.group(0)))

    if st.button("Generate Individual Files"):
        if not selected_slave_ids_numeric:
            st.warning("Please select at least one slave or 'All'.")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()

            total_slaves_to_process = len(selected_slave_ids_numeric)
            zip_path = None

            if not os.path.exists(EXCEL_FILE_BACKUP):
                st.warning("Backup file not found. Performing only the split of the final Excel.")
                with st.spinner("Splitting Excel and generating charts..."):
                    zip_path = split_and_generate_charts_for_slaves(
                        EXCEL_FILE_FINAL,
                        selected_slave_ids_numeric,
                        progress_bar=progress_bar,
                        status_text=status_text,
                        total_slaves=total_slaves_to_process
                    )
            else:
                st.info("Comparing timestamps between final Excel and backup...")
                needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                if needs_merge:
                    st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                    with st.spinner("Merging data from backup and final, and generating individual files..."):
                        zip_path = merge_data_and_generate_charts_for_slaves(
                            EXCEL_FILE_FINAL,
                            EXCEL_FILE_BACKUP,
                            selected_slave_ids_numeric,
                            progress_bar=progress_bar,
                            status_text=status_text,
                            total_slaves=total_slaves_to_process
                        )
                else:
                    st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                    with st.spinner("Splitting Excel and generating charts..."):
                        zip_path = split_and_generate_charts_for_slaves(
                            EXCEL_FILE_FINAL,
                            selected_slave_ids_numeric,
                            progress_bar=progress_bar,
                            status_text=status_text,
                            total_slaves=total_slaves_to_process
                        )

            if zip_path:
                status_text.text("Processing complete!")
                st.success("Individual files generated and available for download.")

if not st.session_state.polling and os.path.exists(EXCEL_FILE_FINAL):
    wb = Workbook()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_FILE_FINAL, read_only=True)
    except Exception:
        pass
    st.subheader("Example of latest reading (first 5 slaves)")
    displayed_slaves = 0
    effective_slave_ids = set()
    for assignment in SLAVE_DEVICE_ASSIGNMENTS:
        if isinstance(assignment["slaves"], range):
            effective_slave_ids.update(assignment["slaves"])
        elif isinstance(assignment["slaves"], list):
            effective_slave_ids.update(assignment["slaves"])

    for sid in sorted(list(effective_slave_ids)):
        if displayed_slaves >= 5:
            break
        if f"Slave_{sid}" not in wb.sheetnames:
            continue
        ws = wb[f"Slave_{sid}"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            st.markdown(f"**Slave {sid}**")
            df = {rows[0][i]: rows[-1][i] for i in range(len(rows[0]))}
            st.write(df)
            displayed_slaves += 1
