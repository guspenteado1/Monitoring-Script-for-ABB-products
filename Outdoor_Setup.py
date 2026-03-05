import os
import re
import time
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from struct import pack, unpack
import zipfile
import pandas as pd
import numpy as np
import streamlit as st
from pymodbus.client import ModbusTcpClient, ModbusSerialClient
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl.chart.axis
from streamlit_autorefresh import st_autorefresh
import warnings

warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", message="missing ScriptRunContext")

# -------------------------
# DECODING FUNCTIONS
# -------------------------
def clean_value(value):
    if isinstance(value, str):
        cleaned = re.sub(r'[^\x20-\x7E]', '', value)
        return cleaned
    elif isinstance(value, list):
        return [clean_value(v) for v in value]
    else:
        return value

def decode_version(registers):
    if not registers:
        return None
    return '.'.join(str((reg >> 12) & 0xF) for reg in registers)

def decode_serial_number(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = b''.join(pack('>H', r) for r in registers)
        serial_num = unpack('>Q', b)[0]
        return str(serial_num)
    except Exception as e:
        print(f"Error decoding serial number: {e}")
        return None

def decode_string(registers):
    try:
        if not registers:
            return None
        byte_array = b''.join(pack('>H', r) for r in registers)
        return byte_array.decode('ascii').strip('\0')
    except Exception as e:
        print(f"Error decoding string: {e}")
        return None

def decode_32bit_signed(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>i', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit signed: {e}")
        return None

def decode_32bit_unsigned(registers):
    try:
        if not registers or len(registers) < 2:
            return None
        b = pack('>HH', registers[0], registers[1])
        return unpack('>I', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit unsigned: {e}")
        return None

def decode_64bit_signed(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit signed: {e}")
        return None

def decode_64bit_unsigned(registers):
    try:
        if not registers or len(registers) < 4:
            return None
        b = pack('>HHHH', registers[0], registers[1], registers[2], registers[3])
        return unpack('>Q', b)[0]
    except Exception as e:
        print(f"Error decoding 64-bit unsigned: {e}")
        return None

def decode_16bit_signed(register):
    try:
        return unpack('>h', pack('>H', register))[0]
    except Exception as e:
        print(f"Error decoding 16-bit signed: {e}")
        return None

def decode_16bit_unsigned(register):
    try:
        return register
    except Exception as e:
        print(f"Error decoding 16-bit unsigned: {e}")
        return None

def decode_device_status(value):
    try:
        return format(int(value), '08b')
    except (TypeError, ValueError) as e:
        print(f"Error decoding device status (value: {value}): {e}")
        return str(value)

def decode_bit(register_value, bit_pos):
    try:
        return (register_value >> bit_pos) & 1
    except Exception as e:
        print(f"Error decoding bit: {e}")
        return None

def decode_bits(register_value, bit_range):
    try:
        start_bit = min(bit_range)
        end_bit = max(bit_range)
        mask = ((1 << (end_bit - start_bit + 1)) - 1) << start_bit
        return (register_value & mask) >> start_bit
    except Exception as e:
        print(f"Error decoding bits: {e}")
        return None

def decode_32bit_little_endian(regs):
    if len(regs) < 2:
        return None
    return (regs[1] << 16) | regs[0]

def decode_32bit_float(regs):
    try:
        if not regs or len(regs) < 2:
            return None
        b = pack('<HH', regs[0], regs[1])
        return unpack('<f', b)[0]
    except Exception as e:
        print(f"Error decoding 32-bit float: {e}")
        return None

def read_coil(resp):
    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
        return None
    return resp.bits[0]

def flatten_value(value):
    if isinstance(value, list):
        return ','.join(str(v) for v in value)
    return value

DECODING_FUNCTIONS_MAP = {
    "decode_version": decode_version,
    "decode_serial_number": decode_serial_number,
    "decode_string": decode_string,
    "decode_32bit_signed": decode_32bit_signed,
    "decode_32bit_unsigned": decode_32bit_unsigned,
    "decode_64bit_signed": decode_64bit_signed,
    "decode_64bit_unsigned": decode_64bit_unsigned,
    "decode_16bit_signed": decode_16bit_signed,
    "decode_16bit_unsigned": decode_16bit_unsigned,
    "decode_device_status": decode_device_status,
    "decode_bit": decode_bit,
    "decode_bits": decode_bits,
    "decode_32bit_little_endian": decode_32bit_little_endian,
    "decode_32bit_float": decode_32bit_float,
    "read_coil": read_coil,
    "flatten_value": flatten_value
}



# -------------------------
# SCRIPT 1: M4M TCP
# -------------------------
def run_tcp():

    # ---------------------------------------- CONFIGURATION SECTION
    PORT = 502
    SENSOR_IPS = [f"192.168.1.{i}" for i in list(range(2, 11)) + list(range(14, 19)) + list(range(20, 21))]
    POLL_INTERVAL = 60
    EXCEL_FILE_FINAL = "M4M_TCP_modbus_poll_log.xlsx"
    EXCEL_FILE_BACKUP = "M4M_TCP_backup_poll_data.xlsx"
    state_prefix = "m4m_tcp_"

    DEVICE_TYPE = {
        "M4M": {
            "variables": [
                (20480, 4, "activeImportEnergyTotal", 3),
                (20484, 4, "activeExportEnergyTotal", 3),
                (20488, 4, "activeNetEnergyTotal", 3),
                (20492, 4, "reactiveImportEnergyTotal", 3),
                (20496, 4, "reactiveExportEnergyTotal", 3),
                (20500, 4, "reactiveNetEnergyTotal", 3),
                (20504, 4, "apparentImportEnergyTotal", 3),
                (20508, 4, "apparentExportEnergyTotal", 3),
                (20512, 4, "apparentNetEnergyTotal", 3),
                (21600, 4, "activeImpEnergyL1", 3),
                (21604, 4, "activeImpEnergyL2", 3),
                (21608, 4, "activeImpEnergyL3", 3),
                (21612, 4, "activeExpEnergyL1", 3),
                (21616, 4, "activeExpEnergyL2", 3),
                (21620, 4, "activeExpEnergyL3", 3),
                (21624, 4, "activeNetEnergyL1", 3),
                (21628, 4, "activeNetEnergyL2", 3),
                (21632, 4, "activeNetEnergyL3", 3),
                (21636, 4, "reactiveImpEnergyL1", 3),
                (21640, 4, "reactiveImpEnergyL2", 3),
                (21644, 4, "reactiveImpEnergyL3", 3),
                (21648, 4, "reactiveExpEnergyL1", 3),
                (21652, 4, "reactiveExpEnergyL2", 3),
                (21656, 4, "reactiveExpEnergyL3", 3),
                (21660, 4, "reactiveNetEnergyL1", 3),
                (21664, 4, "reactiveNetEnergyL2", 3),
                (21668, 4, "reactiveNetEnergyL3", 3),
                (21672, 4, "apparentImpEnergyL1", 3),
                (21676, 4, "apparentImpEnergyL2", 3),
                (21680, 4, "apparentImpEnergyL3", 3),
                (21684, 4, "apparentExpEnergyL1", 3),
                (21688, 4, "apparentExpEnergyL2", 3),
                (21692, 4, "apparentExpEnergyL3", 3),
                (21696, 4, "apparentNetEnergyL1", 3),
                (21700, 4, "apparentNetEnergyL2", 3),
                (21704, 4, "apparentNetEnergyL3", 3),
                (23312, 2, "currentL1", 3),
                (23314, 2, "currentL2", 3),
                (23316, 2, "currentL3", 3),
                (23318, 2, "currentN", 3),
                (23322, 2, "activePowerTotal", 3),
                (23324, 2, "activePowerL1", 3),
                (23326, 2, "activePowerL2", 3),
                (23328, 2, "activePowerL3", 3),
                (23330, 2, "reactivePowerTotal", 3),
                (23332, 2, "reactivePowerL1", 3),
                (23334, 2, "reactivePowerL2", 3),
                (23336, 2, "reactivePowerL3", 3),
                (23338, 2, "apparentPowerTotal", 3),
                (23340, 2, "apparentPowerL1", 3),
                (23342, 2, "apparentPowerL2", 3),
                (23344, 2, "apparentPowerL3", 3)
            ],
            
            "scale_factors": {
                "activeImportEnergyTotal": 0.01,
                "activeExportEnergyTotal": 0.01,
                "activeNetEnergyTotal": 0.01,
                "reactiveImportEnergyTotal": 0.01,
                "reactiveExportEnergyTotal": 0.01,
                "reactiveNetEnergyTotal": 0.01,
                "apparentImportEnergyTotal": 0.01,
                "apparentExportEnergyTotal": 0.01,
                "apparentNetEnergyTotal": 0.01,
                "activeImpEnergyL1": 0.01,
                "activeImpEnergyL2": 0.01,
                "activeImpEnergyL3": 0.01,
                "activeExpEnergyL1": 0.01,
                "activeExpEnergyL2": 0.01,
                "activeExpEnergyL3": 0.01,
                "activeNetEnergyL1": 0.01,
                "activeNetEnergyL2": 0.01,
                "activeNetEnergyL3": 0.01,
                "reactiveImpEnergyL1": 0.01,
                "reactiveImpEnergyL2": 0.01,
                "reactiveImpEnergyL3": 0.01,
                "reactiveExpEnergyL1": 0.01,
                "reactiveExpEnergyL2": 0.01,
                "reactiveExpEnergyL3": 0.01,
                "reactiveNetEnergyL1": 0.01,
                "reactiveNetEnergyL2": 0.01,
                "reactiveNetEnergyL3": 0.01,
                "apparentImpEnergyL1": 0.01,
                "apparentImpEnergyL2": 0.01,
                "apparentImpEnergyL3": 0.01,
                "apparentExpEnergyL1": 0.01,
                "apparentExpEnergyL2": 0.01,
                "apparentExpEnergyL3": 0.01,
                "apparentNetEnergyL1": 0.01,
                "apparentNetEnergyL2": 0.01,
                "apparentNetEnergyL3": 0.01,
                "currentL1": 0.01,
                "currentL2": 0.01,
                "currentL3": 0.01,
                "currentN": 0.01,
                "activePowerTotal": 0.01,
                "activePowerL1": 0.01,
                "activePowerL2": 0.01,
                "activePowerL3": 0.01,
                "reactivePowerTotal": 0.01,
                "reactivePowerL1": 0.01,
                "reactivePowerL2": 0.01,
                "reactivePowerL3": 0.01,
                "apparentPowerTotal": 0.01,
                "apparentPowerL1": 0.01,
                "apparentPowerL2": 0.01,
                "apparentPowerL3": 0.01
            },
            
            "decoding_map": {
                "activeImportEnergyTotal": "decode_64bit_unsigned",
                "activeExportEnergyTotal": "decode_64bit_unsigned",
                "activeNetEnergyTotal": "decode_64bit_signed",
                "reactiveImportEnergyTotal": "decode_64bit_unsigned",
                "reactiveExportEnergyTotal": "decode_64bit_unsigned",
                "reactiveNetEnergyTotal": "decode_64bit_signed",
                "apparentImportEnergyTotal": "decode_64bit_unsigned",
                "apparentExportEnergyTotal": "decode_64bit_unsigned",
                "apparentNetEnergyTotal": "decode_64bit_signed",
                "activeImpEnergyL1": "decode_64bit_unsigned",
                "activeImpEnergyL2": "decode_64bit_unsigned",
                "activeImpEnergyL3": "decode_64bit_unsigned",
                "activeExpEnergyL1": "decode_64bit_unsigned",
                "activeExpEnergyL2": "decode_64bit_unsigned",
                "activeExpEnergyL3": "decode_64bit_unsigned",
                "activeNetEnergyL1": "decode_64bit_signed",
                "activeNetEnergyL2": "decode_64bit_signed",
                "activeNetEnergyL3": "decode_64bit_signed",
                "reactiveImpEnergyL1": "decode_64bit_unsigned",
                "reactiveImpEnergyL2": "decode_64bit_unsigned",
                "reactiveImpEnergyL3": "decode_64bit_unsigned",
                "reactiveExpEnergyL1": "decode_64bit_unsigned",
                "reactiveExpEnergyL2": "decode_64bit_unsigned",
                "reactiveExpEnergyL3": "decode_64bit_unsigned",
                "reactiveNetEnergyL1": "decode_64bit_signed",
                "reactiveNetEnergyL2": "decode_64bit_signed",
                "reactiveNetEnergyL3": "decode_64bit_signed",
                "apparentImpEnergyL1": "decode_64bit_unsigned",
                "apparentImpEnergyL2": "decode_64bit_unsigned",
                "apparentImpEnergyL3": "decode_64bit_unsigned",
                "apparentExpEnergyL1": "decode_64bit_unsigned",
                "apparentExpEnergyL2": "decode_64bit_unsigned",
                "apparentExpEnergyL3": "decode_64bit_unsigned",
                "apparentNetEnergyL1": "decode_64bit_signed",
                "apparentNetEnergyL2": "decode_64bit_signed",
                "apparentNetEnergyL3": "decode_64bit_signed",
                "currentL1": "decode_32bit_unsigned",
                "currentL2": "decode_32bit_unsigned",
                "currentL3": "decode_32bit_unsigned",
                "currentN": "decode_32bit_unsigned",
                "activePowerTotal": "decode_32bit_signed",
                "activePowerL1": "decode_32bit_signed",
                "activePowerL2": "decode_32bit_signed",
                "activePowerL3": "decode_32bit_signed",
                "reactivePowerTotal": "decode_32bit_signed",
                "reactivePowerL1": "decode_32bit_signed",
                "reactivePowerL2": "decode_32bit_signed",
                "reactivePowerL3": "decode_32bit_signed",
                "apparentPowerTotal": "decode_32bit_signed",
                "apparentPowerL1": "decode_32bit_signed",
                "apparentPowerL2": "decode_32bit_signed",
                "apparentPowerL3": "decode_32bit_signed"
            },
            
            "plot_variables": [
                "activeImportEnergyTotal",
                "activeExportEnergyTotal",
                "activeNetEnergyTotal",
                "reactiveImportEnergyTotal",
                "reactiveExportEnergyTotal",
                "reactiveNetEnergyTotal",
                "apparentImportEnergyTotal",
                "apparentExportEnergyTotal",
                "apparentNetEnergyTotal",
                "activeImpEnergyL1",
                "activeImpEnergyL2",
                "activeImpEnergyL3",
                "activeExpEnergyL1",
                "activeExpEnergyL2",
                "activeExpEnergyL3",
                "activeNetEnergyL1",
                "activeNetEnergyL2",
                "activeNetEnergyL3",
                "reactiveImpEnergyL1",
                "reactiveImpEnergyL2",
                "reactiveImpEnergyL3",
                "reactiveExpEnergyL1",
                "reactiveExpEnergyL2",
                "reactiveExpEnergyL3",
                "reactiveNetEnergyL1",
                "reactiveNetEnergyL2",
                "reactiveNetEnergyL3",
                "apparentImpEnergyL1",
                "apparentImpEnergyL2",
                "apparentImpEnergyL3",
                "apparentExpEnergyL1",
                "apparentExpEnergyL2",
                "apparentExpEnergyL3",
                "apparentNetEnergyL1",
                "apparentNetEnergyL2",
                "apparentNetEnergyL3",
                "currentL1",
                "currentL2",
                "currentL3",
                "currentN",
                "activePowerTotal",
                "activePowerL1",
                "activePowerL2",
                "activePowerL3",
                "reactivePowerTotal",
                "reactivePowerL1",
                "reactivePowerL2",
                "reactivePowerL3",
                "apparentPowerTotal",
                "apparentPowerL1",
                "apparentPowerL2",
                "apparentPowerL3"
            ],
            
            "variable_units": {
                "activeImportEnergyTotal": "(kWh)",
                "activeExportEnergyTotal": "(kWh)",
                "activeNetEnergyTotal": "(kWh)",
                "reactiveImportEnergyTotal": "(kVarh)",
                "reactiveExportEnergyTotal": "(kVarh)",
                "reactiveNetEnergyTotal": "(kVarh)",
                "apparentImportEnergyTotal": "(kVAh)",
                "apparentExportEnergyTotal": "(kVAh)",
                "apparentNetEnergyTotal": "(kVAh)",
                "activeImpEnergyL1": "(kWh)",
                "activeImpEnergyL2": "(kWh)",
                "activeImpEnergyL3": "(kWh)",
                "activeExpEnergyL1": "(kWh)",
                "activeExpEnergyL2": "(kWh)",
                "activeExpEnergyL3": "(kWh)",
                "activeNetEnergyL1": "(kWh)",
                "activeNetEnergyL2": "(kWh)",
                "activeNetEnergyL3": "(kWh)",
                "reactiveImpEnergyL1": "(kVarh)",
                "reactiveImpEnergyL2": "(kVarh)",
                "reactiveImpEnergyL3": "(kVarh)",
                "reactiveExpEnergyL1": "(kVarh)",
                "reactiveExpEnergyL2": "(kVarh)",
                "reactiveExpEnergyL3": "(kVarh)",
                "reactiveNetEnergyL1": "(kVarh)",
                "reactiveNetEnergyL2": "(kVarh)",
                "reactiveNetEnergyL3": "(kVarh)",
                "apparentImpEnergyL1": "(kVAh)",
                "apparentImpEnergyL2": "(kVAh)",
                "apparentImpEnergyL3": "(kVAh)",
                "apparentExpEnergyL1": "(kVAh)",
                "apparentExpEnergyL2": "(kVAh)",
                "apparentExpEnergyL3": "(kVAh)",
                "apparentNetEnergyL1": "(kVAh)",
                "apparentNetEnergyL2": "(kVAh)",
                "apparentNetEnergyL3": "(kVAh)",
                "currentL1": "(A)",
                "currentL2": "(A)",
                "currentL3": "(A)",
                "currentN": "(A)",
                "activePowerTotal": "(W)",
                "activePowerL1": "(W)",
                "activePowerL2": "(W)",
                "activePowerL3": "(W)",
                "reactivePowerTotal": "(var)",
                "reactivePowerL1": "(var)",
                "reactivePowerL2": "(var)",
                "reactivePowerL3": "(var)",
                "apparentPowerTotal": "(VA)",
                "apparentPowerL1": "(VA)",
                "apparentPowerL2": "(VA)",
                "apparentPowerL3": "(VA)"
            }
        }
    }

    SLAVE_DEVICE_ASSIGNMENTS = [{"ips": SENSOR_IPS, "device_type_key": "M4M"}]

    if SLAVE_DEVICE_ASSIGNMENTS:
        first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
        if first_assigned_device_type_key in DEVICE_TYPE:
            VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
        else:
            VARIABLES = []
    else:
        VARIABLES = []



    # --------------------------------------------- HELPER FUNCTIONS
    def get_device_config_for_ip(ip_address):
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if "ips" in assignment:
                if ip_address in assignment["ips"]:
                    device_type_key = assignment["device_type_key"]
                    if device_type_key in DEVICE_TYPE:
                        return DEVICE_TYPE[device_type_key]
                    else:
                        print(f"Error: Device type key '{device_type_key}' not found in DEVICE_TYPE.")
                        return None
        print(f"Warning: IP {ip_address} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping.")
        return None



    # ------------------------------------- SENSOR POLLING FUNCTION
    def read_sensor_with_client(ip_address, thread_id, client):
        device_config = get_device_config_for_ip(ip_address)
        if device_config is None:
            dummy_len = len(DEVICE_TYPE.get("M4M", {}).get("variables", []))
            return [None] * dummy_len

        variables = device_config["variables"]
        scale_factors = device_config["scale_factors"]
        decoding_map = device_config["decoding_map"]

        values = []
        thread_prefix = f"[Thread {thread_id}] "

        for var_info in variables:
            addr, size, name, func_code = var_info[:4]
            options = var_info[4] if len(var_info) > 4 else {}

            current_value = None
            try:
                resp = None            
                if func_code == 1:
                    resp = client.read_coils(address=addr, count=size)
                    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                        print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Coils read error or empty bits.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.bits
                elif func_code == 3:
                    resp = client.read_holding_registers(address=addr, count=size)
                    if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                        print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Registers read error or empty registers.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.registers
                else:
                    print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Unsupported function code {func_code}.")
                    values.append(None)
                    continue

                print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

                if name in decoding_map:
                    func_name = decoding_map[name]
                    decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                    if decoder_func:
                        if func_name == "decode_bit":
                            bit_pos = options.get('bit')
                            if bit_pos is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_pos)
                            else:
                                print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                        elif func_name == "decode_bits":
                            bit_range = options.get('bits')
                            if bit_range is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_range)
                            else:
                                print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                        elif func_name == "read_coil":
                            current_value = decoder_func(resp)
                        elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                            current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                        else:
                            current_value = decoder_func(regs_or_bits)
                    else:
                        print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Using raw data.")
                        current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
                else:
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

                if name in scale_factors and isinstance(current_value, (int, float)):
                    current_value *= scale_factors[name]

                final_value = flatten_value(current_value) if isinstance(current_value, list) else current_value
                values.append(final_value)

            except Exception as e:
                print(f"{thread_prefix}Sensor {ip_address} | Addr {addr} | Exception: {e}")
                values.append(None)

        return values



    # ------------------------------------------------ SAVE TO EXCEL
    def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
        wb = Workbook()
        wb.remove(wb.active)

        for ip in SENSOR_IPS:
            device_config = get_device_config_for_ip(ip)
            if device_config is None:
                print(f"Skipping sheet creation for Sensor_{ip} due to missing device configuration.")
                continue

            sheet_name = f"Sensor_{ip.replace('.', '_')}"
            ws = wb.create_sheet(title=sheet_name)
            headers = ['timestamp'] + [var[2] for var in device_config["variables"]]
            ws.append(headers)

        for poll_entry in all_poll_data:
            timestamp = poll_entry['timestamp']
            data = poll_entry['data']
            for ip, values in data.items():
                sheet_name = f"Sensor_{ip.replace('.', '_')}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

        wb.save(filename)

    def append_poll_data_to_backup(poll_data, filename=EXCEL_FILE_BACKUP):
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        for ip, values in poll_data["data"].items():
            device_config = get_device_config_for_ip(ip)
            if device_config is None:
                print(f"Skipping Sensor_{ip} due to missing config.")
                continue

            sheet_name = f"Sensor_{ip.replace('.', '_')}"

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                headers = ['timestamp'] + [var[2] for var in device_config["variables"]]
                ws.append(headers)

            safe_values = [clean_value(v) for v in values]
            ws.append([poll_data["timestamp"]] + safe_values)

        wb.save(filename)

    def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, ip_address, chart_index):
        device_config = get_device_config_for_ip(ip_address)
        if device_config is None:
            print(f"Warning: No device configuration found for sensor {ip_address}. Cannot create chart for {var_name}.")
            return

        variable_units = device_config.get("variable_units", {})
        unit = variable_units.get(var_name, "")

        chart = ScatterChart()
        chart.title = f"{var_name} over time - IP {ip_address}"
        chart.style = 13
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = f"{var_name} {unit}".strip()
        chart.legend = None

        chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.x_axis.majorUnit = 500

        x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
        y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

        title_with_unit = f"{var_name} {unit}".strip()
        series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.smooth = True
        chart.series.append(series)

        chart.x_axis.axId = 10 + chart_index * 2
        chart.y_axis.axId = 20 + chart_index * 2
        chart.y_axis.crossAx = 10 + chart_index * 2
        chart.x_axis.crossAx = 20 + chart_index * 2

        chart.height = 9.5
        chart.width = 20.44

        anchor_col = 'BC'
        anchor_row = 1 + 20 * chart_index
        chart.anchor = f"{anchor_col}{anchor_row}"

        sheet.add_chart(chart)

    def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
        wb = load_workbook(filename)

        for ip in SENSOR_IPS:
            sheet_name = f"Sensor_{ip.replace('.', '_')}"
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            num_rows = ws.max_row

            device_config = get_device_config_for_ip(ip)
            if device_config is None:
                print(f"Skipping chart generation for sensor {ip} due to missing device configuration.")
                continue

            variables_to_plot = device_config.get("plot_variables", [])
            timestamp_col = 1
            chart_index = 0

            for var in variables_to_plot:
                if var not in headers:
                    continue
                var_col = headers.index(var) + 1
                create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, ip, chart_index)
                chart_index += 1

        wb.save(filename)



    # ---------------------- MULTITHREADING POLLING FOR EACH CMS BUS PORT
    def poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix):
        if stop_event.is_set():
            return

        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()

        def poll_ip(ip, thread_id):
            client = ModbusTcpClient(ip, port=PORT)
            if not client.connect():
                print(f"[Thread {thread_id}] Failed to connect to {ip}")
                return
            values = read_sensor_with_client(ip, thread_id, client)
            with lock:
                all_data[ip] = values
            client.close()

        with ThreadPoolExecutor(max_workers=len(SENSOR_IPS)) as executor:
            for i, ip in enumerate(SENSOR_IPS):
                executor.submit(poll_ip, ip, i + 1)
                
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        all_poll_data.append({
            'timestamp': timestamp,
            'data': all_data,
        })

        print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

        st.session_state[f'{state_prefix}last_valid_poll_data'] = {
            'timestamp': timestamp,
            'data': all_data
        }

        try:
            append_poll_data_to_backup({
                'timestamp': timestamp,
                'data': all_data
            })
        except Exception as e:
            print(f"Error saving to backup file: {e}")

    def periodic_polling_loop(stop_event, all_poll_data, state_prefix):
        while not stop_event.is_set():
            start_time = time.time()
            one_cycle = 600
            cycle_duration = 60 * 60

            while time.time() - start_time < one_cycle and not stop_event.is_set():
                poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix)

            if stop_event.is_set():
                break

            all_poll_data.append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': {},
            })

            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = True
            print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
            
            for _ in range(int(cycle_duration)):
                if stop_event.is_set():
                    break
                time.sleep(1)
            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = False



    # ---------------------------------------- SPLIT EXCEL FUNCTIONS
    def split_and_generate_charts_for_ips(path_file_excel, selected_ips, name_output_folder="single_sensors_with_generated_charts", progress_bar=None, status_text=None, total_sensors=0):
        if not os.path.exists(path_file_excel):
            st.error(f"Error: File '{path_file_excel}' not found.")
            return None

        try:
            if not os.path.exists(name_output_folder):
                os.makedirs(name_output_folder)
            else:
                for file_name in os.listdir(name_output_folder):
                    path_old_file = os.path.join(name_output_folder, file_name)
                    if os.path.isfile(path_old_file):
                        os.remove(path_old_file)
                st.info(f"Cleaning folder '{name_output_folder}'.")

            source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
            generated_files = []

            for i, ip in enumerate(selected_ips):
                safe_ip = ip.replace('.', '_')
                sheet_name = f"Sensor_{safe_ip}"
                if status_text:
                    status_text.text(f"Processing Sensor {ip}...")
                if sheet_name not in source_workbook.sheetnames:
                    st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                    continue

                sheet = source_workbook[sheet_name]

                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in sheet.iter_rows(values_only=True):
                    new_ws.append(row)

                num_rows = new_ws.max_row
                num_cols = new_ws.max_column

                if num_rows < 2 or num_cols < 2:
                    st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                    continue

                timestamp_col = 1

                device_config = get_device_config_for_ip(ip)
                if device_config is None:
                    print(f"Skipping chart generation for Sensor {ip} in split function due to missing device configuration.")
                    continue

                variables_to_plot = device_config.get("plot_variables", [])

                for idx, var_name in enumerate(variables_to_plot):
                    if var_name in [cell.value for cell in new_ws[1]]:
                        var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                        create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, ip, idx)
                    else:
                        st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

                output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
                output_filepath = os.path.join(name_output_folder, output_filename)

                new_wb.save(output_filepath)
                generated_files.append(output_filepath)

                if progress_bar and total_sensors > 0:
                    progress_percent = int(((i + 1) / total_sensors) * 100)
                    progress_bar.progress(progress_percent)

            if not generated_files:
                st.warning("No individual files were generated for the selected sensors.")
                return None

            name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
            complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

            with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_to_add in generated_files:
                    zipf.write(file_to_add, os.path.basename(file_to_add))

            st.success(f"All files were zipped into '{complete_zip_path}'")
            return complete_zip_path

        except Exception as e:
            st.error(f"Error splitting and generating charts: {e}")
            return None

    def merge_data_and_generate_charts_for_ips(path_modbus_log, path_backup_data, selected_ips, output_folder="merged_sensors", progress_bar=None, status_text=None, total_sensors=0):
        if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
            st.error("Error: One of the main files (M4M_TCP_modbus_poll_log.xlsx or M4M_TCP_backup_poll_data.xlsx) was not found.")
            return None

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        else:
            for file_name in os.listdir(output_folder):
                path_to_remove = os.path.join(output_folder, file_name)
                if os.path.isfile(path_to_remove):
                    os.remove(path_to_remove)
            st.info(f"Folder '{output_folder}' cleaned.")

        wb_modbus = openpyxl.load_workbook(path_modbus_log)
        wb_backup = openpyxl.load_workbook(path_backup_data)

        generated_files = []

        for i, ip in enumerate(selected_ips):
            safe_ip = ip.replace('.', '_')
            sheet_name = f"Sensor_{safe_ip}"
            if status_text:
                status_text.text(f"Processing Sensor {ip}...")
            if sheet_name not in wb_modbus.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue
            if sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
                continue

            device_config = get_device_config_for_ip(ip)
            if device_config is None:
                print(f"Skipping merge/chart generation for Sensor {ip} due to missing device configuration.")
                continue

            columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

            modbus_sheet = wb_modbus[sheet_name]
            backup_sheet = wb_backup[sheet_name]

            modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
            backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

            modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
            backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

            merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
            merged_data.dropna(subset=["timestamp"], inplace=True)
            merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
            merged_data.sort_values("timestamp", inplace=True)
            merged_data.reset_index(drop=True, inplace=True)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            new_sheet = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in modbus_sheet[1]]
            new_sheet.append(headers)

            for _, row in merged_data.iterrows():
                new_sheet.append(list(row.values))

            num_rows = new_sheet.max_row
            variables_to_plot = device_config.get("plot_variables", [])

            for chart_index, var_name in enumerate(variables_to_plot):
                if var_name in headers:
                    col_idx = headers.index(var_name) + 1
                    create_scatter_chart(
                        new_sheet,
                        timestamp_col=1,
                        var_col=col_idx,
                        num_rows=num_rows,
                        var_name=var_name,
                        ip_address=ip,
                        chart_index=chart_index
                    )
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
            new_wb.save(path_output)
            generated_files.append(path_output)

            if progress_bar and total_sensors > 0:
                progress_percent = int(((i + 1) / total_sensors) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected sensors.")
            return None

        zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_sensor_files.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All merged individual files have been zipped into: {zip_path}")
        return zip_path

    def compare_timestamps_and_decide_action_ips(excel_file_final, excel_file_backup, selected_ips):
        if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
            st.error("Excel files (final or backup) not found for comparison.")
            return False

        wb_final = load_workbook(excel_file_final, data_only=True)
        wb_backup = load_workbook(excel_file_backup, data_only=True)

        needs_merge = False
        for ip in selected_ips:
            sheet_name = f"Sensor_{ip.replace('.', '_')}"
            if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
                continue

            ws_final = wb_final[sheet_name]
            ws_backup = wb_backup[sheet_name]

            timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
            timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

            if not timestamps_backup.issubset(timestamps_final):
                needs_merge = True
                break

        return needs_merge



    # ------------------------------------------ STREAMLIT DASHBOARD
    with st.container():
        st.title("M4M Modbus Ethernet TCP Setup")

        if f'{state_prefix}polling' not in st.session_state:
            st.session_state[f'{state_prefix}last_valid_poll_data'] = {}
            st.session_state[f'{state_prefix}polling'] = False
            st.session_state[f'{state_prefix}stop_event'] = threading.Event()
            st.session_state[f'{state_prefix}thread'] = None
            st.session_state[f'{state_prefix}all_poll_data'] = []
            
            st.session_state[f'{state_prefix}selected_slaves_by_type'] = {
                device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
            }
            st.session_state[f'{state_prefix}realtime_table_placeholders'] = {}
            st.session_state[f'{state_prefix}last_data_length'] = 0
            st.session_state[f'{state_prefix}last_ui_update_time'] = time.time()
            st.session_state[f'{state_prefix}last_table_update_timestamp'] = "N/A"

        if st.button("Start Monitoring", key=f"start_monitoring_{state_prefix}", disabled=st.session_state[f'{state_prefix}polling']):
            if not st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}all_poll_data'] = []
                st.session_state[f'{state_prefix}stop_event'].clear()

                st.session_state[f'{state_prefix}thread'] = threading.Thread(
                    target=periodic_polling_loop,
                    args=(
                        st.session_state[f'{state_prefix}stop_event'],
                        st.session_state[f'{state_prefix}all_poll_data'],
                        state_prefix
                    ),
                    daemon=True
                )

                st.session_state[f'{state_prefix}thread'].start()
                st.session_state[f'{state_prefix}polling'] = True
                st.session_state[f'{state_prefix}last_data_length'] = 0
                st.success("Monitoring started!")

        if st.button("Stop Monitoring", key=f"stop_monitoring_{state_prefix}"):
            if st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}stop_event'].set()
                if st.session_state[f'{state_prefix}thread'] and st.session_state[f'{state_prefix}thread'].is_alive():
                    st.session_state[f'{state_prefix}thread'].join(timeout=10)
                    if st.session_state[f'{state_prefix}thread'].is_alive():
                        st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
                    else:
                        st.info("Polling thread terminated.")
                st.session_state[f'{state_prefix}polling'] = False

                with st.spinner("Saving data to Excel..."):
                    save_all_data_to_excel(st.session_state[f'{state_prefix}all_poll_data'], filename=EXCEL_FILE_FINAL)

                    try:
                        add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                        st.success("Monitoring stopped, data and charts saved to Excel.")
                    except Exception as e:
                        st.error(f"Error adding charts: {e}")
            else:
                st.warning("Monitoring is not running.")

        st.divider()

        if st.session_state.get(f"{state_prefix}all_poll_data"):

            all_poll_data = st.session_state[f"{state_prefix}all_poll_data"]

            st.markdown(
                f"Total of **{int(len(all_poll_data))}** reading(s) completed so far."
            )

            device_counts = {}

            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                key = assignment["device_type_key"]
                normalized_key = re.match(r'^[A-Za-z]+', key)
                normalized_key = normalized_key.group(0) if normalized_key else key
                if "ips" in assignment:
                    count = len(assignment["ips"])
                elif "slaves" in assignment:
                    count = len(assignment["slaves"])
                else:
                    count = 0
                device_counts[key] = device_counts.get(key, 0) + count

            device_count_lines = [f"- **{k}**: {v} device(s)" for k, v in device_counts.items()]
            st.markdown("### Active devices on the system:")
            st.markdown("\n".join(device_count_lines))

            last_poll_data = all_poll_data[-1]
            if "elapsed" not in last_poll_data and len(all_poll_data) >= 2:
                last_poll_data = all_poll_data[-2]

            last_poll_time = last_poll_data.get("elapsed", None)
            last_sleep_time = last_poll_data.get("sleep_time", None)

            if last_poll_time is not None and last_sleep_time is not None:
                st.markdown(
                    f"Completed one polling in **{last_poll_time:.2f}s**, sleeping for **{last_sleep_time:.2f}s**."
                )

            if last_poll_data.get("cycle_completed"):
                cycle_duration = last_poll_data.get("cycle_duration", 0)
                st.markdown(
                    f"Waiting **{int(cycle_duration / 60)} minutes** before next reading cycle."
                )

        st.subheader("Realtime data monitoring by sensor type")
        st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'] = st.empty()

        assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

        for device_type_key in assigned_device_types_for_display:
            device_config = DEVICE_TYPE.get(device_type_key)
            if not device_config:
                continue

            has_slaves_assigned = any(
                assignment["device_type_key"] == device_type_key
                for assignment in SLAVE_DEVICE_ASSIGNMENTS
            )
            if not has_slaves_assigned:
                continue

            available_slaves_for_type = []
            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                if assignment["device_type_key"] == device_type_key:
                    if isinstance(assignment["ips"], range):
                        available_slaves_for_type.extend(list(assignment["ips"]))
                    elif isinstance(assignment["ips"], list):
                        available_slaves_for_type.extend(assignment["ips"])
            available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

            if not available_slaves_for_type:
                continue

            with st.expander(f"Configuration and Monitoring: **{device_type_key}**", expanded=True):
                selected = st.multiselect(
                    f"Select Slave IDs for {device_type_key}:",
                    options=available_slaves_for_type,
                    default=st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, []),
                    key=f"{state_prefix}multiselect_{device_type_key}"
                )

                if st.button(f"Confirm selection for {device_type_key}", key=f"{state_prefix}confirm_button_{device_type_key}"):
                    st.session_state[f'{state_prefix}selected_slaves_by_type'][device_type_key] = selected

                st.session_state[f'{state_prefix}realtime_table_placeholders'][device_type_key] = st.empty()

        def update_realtime_tables_by_type():
            current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

            st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'].write(
                f"Last widget update at **{current_ui_update_timestamp}**"
            )

            if st.session_state[f'{state_prefix}all_poll_data']:
                all_poll_data = st.session_state[f'{state_prefix}all_poll_data']
                latest_poll_entry = all_poll_data[-1]

                if latest_poll_entry.get("data"):
                    raw_data = latest_poll_entry['data']
                    st.session_state[f'{state_prefix}last_valid_poll_data'] = latest_poll_entry
                else:
                    raw_data = st.session_state.get(f'{state_prefix}last_valid_poll_data', {}).get('data', {})

                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    selected_slaves = st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, [])
                    device_config = DEVICE_TYPE.get(device_type_key)

                    placeholder.empty()

                    if not selected_slaves or not device_config:
                        placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                        continue

                    table_data = []
                    variable_names = [var[2] for var in device_config["variables"]]

                    for slave_id in sorted(selected_slaves):
                        if slave_id in raw_data:
                            values = raw_data[slave_id]
                            row_dict = {"Slave ID": f"Slave {slave_id}"}
                            for i, var_name in enumerate(variable_names):
                                if i < len(values) and values[i] is not None:
                                    display_value = values[i]
                                    if isinstance(display_value, float):
                                        display_value = round(display_value, 2)
                                else:
                                    display_value = np.nan
                                row_dict[var_name] = display_value
                            table_data.append(row_dict)
                        else:
                            table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                    if table_data:
                        df = pd.DataFrame(table_data)
                        df = df.set_index("Slave ID")
                        placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                    else:
                        placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
            else:
                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    placeholder.empty()
                    placeholder.info(f"Waiting for first polling data for {device_type_key}...")

        update_realtime_tables_by_type()

        if st.session_state[f'{state_prefix}polling']:
            st_autorefresh(interval=5000, key=f"realtime_autorefresh_{state_prefix}")

        if not os.path.exists(EXCEL_FILE_FINAL):
            if not st.session_state[f'{state_prefix}polling']:
                st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
        else:
            all_possible_slave_ips = sorted(list(set(SENSOR_IPS)))

            display_options = ["All"] + [ip for ip in all_possible_slave_ips]

            if "All" in st.session_state.get('selected_display_options', []):
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=["All"],
                    disabled=True
                )
            else:
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=[]
                )

            st.session_state.selected_display_options = selected_display_options

            selected_slave_ips_for_processing = []
            if "All" in selected_display_options:
                selected_slave_ips_for_processing = all_possible_slave_ips
            elif selected_display_options:
                selected_slave_ips_for_processing = selected_display_options

            if st.button("Generate Individual Files", key=f"generate_individual_files_{state_prefix}"):
                if not selected_slave_ips_for_processing:
                    st.warning("Please select at least one slave or 'All'.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    total_slaves_to_process = len(selected_slave_ips_for_processing)
                    zip_path = None

                    if not os.path.exists(EXCEL_FILE_BACKUP):
                        st.warning("Backup file not found. Performing only the split of the final Excel.")
                        with st.spinner("Splitting Excel and generating charts..."):
                            zip_path = split_and_generate_charts_for_ips(
                                EXCEL_FILE_FINAL,
                                selected_slave_ips_for_processing,
                                progress_bar=progress_bar,
                                status_text=status_text,
                                total_sensors=total_slaves_to_process
                            )
                    else:
                        st.info("Comparing timestamps between final Excel and backup...")
                        needs_merge = compare_timestamps_and_decide_action_ips(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ips_for_processing)

                        if needs_merge:
                            st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                            with st.spinner("Merging data from backup and final, and generating individual files..."):
                                zip_path = merge_data_and_generate_charts_for_ips(
                                    EXCEL_FILE_FINAL,
                                    EXCEL_FILE_BACKUP,
                                    selected_slave_ips_for_processing,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_sensors=total_slaves_to_process
                                )
                        else:
                            st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                            with st.spinner("Splitting Excel and generating charts..."):
                                zip_path = split_and_generate_charts_for_ips(
                                    EXCEL_FILE_FINAL,
                                    selected_slave_ips_for_processing,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_sensors=total_slaves_to_process
                                )

                    if zip_path:
                        status_text.text("Processing complete!")
                        st.success("Individual files generated and available for download.")



# -------------------------
# SCRIPT 2: M4M RTU
# -------------------------
def run_rtu():

    # ---------------------------------------- CONFIGURATION SECTION
    PORT_SERIAL = 'COM4'                        
    BAUDRATE = 19200                          
    PARITY = 'E'                       
    STOPBITS = 1                                       
    BYTESIZE = 8                                 
    TIMEOUT = 1
    SLAVE_IDS = [2, 9, 11]
    POLL_INTERVAL = 60
    EXCEL_FILE_FINAL = "M4M_RTU_modbus_poll_log.xlsx"
    EXCEL_FILE_BACKUP = "M4M_RTU_backup_poll_data.xlsx"
    state_prefix = "m4m_rtu_"
    
    DEVICE_TYPE = {
    "M4M": {
        "variables": [
            (20480, 4, "activeImportEnergyTotal", 3),
            (20484, 4, "activeExportEnergyTotal", 3),
            (20488, 4, "activeNetEnergyTotal", 3),
            (20492, 4, "reactiveImportEnergyTotal", 3),
            (20496, 4, "reactiveExportEnergyTotal", 3),
            (20500, 4, "reactiveNetEnergyTotal", 3),
            (20504, 4, "apparentImportEnergyTotal", 3),
            (20508, 4, "apparentExportEnergyTotal", 3),
            (20512, 4, "apparentNetEnergyTotal", 3),
            (21600, 4, "activeImpEnergyL1", 3),
            (21604, 4, "activeImpEnergyL2", 3),
            (21608, 4, "activeImpEnergyL3", 3),
            (21612, 4, "activeExpEnergyL1", 3),
            (21616, 4, "activeExpEnergyL2", 3),
            (21620, 4, "activeExpEnergyL3", 3),
            (21624, 4, "activeNetEnergyL1", 3),
            (21628, 4, "activeNetEnergyL2", 3),
            (21632, 4, "activeNetEnergyL3", 3),
            (21636, 4, "reactiveImpEnergyL1", 3),
            (21640, 4, "reactiveImpEnergyL2", 3),
            (21644, 4, "reactiveImpEnergyL3", 3),
            (21648, 4, "reactiveExpEnergyL1", 3),
            (21652, 4, "reactiveExpEnergyL2", 3),
            (21656, 4, "reactiveExpEnergyL3", 3),
            (21660, 4, "reactiveNetEnergyL1", 3),
            (21664, 4, "reactiveNetEnergyL2", 3),
            (21668, 4, "reactiveNetEnergyL3", 3),
            (21672, 4, "apparentImpEnergyL1", 3),
            (21676, 4, "apparentImpEnergyL2", 3),
            (21680, 4, "apparentImpEnergyL3", 3),
            (21684, 4, "apparentExpEnergyL1", 3),
            (21688, 4, "apparentExpEnergyL2", 3),
            (21692, 4, "apparentExpEnergyL3", 3),
            (21696, 4, "apparentNetEnergyL1", 3),
            (21700, 4, "apparentNetEnergyL2", 3),
            (21704, 4, "apparentNetEnergyL3", 3),
            (23312, 2, "currentL1", 3),
            (23314, 2, "currentL2", 3),
            (23316, 2, "currentL3", 3),
            (23318, 2, "currentN", 3),
            (23322, 2, "activePowerTotal", 3),
            (23324, 2, "activePowerL1", 3),
            (23326, 2, "activePowerL2", 3),
            (23328, 2, "activePowerL3", 3),
            (23330, 2, "reactivePowerTotal", 3),
            (23332, 2, "reactivePowerL1", 3),
            (23334, 2, "reactivePowerL2", 3),
            (23336, 2, "reactivePowerL3", 3),
            (23338, 2, "apparentPowerTotal", 3),
            (23340, 2, "apparentPowerL1", 3),
            (23342, 2, "apparentPowerL2", 3),
            (23344, 2, "apparentPowerL3", 3)
        ],
        
        "scale_factors": {
            "activeImportEnergyTotal": 0.01,
            "activeExportEnergyTotal": 0.01,
            "activeNetEnergyTotal": 0.01,
            "reactiveImportEnergyTotal": 0.01,
            "reactiveExportEnergyTotal": 0.01,
            "reactiveNetEnergyTotal": 0.01,
            "apparentImportEnergyTotal": 0.01,
            "apparentExportEnergyTotal": 0.01,
            "apparentNetEnergyTotal": 0.01,
            "activeImpEnergyL1": 0.01,
            "activeImpEnergyL2": 0.01,
            "activeImpEnergyL3": 0.01,
            "activeExpEnergyL1": 0.01,
            "activeExpEnergyL2": 0.01,
            "activeExpEnergyL3": 0.01,
            "activeNetEnergyL1": 0.01,
            "activeNetEnergyL2": 0.01,
            "activeNetEnergyL3": 0.01,
            "reactiveImpEnergyL1": 0.01,
            "reactiveImpEnergyL2": 0.01,
            "reactiveImpEnergyL3": 0.01,
            "reactiveExpEnergyL1": 0.01,
            "reactiveExpEnergyL2": 0.01,
            "reactiveExpEnergyL3": 0.01,
            "reactiveNetEnergyL1": 0.01,
            "reactiveNetEnergyL2": 0.01,
            "reactiveNetEnergyL3": 0.01,
            "apparentImpEnergyL1": 0.01,
            "apparentImpEnergyL2": 0.01,
            "apparentImpEnergyL3": 0.01,
            "apparentExpEnergyL1": 0.01,
            "apparentExpEnergyL2": 0.01,
            "apparentExpEnergyL3": 0.01,
            "apparentNetEnergyL1": 0.01,
            "apparentNetEnergyL2": 0.01,
            "apparentNetEnergyL3": 0.01,
            "currentL1": 0.01,
            "currentL2": 0.01,
            "currentL3": 0.01,
            "currentN": 0.01,
            "activePowerTotal": 0.01,
            "activePowerL1": 0.01,
            "activePowerL2": 0.01,
            "activePowerL3": 0.01,
            "reactivePowerTotal": 0.01,
            "reactivePowerL1": 0.01,
            "reactivePowerL2": 0.01,
            "reactivePowerL3": 0.01,
            "apparentPowerTotal": 0.01,
            "apparentPowerL1": 0.01,
            "apparentPowerL2": 0.01,
            "apparentPowerL3": 0.01
        },
        
        "decoding_map": {
            "activeImportEnergyTotal": "decode_64bit_unsigned",
            "activeExportEnergyTotal": "decode_64bit_unsigned",
            "activeNetEnergyTotal": "decode_64bit_signed",
            "reactiveImportEnergyTotal": "decode_64bit_unsigned",
            "reactiveExportEnergyTotal": "decode_64bit_unsigned",
            "reactiveNetEnergyTotal": "decode_64bit_signed",
            "apparentImportEnergyTotal": "decode_64bit_unsigned",
            "apparentExportEnergyTotal": "decode_64bit_unsigned",
            "apparentNetEnergyTotal": "decode_64bit_signed",
            "activeImpEnergyL1": "decode_64bit_unsigned",
            "activeImpEnergyL2": "decode_64bit_unsigned",
            "activeImpEnergyL3": "decode_64bit_unsigned",
            "activeExpEnergyL1": "decode_64bit_unsigned",
            "activeExpEnergyL2": "decode_64bit_unsigned",
            "activeExpEnergyL3": "decode_64bit_unsigned",
            "activeNetEnergyL1": "decode_64bit_signed",
            "activeNetEnergyL2": "decode_64bit_signed",
            "activeNetEnergyL3": "decode_64bit_signed",
            "reactiveImpEnergyL1": "decode_64bit_unsigned",
            "reactiveImpEnergyL2": "decode_64bit_unsigned",
            "reactiveImpEnergyL3": "decode_64bit_unsigned",
            "reactiveExpEnergyL1": "decode_64bit_unsigned",
            "reactiveExpEnergyL2": "decode_64bit_unsigned",
            "reactiveExpEnergyL3": "decode_64bit_unsigned",
            "reactiveNetEnergyL1": "decode_64bit_signed",
            "reactiveNetEnergyL2": "decode_64bit_signed",
            "reactiveNetEnergyL3": "decode_64bit_signed",
            "apparentImpEnergyL1": "decode_64bit_unsigned",
            "apparentImpEnergyL2": "decode_64bit_unsigned",
            "apparentImpEnergyL3": "decode_64bit_unsigned",
            "apparentExpEnergyL1": "decode_64bit_unsigned",
            "apparentExpEnergyL2": "decode_64bit_unsigned",
            "apparentExpEnergyL3": "decode_64bit_unsigned",
            "apparentNetEnergyL1": "decode_64bit_signed",
            "apparentNetEnergyL2": "decode_64bit_signed",
            "apparentNetEnergyL3": "decode_64bit_signed",
            "currentL1": "decode_32bit_unsigned",
            "currentL2": "decode_32bit_unsigned",
            "currentL3": "decode_32bit_unsigned",
            "currentN": "decode_32bit_unsigned",
            "activePowerTotal": "decode_32bit_signed",
            "activePowerL1": "decode_32bit_signed",
            "activePowerL2": "decode_32bit_signed",
            "activePowerL3": "decode_32bit_signed",
            "reactivePowerTotal": "decode_32bit_signed",
            "reactivePowerL1": "decode_32bit_signed",
            "reactivePowerL2": "decode_32bit_signed",
            "reactivePowerL3": "decode_32bit_signed",
            "apparentPowerTotal": "decode_32bit_signed",
            "apparentPowerL1": "decode_32bit_signed",
            "apparentPowerL2": "decode_32bit_signed",
            "apparentPowerL3": "decode_32bit_signed"
        },
        
        "plot_variables": [
            "activeImportEnergyTotal",
            "activeExportEnergyTotal",
            "activeNetEnergyTotal",
            "reactiveImportEnergyTotal",
            "reactiveExportEnergyTotal",
            "reactiveNetEnergyTotal",
            "apparentImportEnergyTotal",
            "apparentExportEnergyTotal",
            "apparentNetEnergyTotal",
            "activeImpEnergyL1",
            "activeImpEnergyL2",
            "activeImpEnergyL3",
            "activeExpEnergyL1",
            "activeExpEnergyL2",
            "activeExpEnergyL3",
            "activeNetEnergyL1",
            "activeNetEnergyL2",
            "activeNetEnergyL3",
            "reactiveImpEnergyL1",
            "reactiveImpEnergyL2",
            "reactiveImpEnergyL3",
            "reactiveExpEnergyL1",
            "reactiveExpEnergyL2",
            "reactiveExpEnergyL3",
            "reactiveNetEnergyL1",
            "reactiveNetEnergyL2",
            "reactiveNetEnergyL3",
            "apparentImpEnergyL1",
            "apparentImpEnergyL2",
            "apparentImpEnergyL3",
            "apparentExpEnergyL1",
            "apparentExpEnergyL2",
            "apparentExpEnergyL3",
            "apparentNetEnergyL1",
            "apparentNetEnergyL2",
            "apparentNetEnergyL3",
            "currentL1",
            "currentL2",
            "currentL3",
            "currentN",
            "activePowerTotal",
            "activePowerL1",
            "activePowerL2",
            "activePowerL3",
            "reactivePowerTotal",
            "reactivePowerL1",
            "reactivePowerL2",
            "reactivePowerL3",
            "apparentPowerTotal",
            "apparentPowerL1",
            "apparentPowerL2",
            "apparentPowerL3"
        ],
        
        "variable_units": {
            "activeImportEnergyTotal": "(kWh)",
            "activeExportEnergyTotal": "(kWh)",
            "activeNetEnergyTotal": "(kWh)",
            "reactiveImportEnergyTotal": "(kVarh)",
            "reactiveExportEnergyTotal": "(kVarh)",
            "reactiveNetEnergyTotal": "(kVarh)",
            "apparentImportEnergyTotal": "(kVAh)",
            "apparentExportEnergyTotal": "(kVAh)",
            "apparentNetEnergyTotal": "(kVAh)",
            "activeImpEnergyL1": "(kWh)",
            "activeImpEnergyL2": "(kWh)",
            "activeImpEnergyL3": "(kWh)",
            "activeExpEnergyL1": "(kWh)",
            "activeExpEnergyL2": "(kWh)",
            "activeExpEnergyL3": "(kWh)",
            "activeNetEnergyL1": "(kWh)",
            "activeNetEnergyL2": "(kWh)",
            "activeNetEnergyL3": "(kWh)",
            "reactiveImpEnergyL1": "(kVarh)",
            "reactiveImpEnergyL2": "(kVarh)",
            "reactiveImpEnergyL3": "(kVarh)",
            "reactiveExpEnergyL1": "(kVarh)",
            "reactiveExpEnergyL2": "(kVarh)",
            "reactiveExpEnergyL3": "(kVarh)",
            "reactiveNetEnergyL1": "(kVarh)",
            "reactiveNetEnergyL2": "(kVarh)",
            "reactiveNetEnergyL3": "(kVarh)",
            "apparentImpEnergyL1": "(kVAh)",
            "apparentImpEnergyL2": "(kVAh)",
            "apparentImpEnergyL3": "(kVAh)",
            "apparentExpEnergyL1": "(kVAh)",
            "apparentExpEnergyL2": "(kVAh)",
            "apparentExpEnergyL3": "(kVAh)",
            "apparentNetEnergyL1": "(kVAh)",
            "apparentNetEnergyL2": "(kVAh)",
            "apparentNetEnergyL3": "(kVAh)",
            "currentL1": "(A)",
            "currentL2": "(A)",
            "currentL3": "(A)",
            "currentN": "(A)",
            "activePowerTotal": "(W)",
            "activePowerL1": "(W)",
            "activePowerL2": "(W)",
            "activePowerL3": "(W)",
            "reactivePowerTotal": "(var)",
            "reactivePowerL1": "(var)",
            "reactivePowerL2": "(var)",
            "reactivePowerL3": "(var)",
            "apparentPowerTotal": "(VA)",
            "apparentPowerL1": "(VA)",
            "apparentPowerL2": "(VA)",
            "apparentPowerL3": "(VA)"
        }
        }
    }

    SLAVE_DEVICE_ASSIGNMENTS = [{"slaves": SLAVE_IDS, "device_type_key": "M4M"}]

    if SLAVE_DEVICE_ASSIGNMENTS:
        first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
        if first_assigned_device_type_key in DEVICE_TYPE:
            VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
        else:
            VARIABLES = []
    else:
        VARIABLES = []



    # --------------------------------------------- HELPER FUNCTIONS
    def get_device_config_for_slave(slave_id):
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if isinstance(assignment["slaves"], range):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
            elif isinstance(assignment["slaves"], list):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
        else:
            print(f"Warning: Slave ID {slave_id} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping this slave.")
            return None

        if device_type_key not in DEVICE_TYPE:
            print(f"Error: Device type key '{device_type_key}' assigned to slave {slave_id} not found in DEVICE_TYPE configuration.")
            return None
            
        return DEVICE_TYPE[device_type_key]



    # -------------------------------------- SENSOR POLLING FUNCTION 
    def read_slave_with_client(slave_id, thread_id, client):
        device_config = get_device_config_for_slave(slave_id)
        if device_config is None:
            dummy_len = len(DEVICE_TYPE.get("M4M", {}).get("variables", []))
            return [None] * dummy_len

        variables = device_config["variables"]
        scale_factors = device_config["scale_factors"]
        decoding_map = device_config["decoding_map"]

        values = []
        thread_prefix = f"[Thread {thread_id}] "

        for var_info in variables:
            addr, size, name, func_code = var_info[:4]
            options = var_info[4] if len(var_info) > 4 else {}

            current_value = None
            try:
                resp = None            
                if func_code == 1:
                    resp = client.read_coils(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Coils read error or empty bits.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.bits
                elif func_code == 3:
                    resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Registers read error or empty registers.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.registers
                else:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Unsupported function code {func_code}.")
                    values.append(None)
                    continue

                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

                if name in decoding_map:
                    func_name = decoding_map[name]
                    decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)
                    
                    if decoder_func:
                        if func_name == "decode_bit":
                            bit_pos = options.get('bit')
                            if bit_pos is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_pos)
                            else:
                                print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                        elif func_name == "decode_bits":
                            bit_range = options.get('bits')
                            if bit_range is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_range)
                            else:
                                print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                        elif func_name == "read_coil":
                            current_value = decoder_func(resp)
                        elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                            current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                        else:
                            current_value = decoder_func(regs_or_bits)
                    else:
                        print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Falling back to raw value processing.")
                        current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
                else:
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

                if name in scale_factors and isinstance(current_value, (int, float)):
                    current_value *= scale_factors[name]
                
                if isinstance(current_value, list):
                    final_value = flatten_value(current_value)
                else:
                    final_value = current_value

                values.append(final_value)

            except Exception as e:
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
                values.append(None)

        return values



    # ------------------------------------------------ SAVE TO EXCEL
    def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
        wb = Workbook()
        wb.remove(wb.active)
        
        for slave_id in SLAVE_IDS:
            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping sheet creation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            ws = wb.create_sheet(title=f"Slave_{slave_id}")
            headers = ['timestamp'] + [var[2] for var in device_config["variables"]]
            ws.append(headers)

        for poll_entry in all_poll_data:
            timestamp = poll_entry['timestamp']
            data = poll_entry['data']
            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

        wb.save(filename)

    def append_poll_data_to_backup(poll_entry, filename=EXCEL_FILE_BACKUP):
        from openpyxl import load_workbook
        from openpyxl import Workbook
        import os

        timestamp = poll_entry['timestamp']
        data = poll_entry['data']

        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        for slave_id, values in data.items():
            sheet_name = f"Slave_{slave_id}"

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                device_config = get_device_config_for_slave(slave_id)
                if device_config is None:
                    print(f"Skipping backup append for Slave_{slave_id} due to missing device configuration.")
                    continue
                headers = ['timestamp'] + [var[2] for var in device_config["variables"]]
                ws.append(headers)
            else:
                ws = wb[sheet_name]

            existing_timestamps = {row[0].value for row in ws.iter_rows(min_row=2, max_col=1)}
            if timestamp in existing_timestamps:
                continue

            safe_values = [clean_value(v) for v in values]
            ws.append([timestamp] + safe_values)

        wb.save(filename)

    def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
        device_config = get_device_config_for_slave(slave_id)
        if device_config is None:
            print(f"Warning: No device configuration found for slave {slave_id}. Cannot create chart for {var_name}.")
            return

        variable_units = device_config.get("variable_units", {})
        unit = variable_units.get(var_name, "")

        chart = ScatterChart()
        chart.title = f"{var_name} over time - ID {slave_id}"
        chart.style = 13
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = f"{var_name} {unit}".strip()
        chart.legend = None

        chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.x_axis.majorUnit = 500

        x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
        y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

        title_with_unit = f"{var_name} {unit}".strip()
        series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.smooth = True
        chart.series.append(series)

        chart.x_axis.axId = 10 + chart_index * 2
        chart.y_axis.axId = 20 + chart_index * 2
        chart.y_axis.crossAx = 10 + chart_index * 2
        chart.x_axis.crossAx = 20 + chart_index * 2

        chart.height = 9.5
        chart.width = 20.44

        anchor_col = 'BC'
        anchor_row = 1 + 20 * chart_index
        chart.anchor = f"{anchor_col}{anchor_row}"

        sheet.add_chart(chart)

    def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
        wb = load_workbook(filename)

        for slave_id in SLAVE_IDS:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            num_rows = ws.max_row
            
            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            variables_to_plot_for_slave = device_config["plot_variables"]

            chart_index = 0
            timestamp_col = 1

            for var in variables_to_plot_for_slave:
                if var not in headers:
                    continue

                var_col = headers.index(var) + 1
                create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
                chart_index += 1

        wb.save(filename)



    # ------------------------------------------------- POLLING LOOP
    def poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix):
        if stop_event.is_set():
            return

        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()
        threads = []

        slave_ids = [2, 9, 11]

        def poll_slave_list(slave_ids, all_data, lock, thread_id):
            client = ModbusSerialClient(
                port=PORT_SERIAL,
                baudrate=BAUDRATE,
                parity=PARITY,
                stopbits=STOPBITS,
                bytesize=BYTESIZE,
                timeout=TIMEOUT
            )
            if not client.connect():
                print(f"[Thread {thread_id}] Failed to connect to RTU port")
                return

            def task(slave_id):
                values = read_slave_with_client(slave_id, thread_id, client)
                with lock:
                    all_data[slave_id] = values

            with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
                executor.map(task, slave_ids)

            client.close()

        t = threading.Thread(target=poll_slave_list, args=(slave_ids, all_data, lock, 1))
        threads.append(t)
        t.start()

        for t in threads:
            t.join()

        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        all_poll_data.append({
            'timestamp': timestamp,
            'data': all_data,
        })

        print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

        st.session_state[f'{state_prefix}last_valid_poll_data'] = {
            'timestamp': timestamp,
            'data': all_data
        }

        try:
            append_poll_data_to_backup({
                'timestamp': timestamp,
                'data': all_data
            })
        except Exception as e:
            print(f"Error saving to backup file: {e}")

    def periodic_polling_loop(stop_event, all_poll_data, state_prefix):
        while not stop_event.is_set():
            start_time = time.time()
            one_cycle = 600
            cycle_duration = 60 * 60

            while time.time() - start_time < one_cycle and not stop_event.is_set():
                poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix)

            if stop_event.is_set():
                break

            all_poll_data.append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': {},
            })

            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = True
            print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
            
            for _ in range(int(cycle_duration)):
                if stop_event.is_set():
                    break
                time.sleep(1)
            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = False



    # ---------------------------------------- SPLIT EXCEL FUNCTIONS
    def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_file_excel):
            st.error(f"Error: File '{path_file_excel}' not found.")
            return None

        try:
            if not os.path.exists(name_output_folder):
                os.makedirs(name_output_folder)
            else:
                for file_name in os.listdir(name_output_folder):
                    path_old_file = os.path.join(name_output_folder, file_name)
                    if os.path.isfile(path_old_file):
                        os.remove(path_old_file)
                st.info(f"Cleaning folder '{name_output_folder}'.")

            source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
            generated_files = []

            for i, slave_id in enumerate(selected_slave_ids):
                sheet_name = f"Slave_{slave_id}"
                if status_text:
                    status_text.text(f"Processing Slave_{slave_id}...")
                if sheet_name not in source_workbook.sheetnames:
                    st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                    continue

                sheet = source_workbook[sheet_name]

                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in sheet.iter_rows(values_only=True):
                    new_ws.append(row)

                num_rows = new_ws.max_row
                num_cols = new_ws.max_column

                if num_rows < 2 or num_cols < 2:
                    st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                    continue

                timestamp_col = 1
                
                device_config = get_device_config_for_slave(slave_id)
                if device_config is None:
                    print(f"Skipping chart generation for Slave_{slave_id} in split function due to missing device configuration.")
                    continue

                variables_to_plot_for_slave = device_config["plot_variables"]

                for idx, var_name in enumerate(variables_to_plot_for_slave):
                    if var_name in [cell.value for cell in new_ws[1]]:
                        var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                        create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                    else:
                        st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

                output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
                output_filepath = os.path.join(name_output_folder, output_filename)

                new_wb.save(output_filepath)
                generated_files.append(output_filepath)
                
                if progress_bar and total_slaves > 0:
                    progress_percent = int(((i + 1) / total_slaves) * 100)
                    progress_bar.progress(progress_percent)

            if not generated_files:
                st.warning("No individual files were generated for the selected slaves.")
                return None

            name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
            complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

            with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_to_add in generated_files:
                    zipf.write(file_to_add, os.path.basename(file_to_add))

            st.success(f"All files were zipped into '{complete_zip_path}'")
            return complete_zip_path

        except Exception as e:
            st.error(f"Error splitting and generating charts: {e}")
            return None

    def extract_slave_id(sheet_name):
        match = re.search(r'\d+', sheet_name)
        return int(match.group(0)) if match else None

    def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
            st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
            return None

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        else:
            for file_name in os.listdir(output_folder):
                path_to_remove = os.path.join(output_folder, file_name)
                if os.path.isfile(path_to_remove):
                    os.remove(path_to_remove)
            st.info(f"Folder '{output_folder}' cleaned.")

        wb_modbus = openpyxl.load_workbook(path_modbus_log)
        wb_backup = openpyxl.load_workbook(path_backup_data)
        
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in wb_modbus.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue
            if sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
                continue

            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping merge/chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

            modbus_sheet = wb_modbus[sheet_name]
            backup_sheet = wb_backup[sheet_name]

            modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
            backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

            modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
            backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

            merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
            merged_data.dropna(subset=["timestamp"], inplace=True)
            merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
            merged_data.sort_values("timestamp", inplace=True)
            merged_data.reset_index(drop=True, inplace=True)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            new_sheet = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in modbus_sheet[1]]
            new_sheet.append(headers)

            for r_idx, row in merged_data.iterrows():
                new_sheet.append(list(row.values))

            num_rows = new_sheet.max_row
            current_slave_id = extract_slave_id(sheet_name)
            
            variables_to_plot_for_slave = device_config["plot_variables"]

            for chart_index, var_name in enumerate(variables_to_plot_for_slave):
                if var_name in headers:
                    col_idx = headers.index(var_name) + 1
                    create_scatter_chart(
                        new_sheet,
                        timestamp_col=1,
                        var_col=col_idx,
                        num_rows=num_rows,
                        var_name=var_name,
                        slave_id=current_slave_id,
                        chart_index=chart_index
                    )
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
            new_wb.save(path_output)
            generated_files.append(path_output)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All merged individual files have been zipped into: {zip_path}")
        return zip_path

    def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
        if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
            st.error("Excel files (final or backup) not found for comparison.")
            return False
        
        wb_final = load_workbook(excel_file_final, data_only=True)
        wb_backup = load_workbook(excel_file_backup, data_only=True)

        needs_merge = False
        for slave_id in selected_slave_ids:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
                continue

            ws_final = wb_final[sheet_name]
            ws_backup = wb_backup[sheet_name]

            timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
            timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

            if not timestamps_backup.issubset(timestamps_final):
                needs_merge = True
                break
        
        return needs_merge



    # ------------------------------------------ STREAMLIT DASHBOARD
    with st.container():
        st.title("M4M Modbus Serial RTU Setup")

        if f'{state_prefix}polling' not in st.session_state:
            st.session_state[f'{state_prefix}last_valid_poll_data'] = {}
            st.session_state[f'{state_prefix}polling'] = False
            st.session_state[f'{state_prefix}stop_event'] = threading.Event()
            st.session_state[f'{state_prefix}thread'] = None
            st.session_state[f'{state_prefix}all_poll_data'] = []
            
            st.session_state[f'{state_prefix}selected_slaves_by_type'] = {
                device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
            }
            st.session_state[f'{state_prefix}realtime_table_placeholders'] = {}
            st.session_state[f'{state_prefix}last_data_length'] = 0
            st.session_state[f'{state_prefix}last_ui_update_time'] = time.time()
            st.session_state[f'{state_prefix}last_table_update_timestamp'] = "N/A"

        if st.button("Start Monitoring", key=f"start_monitoring_{state_prefix}", disabled=st.session_state[f'{state_prefix}polling']):
            if not st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}all_poll_data'] = []
                st.session_state[f'{state_prefix}stop_event'].clear()

                st.session_state[f'{state_prefix}thread'] = threading.Thread(
                    target=periodic_polling_loop,
                    args=(
                        st.session_state[f'{state_prefix}stop_event'],
                        st.session_state[f'{state_prefix}all_poll_data'],
                        state_prefix
                    ),
                    daemon=True
                )

                st.session_state[f'{state_prefix}thread'].start()
                st.session_state[f'{state_prefix}polling'] = True
                st.session_state[f'{state_prefix}last_data_length'] = 0
                st.success("Monitoring started!")

        if st.button("Stop Monitoring", key=f"stop_monitoring_{state_prefix}"):
            if st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}stop_event'].set()
                if st.session_state[f'{state_prefix}thread'] and st.session_state[f'{state_prefix}thread'].is_alive():
                    st.session_state[f'{state_prefix}thread'].join(timeout=10)
                    if st.session_state[f'{state_prefix}thread'].is_alive():
                        st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
                    else:
                        st.info("Polling thread terminated.")
                st.session_state[f'{state_prefix}polling'] = False

                with st.spinner("Saving data to Excel..."):
                    save_all_data_to_excel(st.session_state[f'{state_prefix}all_poll_data'], filename=EXCEL_FILE_FINAL)

                    try:
                        add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                        st.success("Monitoring stopped, data and charts saved to Excel.")
                    except Exception as e:
                        st.error(f"Error adding charts: {e}")
            else:
                st.warning("Monitoring is not running.")

        st.divider()

        if st.session_state.get(f"{state_prefix}all_poll_data"):
            all_poll_data = st.session_state[f"{state_prefix}all_poll_data"]

            st.markdown(
                f"Total of **{int(len(all_poll_data))}** reading(s) completed so far."
            )

            device_counts = {}

            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                key = assignment["device_type_key"]
                normalized_key = re.match(r'^[A-Za-z]+', key)
                normalized_key = normalized_key.group(0) if normalized_key else key
                if "ips" in assignment:
                    count = len(assignment["ips"])
                elif "slaves" in assignment:
                    count = len(assignment["slaves"])
                else:
                    count = 0
                device_counts[key] = device_counts.get(key, 0) + count

            device_count_lines = [f"- **{k}**: {v} device(s)" for k, v in device_counts.items()]
            st.markdown("### Active devices on the system:")
            st.markdown("\n".join(device_count_lines))

            last_poll_data = all_poll_data[-1]
            if "elapsed" not in last_poll_data and len(all_poll_data) >= 2:
                last_poll_data = all_poll_data[-2]

            last_poll_time = last_poll_data.get("elapsed", None)
            last_sleep_time = last_poll_data.get("sleep_time", None)

            if last_poll_time is not None and last_sleep_time is not None:
                st.markdown(
                    f"Completed one polling in **{last_poll_time:.2f}s**, sleeping for **{last_sleep_time:.2f}s**."
                )

            if last_poll_data.get("cycle_completed"):
                cycle_duration = last_poll_data.get("cycle_duration", 0)
                st.markdown(
                    f"Waiting **{int(cycle_duration / 60)} minutes** before next reading cycle."
                )

        st.subheader("Realtime data monitoring by sensor type")

        st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'] = st.empty()

        assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

        for device_type_key in assigned_device_types_for_display:
            device_config = DEVICE_TYPE.get(device_type_key)
            if not device_config:
                continue

            has_slaves_assigned = any(
                assignment["device_type_key"] == device_type_key
                for assignment in SLAVE_DEVICE_ASSIGNMENTS
            )
            if not has_slaves_assigned:
                continue

            available_slaves_for_type = []
            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                if assignment["device_type_key"] == device_type_key:
                    if isinstance(assignment["slaves"], range):
                        available_slaves_for_type.extend(list(assignment["slaves"]))
                    elif isinstance(assignment["slaves"], list):
                        available_slaves_for_type.extend(assignment["slaves"])
            available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

            if not available_slaves_for_type:
                continue

            with st.expander(f"Configuration and Monitoring: **{device_type_key}**", expanded=True):
                selected = st.multiselect(
                    f"Select Slave IDs for {device_type_key}:",
                    options=available_slaves_for_type,
                    default=st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, []),
                    key=f"{state_prefix}multiselect_{device_type_key}"
                )

                if st.button(f"Confirm selection for {device_type_key}", key=f"{state_prefix}confirm_button_{device_type_key}"):
                    st.session_state[f'{state_prefix}selected_slaves_by_type'][device_type_key] = selected

                st.session_state[f'{state_prefix}realtime_table_placeholders'][device_type_key] = st.empty()

        def update_realtime_tables_by_type():
            current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

            st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'].write(
                f"Last widget update at **{current_ui_update_timestamp}**"
            )

            if st.session_state[f'{state_prefix}all_poll_data']:
                all_poll_data = st.session_state[f'{state_prefix}all_poll_data']
                latest_poll_entry = all_poll_data[-1]

                if latest_poll_entry.get("data"):
                    raw_data = latest_poll_entry['data']
                    st.session_state[f'{state_prefix}last_valid_poll_data'] = latest_poll_entry
                else:
                    raw_data = st.session_state.get(f'{state_prefix}last_valid_poll_data', {}).get('data', {})

                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    selected_slaves = st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, [])
                    device_config = DEVICE_TYPE.get(device_type_key)

                    placeholder.empty()

                    if not selected_slaves or not device_config:
                        placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                        continue

                    table_data = []
                    variable_names = [var[2] for var in device_config["variables"]]

                    for slave_id in sorted(selected_slaves):
                        if slave_id in raw_data:
                            values = raw_data[slave_id]
                            row_dict = {"Slave ID": f"Slave {slave_id}"}
                            for i, var_name in enumerate(variable_names):
                                if i < len(values) and values[i] is not None:
                                    display_value = values[i]
                                    if isinstance(display_value, float):
                                        display_value = round(display_value, 2)
                                else:
                                    display_value = np.nan
                                row_dict[var_name] = display_value
                            table_data.append(row_dict)
                        else:
                            table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                    if table_data:
                        df = pd.DataFrame(table_data)
                        df = df.set_index("Slave ID")

                        placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                    else:
                        placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
            else:
                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    placeholder.empty()
                    placeholder.info(f"Waiting for first polling data for {device_type_key}...")

        update_realtime_tables_by_type()

        if st.session_state[f'{state_prefix}polling']:
            st_autorefresh(interval=5000, key=f"realtime_autorefresh_{state_prefix}")

        if not os.path.exists(EXCEL_FILE_FINAL):
            if not st.session_state[f'{state_prefix}polling']:
                st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
        else:
            st.divider()
            all_possible_slave_ids = sorted(list(set(SLAVE_IDS)))

            display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

            if "All" in st.session_state.get('selected_display_options', []):
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=["All"],
                    disabled=True
                )
            else:
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=[]
                )

            st.session_state.selected_display_options = selected_display_options

            selected_slave_ids_numeric = []
            if "All" in selected_display_options:
                selected_slave_ids_numeric = all_possible_slave_ids
            elif selected_display_options:
                for ds_id in selected_display_options:
                    match = re.search(r'\d+', ds_id)
                    if match:
                        selected_slave_ids_numeric.append(int(match.group(0)))

            if st.button("Generate Individual Files", key=f"generate_individual_files_{state_prefix}"):
                if not selected_slave_ids_numeric:
                    st.warning("Please select at least one slave or 'All'.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    total_slaves_to_process = len(selected_slave_ids_numeric)
                    zip_path = None

                    if not os.path.exists(EXCEL_FILE_BACKUP):
                        st.warning("Backup file not found. Performing only the split of the final Excel.")
                        with st.spinner("Splitting Excel and generating charts..."):
                            zip_path = split_and_generate_charts_for_slaves(
                                EXCEL_FILE_FINAL,
                                selected_slave_ids_numeric,
                                progress_bar=progress_bar,
                                status_text=status_text,
                                total_slaves=total_slaves_to_process
                            )
                    else:
                        st.info("Comparing timestamps between final Excel and backup...")
                        needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                        if needs_merge:
                            st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                            with st.spinner("Merging data from backup and final, and generating individual files..."):
                                zip_path = merge_data_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    EXCEL_FILE_BACKUP,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )
                        else:
                            st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                            with st.spinner("Splitting Excel and generating charts..."):
                                zip_path = split_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )

                    if zip_path:
                        status_text.text("Processing complete!")
                        st.success("Individual files generated and available for download.")



# -------------------------
# SCRIPT 3: GATEWAY
# -------------------------
def run_gateway():
    
    # ---------------------------------------- CONFIGURATION SECTION
    IP = '192.168.1.50'
    PORT = 502
    SLAVE_IDS = [4, 5, 15]
    POLL_INTERVAL = 60
    EXCEL_FILE_FINAL = "EdgeIndustrialGateway_modbus_poll_log.xlsx"
    EXCEL_FILE_BACKUP = "EdgeIndustrialGateway_backup_poll_data.xlsx"
    state_prefix = "gateway_"

    DEVICE_TYPE = {
        "M4M": {
            "variables": [
                (20480, 4, "activeImportEnergyTotal", 3),
                (20484, 4, "activeExportEnergyTotal", 3),
                (20488, 4, "activeNetEnergyTotal", 3),
                (20492, 4, "reactiveImportEnergyTotal", 3),
                (20496, 4, "reactiveExportEnergyTotal", 3),
                (20500, 4, "reactiveNetEnergyTotal", 3),
                (20504, 4, "apparentImportEnergyTotal", 3),
                (20508, 4, "apparentExportEnergyTotal", 3),
                (20512, 4, "apparentNetEnergyTotal", 3),
                (21600, 4, "activeImpEnergyL1", 3),
                (21604, 4, "activeImpEnergyL2", 3),
                (21608, 4, "activeImpEnergyL3", 3),
                (21612, 4, "activeExpEnergyL1", 3),
                (21616, 4, "activeExpEnergyL2", 3),
                (21620, 4, "activeExpEnergyL3", 3),
                (21624, 4, "activeNetEnergyL1", 3),
                (21628, 4, "activeNetEnergyL2", 3),
                (21632, 4, "activeNetEnergyL3", 3),
                (21636, 4, "reactiveImpEnergyL1", 3),
                (21640, 4, "reactiveImpEnergyL2", 3),
                (21644, 4, "reactiveImpEnergyL3", 3),
                (21648, 4, "reactiveExpEnergyL1", 3),
                (21652, 4, "reactiveExpEnergyL2", 3),
                (21656, 4, "reactiveExpEnergyL3", 3),
                (21660, 4, "reactiveNetEnergyL1", 3),
                (21664, 4, "reactiveNetEnergyL2", 3),
                (21668, 4, "reactiveNetEnergyL3", 3),
                (21672, 4, "apparentImpEnergyL1", 3),
                (21676, 4, "apparentImpEnergyL2", 3),
                (21680, 4, "apparentImpEnergyL3", 3),
                (21684, 4, "apparentExpEnergyL1", 3),
                (21688, 4, "apparentExpEnergyL2", 3),
                (21692, 4, "apparentExpEnergyL3", 3),
                (21696, 4, "apparentNetEnergyL1", 3),
                (21700, 4, "apparentNetEnergyL2", 3),
                (21704, 4, "apparentNetEnergyL3", 3),
                (23312, 2, "currentL1", 3),
                (23314, 2, "currentL2", 3),
                (23316, 2, "currentL3", 3),
                (23318, 2, "currentN", 3),
                (23322, 2, "activePowerTotal", 3),
                (23324, 2, "activePowerL1", 3),
                (23326, 2, "activePowerL2", 3),
                (23328, 2, "activePowerL3", 3),
                (23330, 2, "reactivePowerTotal", 3),
                (23332, 2, "reactivePowerL1", 3),
                (23334, 2, "reactivePowerL2", 3),
                (23336, 2, "reactivePowerL3", 3),
                (23338, 2, "apparentPowerTotal", 3),
                (23340, 2, "apparentPowerL1", 3),
                (23342, 2, "apparentPowerL2", 3),
                (23344, 2, "apparentPowerL3", 3)
            ],
            
            "scale_factors": {
                "activeImportEnergyTotal": 0.01,
                "activeExportEnergyTotal": 0.01,
                "activeNetEnergyTotal": 0.01,
                "reactiveImportEnergyTotal": 0.01,
                "reactiveExportEnergyTotal": 0.01,
                "reactiveNetEnergyTotal": 0.01,
                "apparentImportEnergyTotal": 0.01,
                "apparentExportEnergyTotal": 0.01,
                "apparentNetEnergyTotal": 0.01,
                "activeImpEnergyL1": 0.01,
                "activeImpEnergyL2": 0.01,
                "activeImpEnergyL3": 0.01,
                "activeExpEnergyL1": 0.01,
                "activeExpEnergyL2": 0.01,
                "activeExpEnergyL3": 0.01,
                "activeNetEnergyL1": 0.01,
                "activeNetEnergyL2": 0.01,
                "activeNetEnergyL3": 0.01,
                "reactiveImpEnergyL1": 0.01,
                "reactiveImpEnergyL2": 0.01,
                "reactiveImpEnergyL3": 0.01,
                "reactiveExpEnergyL1": 0.01,
                "reactiveExpEnergyL2": 0.01,
                "reactiveExpEnergyL3": 0.01,
                "reactiveNetEnergyL1": 0.01,
                "reactiveNetEnergyL2": 0.01,
                "reactiveNetEnergyL3": 0.01,
                "apparentImpEnergyL1": 0.01,
                "apparentImpEnergyL2": 0.01,
                "apparentImpEnergyL3": 0.01,
                "apparentExpEnergyL1": 0.01,
                "apparentExpEnergyL2": 0.01,
                "apparentExpEnergyL3": 0.01,
                "apparentNetEnergyL1": 0.01,
                "apparentNetEnergyL2": 0.01,
                "apparentNetEnergyL3": 0.01,
                "currentL1": 0.01,
                "currentL2": 0.01,
                "currentL3": 0.01,
                "currentN": 0.01,
                "activePowerTotal": 0.01,
                "activePowerL1": 0.01,
                "activePowerL2": 0.01,
                "activePowerL3": 0.01,
                "reactivePowerTotal": 0.01,
                "reactivePowerL1": 0.01,
                "reactivePowerL2": 0.01,
                "reactivePowerL3": 0.01,
                "apparentPowerTotal": 0.01,
                "apparentPowerL1": 0.01,
                "apparentPowerL2": 0.01,
                "apparentPowerL3": 0.01
            },
            
            "decoding_map": {
                "activeImportEnergyTotal": "decode_64bit_unsigned",
                "activeExportEnergyTotal": "decode_64bit_unsigned",
                "activeNetEnergyTotal": "decode_64bit_signed",
                "reactiveImportEnergyTotal": "decode_64bit_unsigned",
                "reactiveExportEnergyTotal": "decode_64bit_unsigned",
                "reactiveNetEnergyTotal": "decode_64bit_signed",
                "apparentImportEnergyTotal": "decode_64bit_unsigned",
                "apparentExportEnergyTotal": "decode_64bit_unsigned",
                "apparentNetEnergyTotal": "decode_64bit_signed",
                "activeImpEnergyL1": "decode_64bit_unsigned",
                "activeImpEnergyL2": "decode_64bit_unsigned",
                "activeImpEnergyL3": "decode_64bit_unsigned",
                "activeExpEnergyL1": "decode_64bit_unsigned",
                "activeExpEnergyL2": "decode_64bit_unsigned",
                "activeExpEnergyL3": "decode_64bit_unsigned",
                "activeNetEnergyL1": "decode_64bit_signed",
                "activeNetEnergyL2": "decode_64bit_signed",
                "activeNetEnergyL3": "decode_64bit_signed",
                "reactiveImpEnergyL1": "decode_64bit_unsigned",
                "reactiveImpEnergyL2": "decode_64bit_unsigned",
                "reactiveImpEnergyL3": "decode_64bit_unsigned",
                "reactiveExpEnergyL1": "decode_64bit_unsigned",
                "reactiveExpEnergyL2": "decode_64bit_unsigned",
                "reactiveExpEnergyL3": "decode_64bit_unsigned",
                "reactiveNetEnergyL1": "decode_64bit_signed",
                "reactiveNetEnergyL2": "decode_64bit_signed",
                "reactiveNetEnergyL3": "decode_64bit_signed",
                "apparentImpEnergyL1": "decode_64bit_unsigned",
                "apparentImpEnergyL2": "decode_64bit_unsigned",
                "apparentImpEnergyL3": "decode_64bit_unsigned",
                "apparentExpEnergyL1": "decode_64bit_unsigned",
                "apparentExpEnergyL2": "decode_64bit_unsigned",
                "apparentExpEnergyL3": "decode_64bit_unsigned",
                "apparentNetEnergyL1": "decode_64bit_signed",
                "apparentNetEnergyL2": "decode_64bit_signed",
                "apparentNetEnergyL3": "decode_64bit_signed",
                "currentL1": "decode_32bit_unsigned",
                "currentL2": "decode_32bit_unsigned",
                "currentL3": "decode_32bit_unsigned",
                "currentN": "decode_32bit_unsigned",
                "activePowerTotal": "decode_32bit_signed",
                "activePowerL1": "decode_32bit_signed",
                "activePowerL2": "decode_32bit_signed",
                "activePowerL3": "decode_32bit_signed",
                "reactivePowerTotal": "decode_32bit_signed",
                "reactivePowerL1": "decode_32bit_signed",
                "reactivePowerL2": "decode_32bit_signed",
                "reactivePowerL3": "decode_32bit_signed",
                "apparentPowerTotal": "decode_32bit_signed",
                "apparentPowerL1": "decode_32bit_signed",
                "apparentPowerL2": "decode_32bit_signed",
                "apparentPowerL3": "decode_32bit_signed"
            },
            
            "plot_variables": [
                "activeImportEnergyTotal",
                "activeExportEnergyTotal",
                "activeNetEnergyTotal",
                "reactiveImportEnergyTotal",
                "reactiveExportEnergyTotal",
                "reactiveNetEnergyTotal",
                "apparentImportEnergyTotal",
                "apparentExportEnergyTotal",
                "apparentNetEnergyTotal",
                "activeImpEnergyL1",
                "activeImpEnergyL2",
                "activeImpEnergyL3",
                "activeExpEnergyL1",
                "activeExpEnergyL2",
                "activeExpEnergyL3",
                "activeNetEnergyL1",
                "activeNetEnergyL2",
                "activeNetEnergyL3",
                "reactiveImpEnergyL1",
                "reactiveImpEnergyL2",
                "reactiveImpEnergyL3",
                "reactiveExpEnergyL1",
                "reactiveExpEnergyL2",
                "reactiveExpEnergyL3",
                "reactiveNetEnergyL1",
                "reactiveNetEnergyL2",
                "reactiveNetEnergyL3",
                "apparentImpEnergyL1",
                "apparentImpEnergyL2",
                "apparentImpEnergyL3",
                "apparentExpEnergyL1",
                "apparentExpEnergyL2",
                "apparentExpEnergyL3",
                "apparentNetEnergyL1",
                "apparentNetEnergyL2",
                "apparentNetEnergyL3",
                "currentL1",
                "currentL2",
                "currentL3",
                "currentN",
                "activePowerTotal",
                "activePowerL1",
                "activePowerL2",
                "activePowerL3",
                "reactivePowerTotal",
                "reactivePowerL1",
                "reactivePowerL2",
                "reactivePowerL3",
                "apparentPowerTotal",
                "apparentPowerL1",
                "apparentPowerL2",
                "apparentPowerL3"
            ],
            
            "variable_units": {
                "activeImportEnergyTotal": "(kWh)",
                "activeExportEnergyTotal": "(kWh)",
                "activeNetEnergyTotal": "(kWh)",
                "reactiveImportEnergyTotal": "(kVarh)",
                "reactiveExportEnergyTotal": "(kVarh)",
                "reactiveNetEnergyTotal": "(kVarh)",
                "apparentImportEnergyTotal": "(kVAh)",
                "apparentExportEnergyTotal": "(kVAh)",
                "apparentNetEnergyTotal": "(kVAh)",
                "activeImpEnergyL1": "(kWh)",
                "activeImpEnergyL2": "(kWh)",
                "activeImpEnergyL3": "(kWh)",
                "activeExpEnergyL1": "(kWh)",
                "activeExpEnergyL2": "(kWh)",
                "activeExpEnergyL3": "(kWh)",
                "activeNetEnergyL1": "(kWh)",
                "activeNetEnergyL2": "(kWh)",
                "activeNetEnergyL3": "(kWh)",
                "reactiveImpEnergyL1": "(kVarh)",
                "reactiveImpEnergyL2": "(kVarh)",
                "reactiveImpEnergyL3": "(kVarh)",
                "reactiveExpEnergyL1": "(kVarh)",
                "reactiveExpEnergyL2": "(kVarh)",
                "reactiveExpEnergyL3": "(kVarh)",
                "reactiveNetEnergyL1": "(kVarh)",
                "reactiveNetEnergyL2": "(kVarh)",
                "reactiveNetEnergyL3": "(kVarh)",
                "apparentImpEnergyL1": "(kVAh)",
                "apparentImpEnergyL2": "(kVAh)",
                "apparentImpEnergyL3": "(kVAh)",
                "apparentExpEnergyL1": "(kVAh)",
                "apparentExpEnergyL2": "(kVAh)",
                "apparentExpEnergyL3": "(kVAh)",
                "apparentNetEnergyL1": "(kVAh)",
                "apparentNetEnergyL2": "(kVAh)",
                "apparentNetEnergyL3": "(kVAh)",
                "currentL1": "(A)",
                "currentL2": "(A)",
                "currentL3": "(A)",
                "currentN": "(A)",
                "activePowerTotal": "(W)",
                "activePowerL1": "(W)",
                "activePowerL2": "(W)",
                "activePowerL3": "(W)",
                "reactivePowerTotal": "(var)",
                "reactivePowerL1": "(var)",
                "reactivePowerL2": "(var)",
                "reactivePowerL3": "(var)",
                "apparentPowerTotal": "(VA)",
                "apparentPowerL1": "(VA)",
                "apparentPowerL2": "(VA)",
                "apparentPowerL3": "(VA)"
            }
        }
    }

    SLAVE_DEVICE_ASSIGNMENTS = [{"slaves": SLAVE_IDS, "device_type_key": "M4M"}]

    if SLAVE_DEVICE_ASSIGNMENTS:
        first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
        if first_assigned_device_type_key in DEVICE_TYPE:
            VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
        else:
            VARIABLES = []
    else:
        VARIABLES = []

    # --------------------------------------------- HELPER FUNCTIONS
    def get_device_config_for_slave(slave_id):
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if isinstance(assignment["slaves"], range):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
            elif isinstance(assignment["slaves"], list):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
        else:
            print(f"Warning: Slave ID {slave_id} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping this slave.")
            return None

        if device_type_key not in DEVICE_TYPE:
            print(f"Error: Device type key '{device_type_key}' assigned to slave {slave_id} not found in DEVICE_TYPE configuration.")
            return None
            
        return DEVICE_TYPE[device_type_key], device_type_key



    # -------------------------------------- SENSOR POLLING FUNCTION
    def read_slave_with_client(slave_id, thread_id, client):
        device_config, device_type_key = get_device_config_for_slave(slave_id)
        if device_config is None:
            dummy_len = len(DEVICE_TYPE.get("CurrentSensor", {}).get("variables", []))
            return [None] * dummy_len

        variables = device_config["variables"]
        scale_factors = device_config["scale_factors"]
        decoding_map = device_config["decoding_map"]

        values = []
        thread_prefix = f"[Thread {thread_id}] "

        for var_info in variables:
            addr, size, name, func_code = var_info[:4]
            options = var_info[4] if len(var_info) > 4 else {}

            current_value = None
            try:
                resp = None
                if func_code == 1:
                    resp = client.read_coils(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Coils read error or empty bits.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.bits
                elif func_code == 3:
                    resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Registers read error or empty registers.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.registers
                else:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Unsupported function code {func_code}.")
                    values.append(None)
                    continue

                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

                if name in decoding_map:
                    func_name = decoding_map[name]
                    decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                    if decoder_func:
                        if func_name == "decode_bit":
                            bit_pos = options.get('bit')
                            if bit_pos is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_pos)
                            else:
                                print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                        elif func_name == "decode_bits":
                            bit_range = options.get('bits')
                            if bit_range is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_range)
                            else:
                                print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                        elif func_name == "read_coil":
                            current_value = decoder_func(resp)
                        elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                            current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                        else:
                            current_value = decoder_func(regs_or_bits)
                    else:
                        print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Falling back to raw value processing.")
                        current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
                else:
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

                if name in scale_factors and isinstance(current_value, (int, float)):
                    current_value *= scale_factors[name]

                if isinstance(current_value, list):
                    final_value = flatten_value(current_value)
                else:
                    final_value = current_value

                values.append(final_value)

            except Exception as e:
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
                values.append(None)

        return values


    
    # ------------------------------------------------ SAVE TO EXCEL
    def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
        wb = Workbook()
        wb.remove(wb.active)
        
        for slave_id in SLAVE_IDS:
            config_result = get_device_config_for_slave(slave_id)

            if config_result is None:
                print(f"Skipping sheet creation for Slave_{slave_id} due to missing config.")
                continue

            device_config, _ = config_result

            if not isinstance(device_config, dict):
                print(f"Skipping sheet creation for Slave_{slave_id} due to invalid device config type.")
                continue

            variables = device_config.get("variables", [])

            headers = ['timestamp']
            for var in variables:
                if isinstance(var, dict):
                    headers.append(var.get("name", "unknown_var"))
                elif isinstance(var, (list, tuple)):
                    headers.append(var[2] if len(var) > 2 else "unknown_var")
                else:
                    headers.append("unknown_var")

            ws = wb.create_sheet(title=f"Slave_{slave_id}")
            ws.append(headers)

        for poll_entry in all_poll_data:
            timestamp = poll_entry['timestamp']
            data = poll_entry['data']
            
            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

        wb.save(filename)

    def append_poll_data_to_backup(poll_entry, filename=EXCEL_FILE_BACKUP):
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']

        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
            else:
                wb = Workbook()
                wb.remove(wb.active)
                for slave_id in SLAVE_IDS:
                    config_result = get_device_config_for_slave(slave_id)
                    if config_result is None:
                        continue
                    device_config, _ = config_result
                    variables = device_config.get("variables", [])
                    headers = ['timestamp']
                    for var in variables:
                        if isinstance(var, dict):
                            headers.append(var.get("name", "unknown_var"))
                        elif isinstance(var, (list, tuple)):
                            headers.append(var[2] if len(var) > 2 else "unknown_var")
                        else:
                            headers.append("unknown_var")
                    ws = wb.create_sheet(title=f"Slave_{slave_id}")
                    ws.append(headers)

            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

            wb.save(filename)

        except Exception as e:
            print(f"Error appending to backup: {e}")

    def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
        config_result = get_device_config_for_slave(slave_id)
        if config_result is None:
            print(f"Warning: No device configuration found for slave {slave_id}. Cannot create chart for {var_name}.")
            return
        
        device_config, _ = config_result 

        variable_units = device_config.get("variable_units", {})
        unit = variable_units.get(var_name, "")

        chart = ScatterChart()
        chart.title = f"{var_name} over time - ID {slave_id}"
        chart.style = 13
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = f"{var_name} {unit}".strip()
        chart.legend = None

        chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.x_axis.majorUnit = 500

        x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
        y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

        title_with_unit = f"{var_name} {unit}".strip()
        series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.smooth = True
        chart.series.append(series)

        chart.x_axis.axId = 10 + chart_index * 2
        chart.y_axis.axId = 20 + chart_index * 2
        chart.y_axis.crossAx = 10 + chart_index * 2
        chart.x_axis.crossAx = 20 + chart_index * 2

        chart.height = 9.5
        chart.width = 20.44

        anchor_col = 'BC'
        anchor_row = 1 + 20 * chart_index
        chart.anchor = f"{anchor_col}{anchor_row}"

        sheet.add_chart(chart)

    def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
        wb = load_workbook(filename)

        for slave_id in SLAVE_IDS:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            num_rows = ws.max_row
            
            config_result = get_device_config_for_slave(slave_id)
            if config_result is None:
                print(f"Skipping chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            device_config, _ = config_result

            variables_to_plot_for_slave = device_config.get("plot_variables", [])

            chart_index = 0
            timestamp_col = 1

            for var in variables_to_plot_for_slave:
                if var not in headers:
                    continue

                var_col = headers.index(var) + 1
                create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
                chart_index += 1

        wb.save(filename)



    # ------------------------------------------------ POLLING LOOP
    def poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix):
        if stop_event.is_set():
            return
    
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()
    
        def poll_slave_ids(slave_ids, all_data, lock, thread_id):
            client = ModbusTcpClient(IP, port=PORT)
            if not client.connect():
                print(f"[Thread {thread_id}] Failed to connect")
                return
    
            def task(slave_id):
                values = read_slave_with_client(slave_id, thread_id, client)
                with lock:
                    all_data[slave_id] = values
    
            with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
                executor.map(task, slave_ids)
    
            client.close()
    
        slave_ids = SLAVE_IDS
        t = threading.Thread(target=poll_slave_ids, args=(slave_ids, all_data, lock, 1))
        t.start()
        t.join()
    
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        all_poll_data.append({
            'timestamp': timestamp,
            'data': all_data,
        })

        print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

        st.session_state[f'{state_prefix}last_valid_poll_data'] = {
            'timestamp': timestamp,
            'data': all_data
        }

        try:
            append_poll_data_to_backup({
                'timestamp': timestamp,
                'data': all_data
            })
        except Exception as e:
            print(f"Error saving to backup file: {e}")

    def periodic_polling_loop(stop_event, all_poll_data, state_prefix):
        while not stop_event.is_set():
            start_time = time.time()
            one_cycle = 600
            cycle_duration = 60 * 60

            while time.time() - start_time < one_cycle and not stop_event.is_set():
                poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix)

            if stop_event.is_set():
                break

            all_poll_data.append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': {},
            })

            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = True
            print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
            
            for _ in range(int(cycle_duration)):
                if stop_event.is_set():
                    break
                time.sleep(1)
            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = False



    # ---------------------------------------- SPLIT EXCEL FUNCTIONS
    def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_file_excel):
            st.error(f"Error: File '{path_file_excel}' not found.")
            return None

        try:
            if not os.path.exists(name_output_folder):
                os.makedirs(name_output_folder)
            else:
                for file_name in os.listdir(name_output_folder):
                    path_old_file = os.path.join(name_output_folder, file_name)
                    if os.path.isfile(path_old_file):
                        os.remove(path_old_file)
                st.info(f"Cleaning folder '{name_output_folder}'.")

            source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
            generated_files = []

            for i, slave_id in enumerate(selected_slave_ids):
                sheet_name = f"Slave_{slave_id}"
                if status_text:
                    status_text.text(f"Processing Slave_{slave_id}...")
                if sheet_name not in source_workbook.sheetnames:
                    st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                    continue

                sheet = source_workbook[sheet_name]

                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in sheet.iter_rows(values_only=True):
                    new_ws.append(row)

                num_rows = new_ws.max_row
                num_cols = new_ws.max_column

                if num_rows < 2 or num_cols < 2:
                    st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                    continue

                timestamp_col = 1
                
                config_result = get_device_config_for_slave(slave_id)
                if config_result is None:
                    print(f"Skipping chart generation for Slave_{slave_id} in split function due to missing device configuration.")
                    continue

                device_config, _ = config_result

                variables_to_plot_for_slave = device_config["plot_variables"]

                for idx, var_name in enumerate(variables_to_plot_for_slave):
                    if var_name in [cell.value for cell in new_ws[1]]:
                        var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                        create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                    else:
                        st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

                output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
                output_filepath = os.path.join(name_output_folder, output_filename)

                new_wb.save(output_filepath)
                generated_files.append(output_filepath)
                
                if progress_bar and total_slaves > 0:
                    progress_percent = int(((i + 1) / total_slaves) * 100)
                    progress_bar.progress(progress_percent)

            if not generated_files:
                st.warning("No individual files were generated for the selected slaves.")
                return None

            name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
            complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

            with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_to_add in generated_files:
                    zipf.write(file_to_add, os.path.basename(file_to_add))

            st.success(f"All files were zipped into '{complete_zip_path}'")
            return complete_zip_path

        except Exception as e:
            st.error(f"Error splitting and generating charts: {e}")
            return None

    def extract_slave_id(sheet_name):
        match = re.search(r'\d+', sheet_name)
        return int(match.group(0)) if match else None

    def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
            st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
            return None

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        else:
            for file_name in os.listdir(output_folder):
                path_to_remove = os.path.join(output_folder, file_name)
                if os.path.isfile(path_to_remove):
                    os.remove(path_to_remove)
            st.info(f"Folder '{output_folder}' cleaned.")

        wb_modbus = openpyxl.load_workbook(path_modbus_log)
        wb_backup = openpyxl.load_workbook(path_backup_data)
        
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in wb_modbus.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue
            if sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
                continue

            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping merge/chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

            modbus_sheet = wb_modbus[sheet_name]
            backup_sheet = wb_backup[sheet_name]

            modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
            backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

            modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
            backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

            merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
            merged_data.dropna(subset=["timestamp"], inplace=True)
            merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
            merged_data.sort_values("timestamp", inplace=True)
            merged_data.reset_index(drop=True, inplace=True)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            new_sheet = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in modbus_sheet[1]]
            new_sheet.append(headers)

            for r_idx, row in merged_data.iterrows():
                new_sheet.append(list(row.values))

            num_rows = new_sheet.max_row
            current_slave_id = extract_slave_id(sheet_name)
            
            variables_to_plot_for_slave = device_config["plot_variables"]

            for chart_index, var_name in enumerate(variables_to_plot_for_slave):
                if var_name in headers:
                    col_idx = headers.index(var_name) + 1
                    create_scatter_chart(
                        new_sheet,
                        timestamp_col=1,
                        var_col=col_idx,
                        num_rows=num_rows,
                        var_name=var_name,
                        slave_id=current_slave_id,
                        chart_index=chart_index
                    )
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
            new_wb.save(path_output)
            generated_files.append(path_output)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All merged individual files have been zipped into: {zip_path}")
        return zip_path

    def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
        if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
            st.error("Excel files (final or backup) not found for comparison.")
            return False

        wb_final = load_workbook(excel_file_final, data_only=True)
        wb_backup = load_workbook(excel_file_backup, data_only=True)

        needs_merge = False
        for slave_id in selected_slave_ids:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
                continue

            ws_final = wb_final[sheet_name]
            ws_backup = wb_backup[sheet_name]

            timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
            timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

            if not timestamps_backup.issubset(timestamps_final):
                needs_merge = True
                break
        
        return needs_merge


    # ----------------------------------------- STREAMLIT DASHBOARD
    with st.container():
        st.title("Long Term Test for ABB Ability™ Edge Industrial Gateway")

        if f'{state_prefix}polling' not in st.session_state:
            st.session_state[f'{state_prefix}last_valid_poll_data'] = {}
            st.session_state[f'{state_prefix}polling'] = False
            st.session_state[f'{state_prefix}stop_event'] = threading.Event()
            st.session_state[f'{state_prefix}thread'] = None
            st.session_state[f'{state_prefix}all_poll_data'] = []
            
            st.session_state[f'{state_prefix}selected_slaves_by_type'] = {
                device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
            }
            st.session_state[f'{state_prefix}realtime_table_placeholders'] = {}
            st.session_state[f'{state_prefix}last_data_length'] = 0
            st.session_state[f'{state_prefix}last_ui_update_time'] = time.time()
            st.session_state[f'{state_prefix}last_table_update_timestamp'] = "N/A"

        if st.button("Start Monitoring", key=f"start_monitoring_{state_prefix}", disabled=st.session_state[f'{state_prefix}polling']):
            if not st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}all_poll_data'] = []
                st.session_state[f'{state_prefix}stop_event'].clear()

                st.session_state[f'{state_prefix}thread'] = threading.Thread(
                    target=periodic_polling_loop,
                    args=(
                        st.session_state[f'{state_prefix}stop_event'],
                        st.session_state[f'{state_prefix}all_poll_data'],
                        state_prefix
                    ),
                    daemon=True
                )

                st.session_state[f'{state_prefix}thread'].start()
                st.session_state[f'{state_prefix}polling'] = True
                st.session_state[f'{state_prefix}last_data_length'] = 0
                st.success("Monitoring started!")

        if st.button("Stop Monitoring", key=f"stop_monitoring_{state_prefix}"):
            if st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}stop_event'].set()
                if st.session_state[f'{state_prefix}thread'] and st.session_state[f'{state_prefix}thread'].is_alive():
                    st.session_state[f'{state_prefix}thread'].join(timeout=10)
                    if st.session_state[f'{state_prefix}thread'].is_alive():
                        st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
                    else:
                        st.info("Polling thread terminated.")
                st.session_state[f'{state_prefix}polling'] = False

                with st.spinner("Saving data to Excel..."):
                    save_all_data_to_excel(st.session_state[f'{state_prefix}all_poll_data'], filename=EXCEL_FILE_FINAL)

                    try:
                        add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                        st.success("Monitoring stopped, data and charts saved to Excel.")
                    except Exception as e:
                        st.error(f"Error adding charts: {e}")
            else:
                st.warning("Monitoring is not running.")

        st.divider()

        if st.session_state.get(f"{state_prefix}all_poll_data"):
            all_poll_data = st.session_state[f"{state_prefix}all_poll_data"]

            st.markdown(
                f"Total of **{int(len(all_poll_data))}** reading(s) completed so far."
            )

            device_counts = {}

            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                key = assignment["device_type_key"]
                normalized_key = re.match(r'^[A-Za-z]+', key)
                normalized_key = normalized_key.group(0) if normalized_key else key
                if "ips" in assignment:
                    count = len(assignment["ips"])
                elif "slaves" in assignment:
                    count = len(assignment["slaves"])
                else:
                    count = 0
                device_counts[key] = device_counts.get(key, 0) + count

            device_count_lines = [f"- **{k}**: {v} device(s)" for k, v in device_counts.items()]
            st.markdown("### Active devices on the system:")
            st.markdown("\n".join(device_count_lines))

            last_poll_data = all_poll_data[-1]
            if "elapsed" not in last_poll_data and len(all_poll_data) >= 2:
                last_poll_data = all_poll_data[-2]

            last_poll_time = last_poll_data.get("elapsed", None)
            last_sleep_time = last_poll_data.get("sleep_time", None)

            if last_poll_time is not None and last_sleep_time is not None:
                st.markdown(
                    f"Completed one polling in **{last_poll_time:.2f}s**, sleeping for **{last_sleep_time:.2f}s**."
                )

            if last_poll_data.get("cycle_completed"):
                cycle_duration = last_poll_data.get("cycle_duration", 0)
                st.markdown(
                    f"Waiting **{int(cycle_duration / 60)} minutes** before next reading cycle."
                )

        st.subheader("Realtime data monitoring by sensor type")

        st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'] = st.empty()

        assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

        for device_type_key in assigned_device_types_for_display:
            device_config = DEVICE_TYPE.get(device_type_key)
            if not device_config:
                continue

            has_slaves_assigned = any(
                assignment["device_type_key"] == device_type_key
                for assignment in SLAVE_DEVICE_ASSIGNMENTS
            )
            if not has_slaves_assigned:
                continue

            available_slaves_for_type = []
            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                if assignment["device_type_key"] == device_type_key:
                    if isinstance(assignment["slaves"], range):
                        available_slaves_for_type.extend(list(assignment["slaves"]))
                    elif isinstance(assignment["slaves"], list):
                        available_slaves_for_type.extend(assignment["slaves"])
            available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

            if not available_slaves_for_type:
                continue

            with st.expander(f"Configuration and Monitoring: **{device_type_key}**", expanded=True):
                selected = st.multiselect(
                    f"Select Slave IDs for {device_type_key}:",
                    options=available_slaves_for_type,
                    default=st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, []),
                    key=f"{state_prefix}multiselect_{device_type_key}"
                )

                if st.button(f"Confirm selection for {device_type_key}", key=f"{state_prefix}confirm_button_{device_type_key}"):
                    st.session_state[f'{state_prefix}selected_slaves_by_type'][device_type_key] = selected

                st.session_state[f'{state_prefix}realtime_table_placeholders'][device_type_key] = st.empty()

        def update_realtime_tables_by_type():
            current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

            st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'].write(
                f"Last widget update at **{current_ui_update_timestamp}**"
            )

            if st.session_state[f'{state_prefix}all_poll_data']:
                all_poll_data = st.session_state[f'{state_prefix}all_poll_data']
                latest_poll_entry = all_poll_data[-1]

                if latest_poll_entry.get("data"):
                    raw_data = latest_poll_entry['data']
                    st.session_state[f'{state_prefix}last_valid_poll_data'] = latest_poll_entry
                else:
                    raw_data = st.session_state.get(f'{state_prefix}last_valid_poll_data', {}).get('data', {})

                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    selected_slaves = st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, [])
                    device_config = DEVICE_TYPE.get(device_type_key)

                    placeholder.empty()

                    if not selected_slaves or not device_config:
                        placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                        continue

                    table_data = []
                    variable_names = [var[2] for var in device_config["variables"]]

                    for slave_id in sorted(selected_slaves):
                        if slave_id in raw_data:
                            values = raw_data[slave_id]
                            row_dict = {"Slave ID": f"Slave {slave_id}"}
                            for i, var_name in enumerate(variable_names):
                                if i < len(values) and values[i] is not None:
                                    display_value = values[i]
                                    if isinstance(display_value, float):
                                        display_value = round(display_value, 2)
                                else:
                                    display_value = np.nan
                                row_dict[var_name] = display_value
                            table_data.append(row_dict)
                        else:
                            table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                    if table_data:
                        df = pd.DataFrame(table_data)
                        df = df.set_index("Slave ID")

                        placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                    else:
                        placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
            else:
                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    placeholder.empty()
                    placeholder.info(f"Waiting for first polling data for {device_type_key}...")

        update_realtime_tables_by_type()

        if st.session_state[f'{state_prefix}polling']:
            st_autorefresh(interval=5000, key=f"realtime_autorefresh_{state_prefix}")

        if not os.path.exists(EXCEL_FILE_FINAL):
            if not st.session_state[f'{state_prefix}polling']:
                st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
        else:
            st.divider()
            all_possible_slave_ids = sorted(list(set(SLAVE_IDS)))

            display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

            if "All" in st.session_state.get('selected_display_options', []):
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=["All"],
                    disabled=True
                )
            else:
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=[]
                )

            st.session_state.selected_display_options = selected_display_options

            selected_slave_ids_numeric = []
            if "All" in selected_display_options:
                selected_slave_ids_numeric = all_possible_slave_ids
            elif selected_display_options:
                for ds_id in selected_display_options:
                    match = re.search(r'\d+', ds_id)
                    if match:
                        selected_slave_ids_numeric.append(int(match.group(0)))

            if st.button("Generate Individual Files", key=f"generate_individual_files_{state_prefix}"):
                if not selected_slave_ids_numeric:
                    st.warning("Please select at least one slave or 'All'.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    total_slaves_to_process = len(selected_slave_ids_numeric)
                    zip_path = None

                    if not os.path.exists(EXCEL_FILE_BACKUP):
                        st.warning("Backup file not found. Performing only the split of the final Excel.")
                        with st.spinner("Splitting Excel and generating charts..."):
                            zip_path = split_and_generate_charts_for_slaves(
                                EXCEL_FILE_FINAL,
                                selected_slave_ids_numeric,
                                progress_bar=progress_bar,
                                status_text=status_text,
                                total_slaves=total_slaves_to_process
                            )
                    else:
                        st.info("Comparing timestamps between final Excel and backup...")
                        needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                        if needs_merge:
                            st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                            with st.spinner("Merging data from backup and final, and generating individual files..."):
                                zip_path = merge_data_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    EXCEL_FILE_BACKUP,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )
                        else:
                            st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                            with st.spinner("Splitting Excel and generating charts..."):
                                zip_path = split_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )

                    if zip_path:
                        status_text.text("Processing complete!")
                        st.success("Individual files generated and available for download.")



# -------------------------
# SCRIPT 4: SCU100
# -------------------------
def run_SCU100():
    
    # ---------------------------------------- CONFIGURATION SECTION
    IP = '192.168.1.100'
    PORT = 502
    SLAVE_IDS = [3, 10, 12]
    POLL_INTERVAL = 60
    EXCEL_FILE_FINAL = "SCU100_modbus_poll_log.xlsx"
    EXCEL_FILE_BACKUP = "SCU100_backup_poll_data.xlsx"
    state_prefix = "SCU100_"

    DEVICE_TYPE = {

        "M4M_30_A2": {
            "variables": [
                (40588, 2, "voltageL1", 3),
                (40590, 2, "voltageL2", 3),
                (40592, 2, "voltageL3", 3),
                (40594, 2, "voltageL1_L2", 3),
                (40596, 2, "voltageL3_L2", 3),
                (40598, 2, "voltageL1_L3", 3),
                (40602, 2, "currentL1", 3),
                (40604, 2, "currentL2", 3),
                (40606, 2, "currentL3", 3),
                (40608, 2, "currentN", 3),
                (40612, 2, "activePowerTotal", 3),
                (40614, 2, "activePowerL1", 3),
                (40616, 2, "activePowerL2", 3),
                (40618, 2, "activePowerL3", 3),
                (40620, 2, "reactivePowerTotal", 3),
                (40622, 2, "reactivePowerL1", 3),
                (40624, 2, "reactivePowerL2", 3),
                (40626, 2, "reactivePowerL3", 3),
                (40628, 2, "apparentPowerTotal", 3),
                (40630, 2, "apparentPowerL1", 3),
                (40632, 2, "apparentPowerL2", 3),
                (40634, 2, "apparentPowerL3", 3)
            ],
            
            "scale_factors": {
                "voltageL1": 0.01,
                "voltageL2": 0.01,
                "voltageL3": 0.01,
                "voltageL1_L2": 0.01,
                "voltageL3_L2": 0.01,
                "voltageL1_L3": 0.01,
                "currentL1": 0.01,
                "currentL2": 0.01,
                "currentL3": 0.01,
                "currentN": 0.01,
                "activePowerTotal": 0.01,
                "activePowerL1": 0.01,
                "activePowerL2": 0.01,
                "activePowerL3": 0.01,
                "reactivePowerTotal": 0.01,
                "reactivePowerL1": 0.01,
                "reactivePowerL2": 0.01,
                "reactivePowerL3": 0.01,
                "apparentPowerTotal": 0.01,
                "apparentPowerL1": 0.01,
                "apparentPowerL2": 0.01,
                "apparentPowerL3": 0.01
            },
            
            "decoding_map": {
                "voltageL1": "decode_32bit_unsigned",
                "voltageL2": "decode_32bit_unsigned",
                "voltageL3": "decode_32bit_unsigned",
                "voltageL1_L2": "decode_32bit_unsigned",
                "voltageL3_L2": "decode_32bit_unsigned",
                "voltageL1_L3": "decode_32bit_unsigned",
                "currentL1": "decode_32bit_unsigned",
                "currentL2": "decode_32bit_unsigned",
                "currentL3": "decode_32bit_unsigned",
                "currentN": "decode_32bit_unsigned",
                "activePowerTotal": "decode_32bit_signed",
                "activePowerL1": "decode_32bit_signed",
                "activePowerL2": "decode_32bit_signed",
                "activePowerL3": "decode_32bit_signed",
                "reactivePowerTotal": "decode_32bit_signed",
                "reactivePowerL1": "decode_32bit_signed",
                "reactivePowerL2": "decode_32bit_signed",
                "reactivePowerL3": "decode_32bit_signed",
                "apparentPowerTotal": "decode_32bit_signed",
                "apparentPowerL1": "decode_32bit_signed",
                "apparentPowerL2": "decode_32bit_signed",
                "apparentPowerL3": "decode_32bit_signed"
            },
            
            "plot_variables": [
                "voltageL1",
                "voltageL2",
                "voltageL3",
                "voltageL1_L2",
                "voltageL3_L2",
                "voltageL1_L3",
                "currentL1",
                "currentL2",
                "currentL3",
                "currentN",
                "activePowerTotal",
                "activePowerL1",
                "activePowerL2",
                "activePowerL3",
                "reactivePowerTotal",
                "reactivePowerL1",
                "reactivePowerL2",
                "reactivePowerL3",
                "apparentPowerTotal",
                "apparentPowerL1",
                "apparentPowerL2",
                "apparentPowerL3"
            ],
            
            "variable_units": {
                "voltageL1": "(V)",
                "voltageL2": "(V)",
                "voltageL3": "(V)",
                "voltageL1_L2": "(V)",
                "voltageL3_L2": "(V)",
                "voltageL1_L3": "(V)",
                "currentL1": "(A)",
                "currentL2": "(A)",
                "currentL3": "(A)",
                "currentN": "(A)",
                "activePowerTotal": "(W)",
                "activePowerL1": "(W)",
                "activePowerL2": "(W)",
                "activePowerL3": "(W)",
                "reactivePowerTotal": "(var)",
                "reactivePowerL1": "(var)",
                "reactivePowerL2": "(var)",
                "reactivePowerL3": "(var)",
                "apparentPowerTotal": "(VA)",
                "apparentPowerL1": "(VA)",
                "apparentPowerL2": "(VA)",
                "apparentPowerL3": "(VA)"
            }
        },

        "M4M_30_F4": {
            "variables": [
                (50346, 4, "activeImportEnergyTotal", 3),
                (50350, 4, "activeExportEnergyTotal", 3),
                (50354, 4, "activeNetEnergyTotal", 3),
                (50358, 4, "reactiveImportEnergyTotal", 3),
                (50362, 4, "reactiveExportEnergyTotal", 3),
                (50366, 4, "reactiveNetEnergyTotal", 3),
                (50370, 4, "apparentImportEnergyTotal", 3),
                (50374, 4, "apparentExportEnergyTotal", 3),
                (50378, 4, "apparentNetEnergyTotal", 3)
            ],
            
            "scale_factors": {
                "activeImportEnergyTotal": 0.01,
                "activeExportEnergyTotal": 0.01,
                "activeNetEnergyTotal": 0.01,
                "reactiveImportEnergyTotal": 0.01,
                "reactiveExportEnergyTotal": 0.01,
                "reactiveNetEnergyTotal": 0.01,
                "apparentImportEnergyTotal": 0.01,
                "apparentExportEnergyTotal": 0.01,
                "apparentNetEnergyTotal": 0.01
            },
            
            "decoding_map": {
                "activeImportEnergyTotal": "decode_64bit_unsigned",
                "activeExportEnergyTotal": "decode_64bit_unsigned",
                "activeNetEnergyTotal": "decode_64bit_signed",
                "reactiveImportEnergyTotal": "decode_64bit_unsigned",
                "reactiveExportEnergyTotal": "decode_64bit_unsigned",
                "reactiveNetEnergyTotal": "decode_64bit_signed",
                "apparentImportEnergyTotal": "decode_64bit_unsigned",
                "apparentExportEnergyTotal": "decode_64bit_unsigned",
                "apparentNetEnergyTotal": "decode_64bit_signed"
            },
            
            "plot_variables": [
                "activeImportEnergyTotal",
                "activeExportEnergyTotal",
                "activeNetEnergyTotal",
                "reactiveImportEnergyTotal",
                "reactiveExportEnergyTotal",
                "reactiveNetEnergyTotal",
                "apparentImportEnergyTotal",
                "apparentExportEnergyTotal",
                "apparentNetEnergyTotal"
            ],
            
            "variable_units": {
                "activeImportEnergyTotal": "(kWh)",
                "activeExportEnergyTotal": "(kWh)",
                "activeNetEnergyTotal": "(kWh)",
                "reactiveImportEnergyTotal": "(kVarh)",
                "reactiveExportEnergyTotal": "(kVarh)",
                "reactiveNetEnergyTotal": "(kVarh)",
                "apparentImportEnergyTotal": "(kVAh)",
                "apparentExportEnergyTotal": "(kVAh)",
                "apparentNetEnergyTotal": "(kVAh)"
            }
        },

        "M4M_20_F5": {
            "variables": [
                (52411, 4, "activeImpEnergyL1", 3),
                (52415, 4, "activeImpEnergyL2", 3),
                (52419, 4, "activeImpEnergyL3", 3),
                (52423, 4, "activeExpEnergyL1", 3),
                (52427, 4, "activeExpEnergyL2", 3),
                (52431, 4, "activeExpEnergyL3", 3),
                (52435, 4, "activeNetEnergyL1", 3),
                (52439, 4, "activeNetEnergyL2", 3),
                (52443, 4, "activeNetEnergyL3", 3),
                (52447, 4, "reactiveImpEnergyL1", 3),
                (52451, 4, "reactiveImpEnergyL2", 3),
                (52455, 4, "reactiveImpEnergyL3", 3),
                (52459, 4, "reactiveExpEnergyL1", 3),
                (52463, 4, "reactiveExpEnergyL2", 3),
                (52467, 4, "reactiveExpEnergyL3", 3),
                (52471, 4, "reactiveNetEnergyL1", 3),
                (52475, 4, "reactiveNetEnergyL2", 3),
                (52479, 4, "reactiveNetEnergyL3", 3),
                (52483, 4, "apparentImpEnergyL1", 3),
                (52487, 4, "apparentImpEnergyL2", 3),
                (52491, 4, "apparentImpEnergyL3", 3),
                (52495, 4, "apparentExpEnergyL1", 3),
                (52499, 4, "apparentExpEnergyL2", 3),
                (52503, 4, "apparentExpEnergyL3", 3)
            ],
            
            "scale_factors": {
                "activeImpEnergyL1": 0.01,
                "activeImpEnergyL2": 0.01,
                "activeImpEnergyL3": 0.01,
                "activeExpEnergyL1": 0.01,
                "activeExpEnergyL2": 0.01,
                "activeExpEnergyL3": 0.01,
                "activeNetEnergyL1": 0.01,
                "activeNetEnergyL2": 0.01,
                "activeNetEnergyL3": 0.01,
                "reactiveImpEnergyL1": 0.01,
                "reactiveImpEnergyL2": 0.01,
                "reactiveImpEnergyL3": 0.01,
                "reactiveExpEnergyL1": 0.01,
                "reactiveExpEnergyL2": 0.01,
                "reactiveExpEnergyL3": 0.01,
                "reactiveNetEnergyL1": 0.01,
                "reactiveNetEnergyL2": 0.01,
                "reactiveNetEnergyL3": 0.01,
                "apparentImpEnergyL1": 0.01,
                "apparentImpEnergyL2": 0.01,
                "apparentImpEnergyL3": 0.01,
                "apparentExpEnergyL1": 0.01,
                "apparentExpEnergyL2": 0.01,
                "apparentExpEnergyL3": 0.01
            },
            
            "decoding_map": {
                "activeImpEnergyL1": "decode_64bit_unsigned",
                "activeImpEnergyL2": "decode_64bit_unsigned",
                "activeImpEnergyL3": "decode_64bit_unsigned",
                "activeExpEnergyL1": "decode_64bit_unsigned",
                "activeExpEnergyL2": "decode_64bit_unsigned",
                "activeExpEnergyL3": "decode_64bit_unsigned",
                "activeNetEnergyL1": "decode_64bit_signed",
                "activeNetEnergyL2": "decode_64bit_signed",
                "activeNetEnergyL3": "decode_64bit_signed",
                "reactiveImpEnergyL1": "decode_64bit_unsigned",
                "reactiveImpEnergyL2": "decode_64bit_unsigned",
                "reactiveImpEnergyL3": "decode_64bit_unsigned",
                "reactiveExpEnergyL1": "decode_64bit_unsigned",
                "reactiveExpEnergyL2": "decode_64bit_unsigned",
                "reactiveExpEnergyL3": "decode_64bit_unsigned",
                "reactiveNetEnergyL1": "decode_64bit_signed",
                "reactiveNetEnergyL2": "decode_64bit_signed",
                "reactiveNetEnergyL3": "decode_64bit_signed",
                "apparentImpEnergyL1": "decode_64bit_unsigned",
                "apparentImpEnergyL2": "decode_64bit_unsigned",
                "apparentImpEnergyL3": "decode_64bit_unsigned",
                "apparentExpEnergyL1": "decode_64bit_unsigned",
                "apparentExpEnergyL2": "decode_64bit_unsigned",
                "apparentExpEnergyL3": "decode_64bit_unsigned"
            },
            
            "plot_variables": [
                "activeImpEnergyL1",
                "activeImpEnergyL2",
                "activeImpEnergyL3",
                "activeExpEnergyL1",
                "activeExpEnergyL2",
                "activeExpEnergyL3",
                "activeNetEnergyL1",
                "activeNetEnergyL2",
                "activeNetEnergyL3",
                "reactiveImpEnergyL1",
                "reactiveImpEnergyL2",
                "reactiveImpEnergyL3",
                "reactiveExpEnergyL1",
                "reactiveExpEnergyL2",
                "reactiveExpEnergyL3",
                "reactiveNetEnergyL1",
                "reactiveNetEnergyL2",
                "reactiveNetEnergyL3",
                "apparentImpEnergyL1",
                "apparentImpEnergyL2",
                "apparentImpEnergyL3",
                "apparentExpEnergyL1",
                "apparentExpEnergyL2",
                "apparentExpEnergyL3"
            ],
            
            "variable_units": {
                "activeImpEnergyL1": "(kWh)",
                "activeImpEnergyL2": "(kWh)",
                "activeImpEnergyL3": "(kWh)",
                "activeExpEnergyL1": "(kWh)",
                "activeExpEnergyL2": "(kWh)",
                "activeExpEnergyL3": "(kWh)",
                "activeNetEnergyL1": "(kWh)",
                "activeNetEnergyL2": "(kWh)",
                "activeNetEnergyL3": "(kWh)",
                "reactiveImpEnergyL1": "(kVarh)",
                "reactiveImpEnergyL2": "(kVarh)",
                "reactiveImpEnergyL3": "(kVarh)",
                "reactiveExpEnergyL1": "(kVarh)",
                "reactiveExpEnergyL2": "(kVarh)",
                "reactiveExpEnergyL3": "(kVarh)",
                "reactiveNetEnergyL1": "(kVarh)",
                "reactiveNetEnergyL2": "(kVarh)",
                "reactiveNetEnergyL3": "(kVarh)",
                "apparentImpEnergyL1": "(kVAh)",
                "apparentImpEnergyL2": "(kVAh)",
                "apparentImpEnergyL3": "(kVAh)",
                "apparentExpEnergyL1": "(kVAh)",
                "apparentExpEnergyL2": "(kVAh)",
                "apparentExpEnergyL3": "(kVAh)"
            }
        }
    }


    SLAVE_DEVICE_ASSIGNMENTS = [{"slaves": [3], "device_type_key": "M4M_30_A2"}, {"slaves": [10], "device_type_key": "M4M_30_F4"}, {"slaves": [12], "device_type_key": "M4M_20_F5"}]

    if SLAVE_DEVICE_ASSIGNMENTS:
        first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
        if first_assigned_device_type_key in DEVICE_TYPE:
            VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
        else:
            VARIABLES = []
    else:
        VARIABLES = []

    # --------------------------------------------- HELPER FUNCTIONS
    def get_device_config_for_slave(slave_id):
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if isinstance(assignment["slaves"], range):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
            elif isinstance(assignment["slaves"], list):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
        else:
            print(f"Warning: Slave ID {slave_id} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping this slave.")
            return None

        if device_type_key not in DEVICE_TYPE:
            print(f"Error: Device type key '{device_type_key}' assigned to slave {slave_id} not found in DEVICE_TYPE configuration.")
            return None
            
        return DEVICE_TYPE[device_type_key], device_type_key



    # -------------------------------------- SENSOR POLLING FUNCTION
    def read_slave_with_client(slave_id, thread_id, client):
        device_config, device_type_key = get_device_config_for_slave(slave_id)
        if device_config is None:
            dummy_len = len(DEVICE_TYPE.get("CurrentSensor", {}).get("variables", []))
            return [None] * dummy_len

        variables = device_config["variables"]
        scale_factors = device_config["scale_factors"]
        decoding_map = device_config["decoding_map"]

        values = []
        thread_prefix = f"[Thread {thread_id}] "

        for var_info in variables:
            addr, size, name, func_code = var_info[:4]
            options = var_info[4] if len(var_info) > 4 else {}

            current_value = None
            try:
                resp = None
                if func_code == 1:
                    resp = client.read_coils(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Coils read error or empty bits.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.bits
                elif func_code == 3:
                    resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Registers read error or empty registers.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.registers
                    if func_code == 3 and size > 1:
                        regs_or_bits = list(reversed(regs_or_bits))
                else:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Unsupported function code {func_code}.")
                    values.append(None)
                    continue

                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

                if name in decoding_map:
                    func_name = decoding_map[name]
                    decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                    if decoder_func:
                        if func_name == "decode_bit":
                            bit_pos = options.get('bit')
                            if bit_pos is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_pos)
                            else:
                                print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                        elif func_name == "decode_bits":
                            bit_range = options.get('bits')
                            if bit_range is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_range)
                            else:
                                print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                        elif func_name == "read_coil":
                            current_value = decoder_func(resp)
                        elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                            current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                        else:
                            current_value = decoder_func(regs_or_bits)
                    else:
                        print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Falling back to raw value processing.")
                        current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
                else:
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

                if name in scale_factors and isinstance(current_value, (int, float)):
                    current_value *= scale_factors[name]

                if isinstance(current_value, list):
                    final_value = flatten_value(current_value)
                else:
                    final_value = current_value

                values.append(final_value)

            except Exception as e:
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
                values.append(None)

        return values


    
    # ------------------------------------------------ SAVE TO EXCEL
    def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
        wb = Workbook()
        wb.remove(wb.active)
        
        for slave_id in SLAVE_IDS:
            config_result = get_device_config_for_slave(slave_id)

            if config_result is None:
                print(f"Skipping sheet creation for Slave_{slave_id} due to missing config.")
                continue

            device_config, _ = config_result

            if not isinstance(device_config, dict):
                print(f"Skipping sheet creation for Slave_{slave_id} due to invalid device config type.")
                continue

            variables = device_config.get("variables", [])

            headers = ['timestamp']
            for var in variables:
                if isinstance(var, dict):
                    headers.append(var.get("name", "unknown_var"))
                elif isinstance(var, (list, tuple)):
                    headers.append(var[2] if len(var) > 2 else "unknown_var")
                else:
                    headers.append("unknown_var")

            ws = wb.create_sheet(title=f"Slave_{slave_id}")
            ws.append(headers)

        for poll_entry in all_poll_data:
            timestamp = poll_entry['timestamp']
            data = poll_entry['data']
            
            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

        wb.save(filename)

    def append_poll_data_to_backup(poll_entry, filename=EXCEL_FILE_BACKUP):
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']

        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
            else:
                wb = Workbook()
                wb.remove(wb.active)
                for slave_id in SLAVE_IDS:
                    config_result = get_device_config_for_slave(slave_id)
                    if config_result is None:
                        continue
                    device_config, _ = config_result
                    variables = device_config.get("variables", [])
                    headers = ['timestamp']
                    for var in variables:
                        if isinstance(var, dict):
                            headers.append(var.get("name", "unknown_var"))
                        elif isinstance(var, (list, tuple)):
                            headers.append(var[2] if len(var) > 2 else "unknown_var")
                        else:
                            headers.append("unknown_var")
                    ws = wb.create_sheet(title=f"Slave_{slave_id}")
                    ws.append(headers)

            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

            wb.save(filename)

        except Exception as e:
            print(f"Error appending to backup: {e}")

    def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
        config_result = get_device_config_for_slave(slave_id)
        if config_result is None:
            print(f"Warning: No device configuration found for slave {slave_id}. Cannot create chart for {var_name}.")
            return
        
        device_config, _ = config_result 

        variable_units = device_config.get("variable_units", {})
        unit = variable_units.get(var_name, "")

        chart = ScatterChart()
        chart.title = f"{var_name} over time - ID {slave_id}"
        chart.style = 13
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = f"{var_name} {unit}".strip()
        chart.legend = None

        chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.x_axis.majorUnit = 500

        x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
        y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

        title_with_unit = f"{var_name} {unit}".strip()
        series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.smooth = True
        chart.series.append(series)

        chart.x_axis.axId = 10 + chart_index * 2
        chart.y_axis.axId = 20 + chart_index * 2
        chart.y_axis.crossAx = 10 + chart_index * 2
        chart.x_axis.crossAx = 20 + chart_index * 2

        chart.height = 9.5
        chart.width = 20.44

        anchor_col = 'BC'
        anchor_row = 1 + 20 * chart_index
        chart.anchor = f"{anchor_col}{anchor_row}"

        sheet.add_chart(chart)

    def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
        wb = load_workbook(filename)

        for slave_id in SLAVE_IDS:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            num_rows = ws.max_row
            
            config_result = get_device_config_for_slave(slave_id)
            if config_result is None:
                print(f"Skipping chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            device_config, _ = config_result

            variables_to_plot_for_slave = device_config.get("plot_variables", [])

            chart_index = 0
            timestamp_col = 1

            for var in variables_to_plot_for_slave:
                if var not in headers:
                    continue

                var_col = headers.index(var) + 1
                create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
                chart_index += 1

        wb.save(filename)



    # ------------------------------------------------ POLLING LOOP
    def poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix):
        if stop_event.is_set():
            return
    
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()
    
        def poll_slave_ids(slave_ids, all_data, lock, thread_id):
            client = ModbusTcpClient(IP, port=PORT)
            if not client.connect():
                print(f"[Thread {thread_id}] Failed to connect")
                return
    
            def task(slave_id):
                values = read_slave_with_client(slave_id, thread_id, client)
                with lock:
                    all_data[slave_id] = values
    
            with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
                executor.map(task, slave_ids)
    
            client.close()
    
        slave_ids = SLAVE_IDS
        t = threading.Thread(target=poll_slave_ids, args=(slave_ids, all_data, lock, 1))
        t.start()
        t.join()
    
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        all_poll_data.append({
            'timestamp': timestamp,
            'data': all_data,
        })

        print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

        st.session_state[f'{state_prefix}last_valid_poll_data'] = {
            'timestamp': timestamp,
            'data': all_data
        }

        try:
            append_poll_data_to_backup({
                'timestamp': timestamp,
                'data': all_data
            })
        except Exception as e:
            print(f"Error saving to backup file: {e}")

    def periodic_polling_loop(stop_event, all_poll_data, state_prefix):
        while not stop_event.is_set():
            start_time = time.time()
            one_cycle = 600
            cycle_duration = 60 * 60

            while time.time() - start_time < one_cycle and not stop_event.is_set():
                poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix)

            if stop_event.is_set():
                break

            all_poll_data.append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': {},
            })

            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = True
            print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
            
            for _ in range(int(cycle_duration)):
                if stop_event.is_set():
                    break
                time.sleep(1)
            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = False



    # ---------------------------------------- SPLIT EXCEL FUNCTIONS
    def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_file_excel):
            st.error(f"Error: File '{path_file_excel}' not found.")
            return None

        try:
            if not os.path.exists(name_output_folder):
                os.makedirs(name_output_folder)
            else:
                for file_name in os.listdir(name_output_folder):
                    path_old_file = os.path.join(name_output_folder, file_name)
                    if os.path.isfile(path_old_file):
                        os.remove(path_old_file)
                st.info(f"Cleaning folder '{name_output_folder}'.")

            source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
            generated_files = []

            for i, slave_id in enumerate(selected_slave_ids):
                sheet_name = f"Slave_{slave_id}"
                if status_text:
                    status_text.text(f"Processing Slave_{slave_id}...")
                if sheet_name not in source_workbook.sheetnames:
                    st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                    continue

                sheet = source_workbook[sheet_name]

                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in sheet.iter_rows(values_only=True):
                    new_ws.append(row)

                num_rows = new_ws.max_row
                num_cols = new_ws.max_column

                if num_rows < 2 or num_cols < 2:
                    st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                    continue

                timestamp_col = 1
                
                config_result = get_device_config_for_slave(slave_id)
                if config_result is None:
                    print(f"Skipping chart generation for Slave_{slave_id} in split function due to missing device configuration.")
                    continue

                device_config, _ = config_result 

                variables_to_plot_for_slave = device_config["plot_variables"]

                for idx, var_name in enumerate(variables_to_plot_for_slave):
                    if var_name in [cell.value for cell in new_ws[1]]:
                        var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                        create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                    else:
                        st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

                output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
                output_filepath = os.path.join(name_output_folder, output_filename)

                new_wb.save(output_filepath)
                generated_files.append(output_filepath)
                
                if progress_bar and total_slaves > 0:
                    progress_percent = int(((i + 1) / total_slaves) * 100)
                    progress_bar.progress(progress_percent)

            if not generated_files:
                st.warning("No individual files were generated for the selected slaves.")
                return None

            name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
            complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

            with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_to_add in generated_files:
                    zipf.write(file_to_add, os.path.basename(file_to_add))

            st.success(f"All files were zipped into '{complete_zip_path}'")
            return complete_zip_path

        except Exception as e:
            st.error(f"Error splitting and generating charts: {e}")
            return None

    def extract_slave_id(sheet_name):
        match = re.search(r'\d+', sheet_name)
        return int(match.group(0)) if match else None

    def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
            st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
            return None

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        else:
            for file_name in os.listdir(output_folder):
                path_to_remove = os.path.join(output_folder, file_name)
                if os.path.isfile(path_to_remove):
                    os.remove(path_to_remove)
            st.info(f"Folder '{output_folder}' cleaned.")

        wb_modbus = openpyxl.load_workbook(path_modbus_log)
        wb_backup = openpyxl.load_workbook(path_backup_data)
        
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in wb_modbus.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue
            if sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
                continue

            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping merge/chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

            modbus_sheet = wb_modbus[sheet_name]
            backup_sheet = wb_backup[sheet_name]

            modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
            backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

            modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
            backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

            merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
            merged_data.dropna(subset=["timestamp"], inplace=True)
            merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
            merged_data.sort_values("timestamp", inplace=True)
            merged_data.reset_index(drop=True, inplace=True)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            new_sheet = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in modbus_sheet[1]]
            new_sheet.append(headers)

            for r_idx, row in merged_data.iterrows():
                new_sheet.append(list(row.values))

            num_rows = new_sheet.max_row
            current_slave_id = extract_slave_id(sheet_name)
            
            variables_to_plot_for_slave = device_config["plot_variables"]

            for chart_index, var_name in enumerate(variables_to_plot_for_slave):
                if var_name in headers:
                    col_idx = headers.index(var_name) + 1
                    create_scatter_chart(
                        new_sheet,
                        timestamp_col=1,
                        var_col=col_idx,
                        num_rows=num_rows,
                        var_name=var_name,
                        slave_id=current_slave_id,
                        chart_index=chart_index
                    )
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
            new_wb.save(path_output)
            generated_files.append(path_output)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All merged individual files have been zipped into: {zip_path}")
        return zip_path

    def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
        if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
            st.error("Excel files (final or backup) not found for comparison.")
            return False

        wb_final = load_workbook(excel_file_final, data_only=True)
        wb_backup = load_workbook(excel_file_backup, data_only=True)

        needs_merge = False
        for slave_id in selected_slave_ids:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
                continue

            ws_final = wb_final[sheet_name]
            ws_backup = wb_backup[sheet_name]

            timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
            timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

            if not timestamps_backup.issubset(timestamps_final):
                needs_merge = True
                break
        
        return needs_merge


    # ----------------------------------------- STREAMLIT DASHBOARD
    with st.container():
        st.title("Long Term Test Setup for SCU100")
        
        if f'{state_prefix}polling' not in st.session_state:
            st.session_state[f'{state_prefix}last_valid_poll_data'] = {}
            st.session_state[f'{state_prefix}polling'] = False
            st.session_state[f'{state_prefix}stop_event'] = threading.Event()
            st.session_state[f'{state_prefix}thread'] = None
            st.session_state[f'{state_prefix}all_poll_data'] = []
            
            st.session_state[f'{state_prefix}selected_slaves_by_type'] = {
                device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
            }
            st.session_state[f'{state_prefix}realtime_table_placeholders'] = {}
            st.session_state[f'{state_prefix}last_data_length'] = 0
            st.session_state[f'{state_prefix}last_ui_update_time'] = time.time()
            st.session_state[f'{state_prefix}last_table_update_timestamp'] = "N/A"

        if st.button("Start Monitoring", key=f"start_monitoring_{state_prefix}", disabled=st.session_state[f'{state_prefix}polling']):
            if not st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}all_poll_data'] = []
                st.session_state[f'{state_prefix}stop_event'].clear()

                st.session_state[f'{state_prefix}thread'] = threading.Thread(
                    target=periodic_polling_loop,
                    args=(
                        st.session_state[f'{state_prefix}stop_event'],
                        st.session_state[f'{state_prefix}all_poll_data'],
                        state_prefix
                    ),
                    daemon=True
                )

                st.session_state[f'{state_prefix}thread'].start()
                st.session_state[f'{state_prefix}polling'] = True
                st.session_state[f'{state_prefix}last_data_length'] = 0
                st.success("Monitoring started!")

        if st.button("Stop Monitoring", key=f"stop_monitoring_{state_prefix}"):
            if st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}stop_event'].set()
                if st.session_state[f'{state_prefix}thread'] and st.session_state[f'{state_prefix}thread'].is_alive():
                    st.session_state[f'{state_prefix}thread'].join(timeout=10)
                    if st.session_state[f'{state_prefix}thread'].is_alive():
                        st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
                    else:
                        st.info("Polling thread terminated.")
                st.session_state[f'{state_prefix}polling'] = False

                with st.spinner("Saving data to Excel..."):
                    save_all_data_to_excel(st.session_state[f'{state_prefix}all_poll_data'], filename=EXCEL_FILE_FINAL)

                    try:
                        add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                        st.success("Monitoring stopped, data and charts saved to Excel.")
                    except Exception as e:
                        st.error(f"Error adding charts: {e}")
            else:
                st.warning("Monitoring is not running.")

        st.divider()

        if st.session_state.get(f"{state_prefix}all_poll_data"):
            all_poll_data = st.session_state[f"{state_prefix}all_poll_data"]

            st.markdown(
                f"Total of **{int(len(all_poll_data))}** reading(s) completed so far."
            )

            device_counts = {}

            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                key = assignment["device_type_key"]
                normalized_key = re.match(r'^[A-Za-z]+', key)
                normalized_key = normalized_key.group(0) if normalized_key else key
                if "ips" in assignment:
                    count = len(assignment["ips"])
                elif "slaves" in assignment:
                    count = len(assignment["slaves"])
                else:
                    count = 0
                device_counts[key] = device_counts.get(key, 0) + count

            device_count_lines = [f"- **{k}**: {v} device(s)" for k, v in device_counts.items()]
            st.markdown("### Active devices on the system:")
            st.markdown("\n".join(device_count_lines))

            last_poll_data = all_poll_data[-1]
            if "elapsed" not in last_poll_data and len(all_poll_data) >= 2:
                last_poll_data = all_poll_data[-2]

            last_poll_time = last_poll_data.get("elapsed", None)
            last_sleep_time = last_poll_data.get("sleep_time", None)

            if last_poll_time is not None and last_sleep_time is not None:
                st.markdown(
                    f"Completed one polling in **{last_poll_time:.2f}s**, sleeping for **{last_sleep_time:.2f}s**."
                )

            if last_poll_data.get("cycle_completed"):
                cycle_duration = last_poll_data.get("cycle_duration", 0)
                st.markdown(
                    f"Waiting **{int(cycle_duration / 60)} minutes** before next reading cycle."
                )

        st.subheader("Realtime data monitoring by sensor type")

        st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'] = st.empty()

        assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

        for device_type_key in assigned_device_types_for_display:
            device_config = DEVICE_TYPE.get(device_type_key)
            if not device_config:
                continue

            has_slaves_assigned = any(
                assignment["device_type_key"] == device_type_key
                for assignment in SLAVE_DEVICE_ASSIGNMENTS
            )
            if not has_slaves_assigned:
                continue

            available_slaves_for_type = []
            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                if assignment["device_type_key"] == device_type_key:
                    if isinstance(assignment["slaves"], range):
                        available_slaves_for_type.extend(list(assignment["slaves"]))
                    elif isinstance(assignment["slaves"], list):
                        available_slaves_for_type.extend(assignment["slaves"])
            available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

            if not available_slaves_for_type:
                continue

            with st.expander(f"Configuration and Monitoring: **{device_type_key}**", expanded=True):
                selected = st.multiselect(
                    f"Select Slave IDs for {device_type_key}:",
                    options=available_slaves_for_type,
                    default=st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, []),
                    key=f"{state_prefix}multiselect_{device_type_key}"
                )

                if st.button(f"Confirm selection for {device_type_key}", key=f"{state_prefix}confirm_button_{device_type_key}"):
                    st.session_state[f'{state_prefix}selected_slaves_by_type'][device_type_key] = selected

                st.session_state[f'{state_prefix}realtime_table_placeholders'][device_type_key] = st.empty()

        def update_realtime_tables_by_type():
            current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

            st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'].write(
                f"Last widget update at **{current_ui_update_timestamp}**"
            )

            if st.session_state[f'{state_prefix}all_poll_data']:
                all_poll_data = st.session_state[f'{state_prefix}all_poll_data']
                latest_poll_entry = all_poll_data[-1]

                if latest_poll_entry.get("data"):
                    raw_data = latest_poll_entry['data']
                    st.session_state[f'{state_prefix}last_valid_poll_data'] = latest_poll_entry
                else:
                    raw_data = st.session_state.get(f'{state_prefix}last_valid_poll_data', {}).get('data', {})

                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    selected_slaves = st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, [])
                    device_config = DEVICE_TYPE.get(device_type_key)

                    placeholder.empty()

                    if not selected_slaves or not device_config:
                        placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                        continue

                    table_data = []
                    variable_names = [var[2] for var in device_config["variables"]]

                    for slave_id in sorted(selected_slaves):
                        if slave_id in raw_data:
                            values = raw_data[slave_id]
                            row_dict = {"Slave ID": f"Slave {slave_id}"}
                            for i, var_name in enumerate(variable_names):
                                if i < len(values) and values[i] is not None:
                                    display_value = values[i]
                                    if isinstance(display_value, float):
                                        display_value = round(display_value, 2)
                                else:
                                    display_value = np.nan
                                row_dict[var_name] = display_value
                            table_data.append(row_dict)
                        else:
                            table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                    if table_data:
                        df = pd.DataFrame(table_data)
                        df = df.set_index("Slave ID")

                        placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                    else:
                        placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
            else:
                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    placeholder.empty()
                    placeholder.info(f"Waiting for first polling data for {device_type_key}...")

        update_realtime_tables_by_type()

        if st.session_state[f'{state_prefix}polling']:
            st_autorefresh(interval=5000, key=f"realtime_autorefresh_{state_prefix}")

        if not os.path.exists(EXCEL_FILE_FINAL):
            if not st.session_state[f'{state_prefix}polling']:
                st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
        else:
            st.divider()
            all_possible_slave_ids = sorted(list(set(SLAVE_IDS)))

            display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

            if "All" in st.session_state.get('selected_display_options', []):
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=["All"],
                    disabled=True
                )
            else:
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=[]
                )

            st.session_state.selected_display_options = selected_display_options

            selected_slave_ids_numeric = []
            if "All" in selected_display_options:
                selected_slave_ids_numeric = all_possible_slave_ids
            elif selected_display_options:
                for ds_id in selected_display_options:
                    match = re.search(r'\d+', ds_id)
                    if match:
                        selected_slave_ids_numeric.append(int(match.group(0)))

            if st.button("Generate Individual Files", key=f"generate_individual_files_{state_prefix}"):
                if not selected_slave_ids_numeric:
                    st.warning("Please select at least one slave or 'All'.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    total_slaves_to_process = len(selected_slave_ids_numeric)
                    zip_path = None

                    if not os.path.exists(EXCEL_FILE_BACKUP):
                        st.warning("Backup file not found. Performing only the split of the final Excel.")
                        with st.spinner("Splitting Excel and generating charts..."):
                            zip_path = split_and_generate_charts_for_slaves(
                                EXCEL_FILE_FINAL,
                                selected_slave_ids_numeric,
                                progress_bar=progress_bar,
                                status_text=status_text,
                                total_slaves=total_slaves_to_process
                            )
                    else:
                        st.info("Comparing timestamps between final Excel and backup...")
                        needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                        if needs_merge:
                            st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                            with st.spinner("Merging data from backup and final, and generating individual files..."):
                                zip_path = merge_data_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    EXCEL_FILE_BACKUP,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )
                        else:
                            st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                            with st.spinner("Splitting Excel and generating charts..."):
                                zip_path = split_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )

                    if zip_path:
                        status_text.text("Processing complete!")
                        st.success("Individual files generated and available for download.")



# -------------------------
# SCRIPT 5: SCU200
# -------------------------
def run_SCU200():
    
    # ---------------------------------------- CONFIGURATION SECTION
    IP = '192.168.1.200'
    PORT = 502
    SLAVE_IDS = list(range(1, 164))
    POLL_INTERVAL = 60
    EXCEL_FILE_FINAL = "SCU200_modbus_poll_log.xlsx"
    EXCEL_FILE_BACKUP = "SCU200_backup_poll_data.xlsx"
    state_prefix = "SCU200_"

    DEVICE_TYPE = {
        "CurrentSensor": {
            "variables": [
                (9, 1, "currentTrms", 3),
                (10, 1, "currentAc", 3),
                (11, 1, "currentDc", 3),
                (100, 2, "activePowerTotal", 3),
                (102, 2, "activeEnergyTotal", 3)
            ],
            
            "scale_factors": {
                "currentTrms": 0.01,
                "currentAc": 0.01,
                "currentDc": 0.01,
                "activePowerTotal": 1,
                "activeEnergyTotal": 1
            },
            
            "decoding_map": {
                "currentTrms": "decode_16bit_unsigned",
                "currentAc": "decode_16bit_unsigned",
                "currentDc": "decode_16bit_signed",
                "activePowerTotal": "decode_32bit_signed",
                "activeEnergyTotal": "decode_32bit_signed"
            },
            
            "plot_variables": [
                "currentTrms", 
                "currentAc", 
                "currentDc", 
                "activePowerTotal", 
                "activeEnergyTotal"],
            
            "variable_units": {
                "currentTrms": "(A)",
                "currentAc": "(A)",
                "currentDc": "(A)",
                "activePowerTotal": "(W)",
                "activeEnergyTotal": "(kWh)"
            }
        },

        "M4M": {
            "variables": [
                (20480, 4, "activeImportEnergyTotal", 3),
                (20484, 4, "activeExportEnergyTotal", 3),
                (20488, 4, "activeNetEnergyTotal", 3),
                (20492, 4, "reactiveImportEnergyTotal", 3),
                (20496, 4, "reactiveExportEnergyTotal", 3),
                (20500, 4, "reactiveNetEnergyTotal", 3),
                (20504, 4, "apparentImportEnergyTotal", 3),
                (20508, 4, "apparentExportEnergyTotal", 3),
                (20512, 4, "apparentNetEnergyTotal", 3),
                (21600, 4, "activeImpEnergyL1", 3),
                (21604, 4, "activeImpEnergyL2", 3),
                (21608, 4, "activeImpEnergyL3", 3),
                (21612, 4, "activeExpEnergyL1", 3),
                (21616, 4, "activeExpEnergyL2", 3),
                (21620, 4, "activeExpEnergyL3", 3),
                (21624, 4, "activeNetEnergyL1", 3),
                (21628, 4, "activeNetEnergyL2", 3),
                (21632, 4, "activeNetEnergyL3", 3),
                (21636, 4, "reactiveImpEnergyL1", 3),
                (21640, 4, "reactiveImpEnergyL2", 3),
                (21644, 4, "reactiveImpEnergyL3", 3),
                (21648, 4, "reactiveExpEnergyL1", 3),
                (21652, 4, "reactiveExpEnergyL2", 3),
                (21656, 4, "reactiveExpEnergyL3", 3),
                (21660, 4, "reactiveNetEnergyL1", 3),
                (21664, 4, "reactiveNetEnergyL2", 3),
                (21668, 4, "reactiveNetEnergyL3", 3),
                (21672, 4, "apparentImpEnergyL1", 3),
                (21676, 4, "apparentImpEnergyL2", 3),
                (21680, 4, "apparentImpEnergyL3", 3),
                (21684, 4, "apparentExpEnergyL1", 3),
                (21688, 4, "apparentExpEnergyL2", 3),
                (21692, 4, "apparentExpEnergyL3", 3),
                (21696, 4, "apparentNetEnergyL1", 3),
                (21700, 4, "apparentNetEnergyL2", 3),
                (21704, 4, "apparentNetEnergyL3", 3),
                (23312, 2, "currentL1", 3),
                (23314, 2, "currentL2", 3),
                (23316, 2, "currentL3", 3),
                (23318, 2, "currentN", 3),
                (23322, 2, "activePowerTotal", 3),
                (23324, 2, "activePowerL1", 3),
                (23326, 2, "activePowerL2", 3),
                (23328, 2, "activePowerL3", 3),
                (23330, 2, "reactivePowerTotal", 3),
                (23332, 2, "reactivePowerL1", 3),
                (23334, 2, "reactivePowerL2", 3),
                (23336, 2, "reactivePowerL3", 3),
                (23338, 2, "apparentPowerTotal", 3),
                (23340, 2, "apparentPowerL1", 3),
                (23342, 2, "apparentPowerL2", 3),
                (23344, 2, "apparentPowerL3", 3)
            ],
            
            "scale_factors": {
                "activeImportEnergyTotal": 0.01,
                "activeExportEnergyTotal": 0.01,
                "activeNetEnergyTotal": 0.01,
                "reactiveImportEnergyTotal": 0.01,
                "reactiveExportEnergyTotal": 0.01,
                "reactiveNetEnergyTotal": 0.01,
                "apparentImportEnergyTotal": 0.01,
                "apparentExportEnergyTotal": 0.01,
                "apparentNetEnergyTotal": 0.01,
                "activeImpEnergyL1": 0.01,
                "activeImpEnergyL2": 0.01,
                "activeImpEnergyL3": 0.01,
                "activeExpEnergyL1": 0.01,
                "activeExpEnergyL2": 0.01,
                "activeExpEnergyL3": 0.01,
                "activeNetEnergyL1": 0.01,
                "activeNetEnergyL2": 0.01,
                "activeNetEnergyL3": 0.01,
                "reactiveImpEnergyL1": 0.01,
                "reactiveImpEnergyL2": 0.01,
                "reactiveImpEnergyL3": 0.01,
                "reactiveExpEnergyL1": 0.01,
                "reactiveExpEnergyL2": 0.01,
                "reactiveExpEnergyL3": 0.01,
                "reactiveNetEnergyL1": 0.01,
                "reactiveNetEnergyL2": 0.01,
                "reactiveNetEnergyL3": 0.01,
                "apparentImpEnergyL1": 0.01,
                "apparentImpEnergyL2": 0.01,
                "apparentImpEnergyL3": 0.01,
                "apparentExpEnergyL1": 0.01,
                "apparentExpEnergyL2": 0.01,
                "apparentExpEnergyL3": 0.01,
                "apparentNetEnergyL1": 0.01,
                "apparentNetEnergyL2": 0.01,
                "apparentNetEnergyL3": 0.01,
                "currentL1": 0.01,
                "currentL2": 0.01,
                "currentL3": 0.01,
                "currentN": 0.01,
                "activePowerTotal": 0.01,
                "activePowerL1": 0.01,
                "activePowerL2": 0.01,
                "activePowerL3": 0.01,
                "reactivePowerTotal": 0.01,
                "reactivePowerL1": 0.01,
                "reactivePowerL2": 0.01,
                "reactivePowerL3": 0.01,
                "apparentPowerTotal": 0.01,
                "apparentPowerL1": 0.01,
                "apparentPowerL2": 0.01,
                "apparentPowerL3": 0.01
            },
            
            "decoding_map": {
                "activeImportEnergyTotal": "decode_64bit_unsigned",
                "activeExportEnergyTotal": "decode_64bit_unsigned",
                "activeNetEnergyTotal": "decode_64bit_signed",
                "reactiveImportEnergyTotal": "decode_64bit_unsigned",
                "reactiveExportEnergyTotal": "decode_64bit_unsigned",
                "reactiveNetEnergyTotal": "decode_64bit_signed",
                "apparentImportEnergyTotal": "decode_64bit_unsigned",
                "apparentExportEnergyTotal": "decode_64bit_unsigned",
                "apparentNetEnergyTotal": "decode_64bit_signed",
                "activeImpEnergyL1": "decode_64bit_unsigned",
                "activeImpEnergyL2": "decode_64bit_unsigned",
                "activeImpEnergyL3": "decode_64bit_unsigned",
                "activeExpEnergyL1": "decode_64bit_unsigned",
                "activeExpEnergyL2": "decode_64bit_unsigned",
                "activeExpEnergyL3": "decode_64bit_unsigned",
                "activeNetEnergyL1": "decode_64bit_signed",
                "activeNetEnergyL2": "decode_64bit_signed",
                "activeNetEnergyL3": "decode_64bit_signed",
                "reactiveImpEnergyL1": "decode_64bit_unsigned",
                "reactiveImpEnergyL2": "decode_64bit_unsigned",
                "reactiveImpEnergyL3": "decode_64bit_unsigned",
                "reactiveExpEnergyL1": "decode_64bit_unsigned",
                "reactiveExpEnergyL2": "decode_64bit_unsigned",
                "reactiveExpEnergyL3": "decode_64bit_unsigned",
                "reactiveNetEnergyL1": "decode_64bit_signed",
                "reactiveNetEnergyL2": "decode_64bit_signed",
                "reactiveNetEnergyL3": "decode_64bit_signed",
                "apparentImpEnergyL1": "decode_64bit_unsigned",
                "apparentImpEnergyL2": "decode_64bit_unsigned",
                "apparentImpEnergyL3": "decode_64bit_unsigned",
                "apparentExpEnergyL1": "decode_64bit_unsigned",
                "apparentExpEnergyL2": "decode_64bit_unsigned",
                "apparentExpEnergyL3": "decode_64bit_unsigned",
                "apparentNetEnergyL1": "decode_64bit_signed",
                "apparentNetEnergyL2": "decode_64bit_signed",
                "apparentNetEnergyL3": "decode_64bit_signed",
                "currentL1": "decode_32bit_unsigned",
                "currentL2": "decode_32bit_unsigned",
                "currentL3": "decode_32bit_unsigned",
                "currentN": "decode_32bit_unsigned",
                "activePowerTotal": "decode_32bit_signed",
                "activePowerL1": "decode_32bit_signed",
                "activePowerL2": "decode_32bit_signed",
                "activePowerL3": "decode_32bit_signed",
                "reactivePowerTotal": "decode_32bit_signed",
                "reactivePowerL1": "decode_32bit_signed",
                "reactivePowerL2": "decode_32bit_signed",
                "reactivePowerL3": "decode_32bit_signed",
                "apparentPowerTotal": "decode_32bit_signed",
                "apparentPowerL1": "decode_32bit_signed",
                "apparentPowerL2": "decode_32bit_signed",
                "apparentPowerL3": "decode_32bit_signed"
            },
            
            "plot_variables": [
                "activeImportEnergyTotal",
                "activeExportEnergyTotal",
                "activeNetEnergyTotal",
                "reactiveImportEnergyTotal",
                "reactiveExportEnergyTotal",
                "reactiveNetEnergyTotal",
                "apparentImportEnergyTotal",
                "apparentExportEnergyTotal",
                "apparentNetEnergyTotal",
                "activeImpEnergyL1",
                "activeImpEnergyL2",
                "activeImpEnergyL3",
                "activeExpEnergyL1",
                "activeExpEnergyL2",
                "activeExpEnergyL3",
                "activeNetEnergyL1",
                "activeNetEnergyL2",
                "activeNetEnergyL3",
                "reactiveImpEnergyL1",
                "reactiveImpEnergyL2",
                "reactiveImpEnergyL3",
                "reactiveExpEnergyL1",
                "reactiveExpEnergyL2",
                "reactiveExpEnergyL3",
                "reactiveNetEnergyL1",
                "reactiveNetEnergyL2",
                "reactiveNetEnergyL3",
                "apparentImpEnergyL1",
                "apparentImpEnergyL2",
                "apparentImpEnergyL3",
                "apparentExpEnergyL1",
                "apparentExpEnergyL2",
                "apparentExpEnergyL3",
                "apparentNetEnergyL1",
                "apparentNetEnergyL2",
                "apparentNetEnergyL3",
                "currentL1",
                "currentL2",
                "currentL3",
                "currentN",
                "activePowerTotal",
                "activePowerL1",
                "activePowerL2",
                "activePowerL3",
                "reactivePowerTotal",
                "reactivePowerL1",
                "reactivePowerL2",
                "reactivePowerL3",
                "apparentPowerTotal",
                "apparentPowerL1",
                "apparentPowerL2",
                "apparentPowerL3"
            ],
            
            "variable_units": {
                "activeImportEnergyTotal": "(kWh)",
                "activeExportEnergyTotal": "(kWh)",
                "activeNetEnergyTotal": "(kWh)",
                "reactiveImportEnergyTotal": "(kVarh)",
                "reactiveExportEnergyTotal": "(kVarh)",
                "reactiveNetEnergyTotal": "(kVarh)",
                "apparentImportEnergyTotal": "(kVAh)",
                "apparentExportEnergyTotal": "(kVAh)",
                "apparentNetEnergyTotal": "(kVAh)",
                "activeImpEnergyL1": "(kWh)",
                "activeImpEnergyL2": "(kWh)",
                "activeImpEnergyL3": "(kWh)",
                "activeExpEnergyL1": "(kWh)",
                "activeExpEnergyL2": "(kWh)",
                "activeExpEnergyL3": "(kWh)",
                "activeNetEnergyL1": "(kWh)",
                "activeNetEnergyL2": "(kWh)",
                "activeNetEnergyL3": "(kWh)",
                "reactiveImpEnergyL1": "(kVarh)",
                "reactiveImpEnergyL2": "(kVarh)",
                "reactiveImpEnergyL3": "(kVarh)",
                "reactiveExpEnergyL1": "(kVarh)",
                "reactiveExpEnergyL2": "(kVarh)",
                "reactiveExpEnergyL3": "(kVarh)",
                "reactiveNetEnergyL1": "(kVarh)",
                "reactiveNetEnergyL2": "(kVarh)",
                "reactiveNetEnergyL3": "(kVarh)",
                "apparentImpEnergyL1": "(kVAh)",
                "apparentImpEnergyL2": "(kVAh)",
                "apparentImpEnergyL3": "(kVAh)",
                "apparentExpEnergyL1": "(kVAh)",
                "apparentExpEnergyL2": "(kVAh)",
                "apparentExpEnergyL3": "(kVAh)",
                "apparentNetEnergyL1": "(kVAh)",
                "apparentNetEnergyL2": "(kVAh)",
                "apparentNetEnergyL3": "(kVAh)",
                "currentL1": "(A)",
                "currentL2": "(A)",
                "currentL3": "(A)",
                "currentN": "(A)",
                "activePowerTotal": "(W)",
                "activePowerL1": "(W)",
                "activePowerL2": "(W)",
                "activePowerL3": "(W)",
                "reactivePowerTotal": "(var)",
                "reactivePowerL1": "(var)",
                "reactivePowerL2": "(var)",
                "reactivePowerL3": "(var)",
                "apparentPowerTotal": "(VA)",
                "apparentPowerL1": "(VA)",
                "apparentPowerL2": "(VA)",
                "apparentPowerL3": "(VA)"
            }
        }
    }

    SLAVE_DEVICE_ASSIGNMENTS = [
        {"slaves": list(range(1, 161)), "device_type_key": "CurrentSensor"},
        {"slaves": list(range(161, 164)), "device_type_key": "M4M"}
        ]

    if SLAVE_DEVICE_ASSIGNMENTS:
        first_assigned_device_type_key = SLAVE_DEVICE_ASSIGNMENTS[0]["device_type_key"]
        if first_assigned_device_type_key in DEVICE_TYPE:
            VARIABLES = [(addr, size, name) for addr, size, name, *rest in DEVICE_TYPE[first_assigned_device_type_key]["variables"]]
        else:
            VARIABLES = []
    else:
        VARIABLES = []

    # --------------------------------------------- HELPER FUNCTIONS
    def get_device_config_for_slave(slave_id):
        for assignment in SLAVE_DEVICE_ASSIGNMENTS:
            if isinstance(assignment["slaves"], range):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
            elif isinstance(assignment["slaves"], list):
                if slave_id in assignment["slaves"]:
                    device_type_key = assignment["device_type_key"]
                    break
        else:
            print(f"Warning: Slave ID {slave_id} not assigned to any device type in SLAVE_DEVICE_ASSIGNMENTS. Skipping this slave.")
            return None

        if device_type_key not in DEVICE_TYPE:
            print(f"Error: Device type key '{device_type_key}' assigned to slave {slave_id} not found in DEVICE_TYPE configuration.")
            return None
            
        return DEVICE_TYPE[device_type_key], device_type_key



    # -------------------------------------- SENSOR POLLING FUNCTION
    def read_slave_with_client(slave_id, thread_id, client):
        device_config, device_type_key = get_device_config_for_slave(slave_id)
        if device_config is None:
            dummy_len = len(DEVICE_TYPE.get("CurrentSensor", {}).get("variables", []))
            return [None] * dummy_len

        variables = device_config["variables"]
        scale_factors = device_config["scale_factors"]
        decoding_map = device_config["decoding_map"]

        values = []
        thread_prefix = f"[Thread {thread_id}] "

        for var_info in variables:
            addr, size, name, func_code = var_info[:4]
            options = var_info[4] if len(var_info) > 4 else {}

            current_value = None
            try:
                resp = None
                if func_code == 1:
                    resp = client.read_coils(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'bits') or not resp.bits:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Coils read error or empty bits.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.bits
                elif func_code == 3:
                    resp = client.read_holding_registers(address=addr, count=size, slave=slave_id)
                    if resp.isError() or not hasattr(resp, 'registers') or not resp.registers:
                        print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Registers read error or empty registers.")
                        values.append(None)
                        continue
                    regs_or_bits = resp.registers
                    if func_code == 3 and size > 1:
                        regs_or_bits = list(reversed(regs_or_bits))
                else:
                    print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Unsupported function code {func_code}.")
                    values.append(None)
                    continue

                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Raw Data (Func {func_code}): {regs_or_bits}")

                if name in decoding_map:
                    func_name = decoding_map[name]
                    decoder_func = DECODING_FUNCTIONS_MAP.get(func_name)

                    if decoder_func:
                        if func_name == "decode_bit":
                            bit_pos = options.get('bit')
                            if bit_pos is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_pos)
                            else:
                                print(f"Error: Missing 'bit' in options for decode_bit of {name} or empty data.")
                        elif func_name == "decode_bits":
                            bit_range = options.get('bits')
                            if bit_range is not None and regs_or_bits:
                                current_value = decoder_func(regs_or_bits[0], bit_range)
                            else:
                                print(f"Error: Missing 'bits' in options for decode_bits of {name} or empty data.")
                        elif func_name == "read_coil":
                            current_value = decoder_func(resp)
                        elif func_name in ["decode_16bit_signed", "decode_16bit_unsigned", "decode_device_status"]:
                            current_value = decoder_func(regs_or_bits[0]) if regs_or_bits else None
                        else:
                            current_value = decoder_func(regs_or_bits)
                    else:
                        print(f"{thread_prefix}Warning: No decoder function found for '{func_name}' for variable '{name}'. Falling back to raw value processing.")
                        current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits
                else:
                    current_value = regs_or_bits[0] if size == 1 and regs_or_bits else regs_or_bits

                if name in scale_factors and isinstance(current_value, (int, float)):
                    current_value *= scale_factors[name]

                if isinstance(current_value, list):
                    final_value = flatten_value(current_value)
                else:
                    final_value = current_value

                values.append(final_value)

            except Exception as e:
                print(f"{thread_prefix}Slave {slave_id} | Addr {addr} | Exception: {e}")
                values.append(None)

        return values


    
    # ------------------------------------------------ SAVE TO EXCEL
    def save_all_data_to_excel(all_poll_data, filename=EXCEL_FILE_FINAL):
        wb = Workbook()
        wb.remove(wb.active)
        
        for slave_id in SLAVE_IDS:
            config_result = get_device_config_for_slave(slave_id)

            if config_result is None:
                print(f"Skipping sheet creation for Slave_{slave_id} due to missing config.")
                continue

            device_config, _ = config_result

            if not isinstance(device_config, dict):
                print(f"Skipping sheet creation for Slave_{slave_id} due to invalid device config type.")
                continue

            variables = device_config.get("variables", [])

            headers = ['timestamp']
            for var in variables:
                if isinstance(var, dict):
                    headers.append(var.get("name", "unknown_var"))
                elif isinstance(var, (list, tuple)):
                    headers.append(var[2] if len(var) > 2 else "unknown_var")
                else:
                    headers.append("unknown_var")

            ws = wb.create_sheet(title=f"Slave_{slave_id}")
            ws.append(headers)

        for poll_entry in all_poll_data:
            timestamp = poll_entry['timestamp']
            data = poll_entry['data']
            
            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

        wb.save(filename)

    def append_poll_data_to_backup(poll_entry, filename=EXCEL_FILE_BACKUP):
        timestamp = poll_entry['timestamp']
        data = poll_entry['data']

        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
            else:
                wb = Workbook()
                wb.remove(wb.active)
                for slave_id in SLAVE_IDS:
                    config_result = get_device_config_for_slave(slave_id)
                    if config_result is None:
                        continue
                    device_config, _ = config_result
                    variables = device_config.get("variables", [])
                    headers = ['timestamp']
                    for var in variables:
                        if isinstance(var, dict):
                            headers.append(var.get("name", "unknown_var"))
                        elif isinstance(var, (list, tuple)):
                            headers.append(var[2] if len(var) > 2 else "unknown_var")
                        else:
                            headers.append("unknown_var")
                    ws = wb.create_sheet(title=f"Slave_{slave_id}")
                    ws.append(headers)

            for slave_id, values in data.items():
                sheet_name = f"Slave_{slave_id}"
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                safe_values = [clean_value(v) for v in values]
                ws.append([timestamp] + safe_values)

            wb.save(filename)

        except Exception as e:
            print(f"Error appending to backup: {e}")

    def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
        config_result = get_device_config_for_slave(slave_id)
        if config_result is None:
            print(f"Warning: No device configuration found for slave {slave_id}. Cannot create chart for {var_name}.")
            return
        
        device_config, _ = config_result 

        variable_units = device_config.get("variable_units", {})
        unit = variable_units.get(var_name, "")

        chart = ScatterChart()
        chart.title = f"{var_name} over time - ID {slave_id}"
        chart.style = 13
        chart.x_axis.title = "Timestamp"
        chart.y_axis.title = f"{var_name} {unit}".strip()
        chart.legend = None

        chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.x_axis.majorUnit = 500

        x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
        y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

        title_with_unit = f"{var_name} {unit}".strip()
        series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.smooth = True
        chart.series.append(series)

        chart.x_axis.axId = 10 + chart_index * 2
        chart.y_axis.axId = 20 + chart_index * 2
        chart.y_axis.crossAx = 10 + chart_index * 2
        chart.x_axis.crossAx = 20 + chart_index * 2

        chart.height = 9.5
        chart.width = 20.44

        anchor_col = 'BC'
        anchor_row = 1 + 20 * chart_index
        chart.anchor = f"{anchor_col}{anchor_row}"

        sheet.add_chart(chart)

    def add_charts_to_excel(filename=EXCEL_FILE_FINAL):
        wb = load_workbook(filename)

        for slave_id in SLAVE_IDS:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            num_rows = ws.max_row
            
            config_result = get_device_config_for_slave(slave_id)
            if config_result is None:
                print(f"Skipping chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            device_config, _ = config_result

            variables_to_plot_for_slave = device_config.get("plot_variables", [])

            chart_index = 0
            timestamp_col = 1

            for var in variables_to_plot_for_slave:
                if var not in headers:
                    continue

                var_col = headers.index(var) + 1
                create_scatter_chart(ws, timestamp_col, var_col, num_rows, var, slave_id, chart_index)
                chart_index += 1

        wb.save(filename)



    # ------------------------------------------------ POLLING LOOP
    def poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix):
        if stop_event.is_set():
            return
    
        start_time = time.time()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        all_data = {}
        lock = threading.Lock()
    
        def poll_slave_ids(slave_ids, all_data, lock, thread_id):
            client = ModbusTcpClient(IP, port=PORT)
            if not client.connect():
                print(f"[Thread {thread_id}] Failed to connect")
                return
    
            def task(slave_id):
                values = read_slave_with_client(slave_id, thread_id, client)
                with lock:
                    all_data[slave_id] = values
    
            with ThreadPoolExecutor(max_workers=len(slave_ids)) as executor:
                executor.map(task, slave_ids)
    
            client.close()
    
        slave_ids = SLAVE_IDS
        t = threading.Thread(target=poll_slave_ids, args=(slave_ids, all_data, lock, 1))
        t.start()
        t.join()
    
        elapsed = time.time() - start_time
        sleep_time = max(0, POLL_INTERVAL - elapsed)
        all_poll_data.append({
            'timestamp': timestamp,
            'data': all_data,
        })

        print(f"\n Completed one polling in {elapsed:.2f}s, sleeping for {sleep_time:.2f}s\n")
        time.sleep(sleep_time)

        st.session_state[f'{state_prefix}last_valid_poll_data'] = {
            'timestamp': timestamp,
            'data': all_data
        }

        try:
            append_poll_data_to_backup({
                'timestamp': timestamp,
                'data': all_data
            })
        except Exception as e:
            print(f"Error saving to backup file: {e}")

    def periodic_polling_loop(stop_event, all_poll_data, state_prefix):
        while not stop_event.is_set():
            start_time = time.time()
            one_cycle = 600
            cycle_duration = 60 * 60

            while time.time() - start_time < one_cycle and not stop_event.is_set():
                poll_slaves_multithread_once(stop_event, all_poll_data, state_prefix)

            if stop_event.is_set():
                break

            all_poll_data.append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': {},
            })

            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = True
            print(f"Waiting {cycle_duration / 60} minutes before next reading cycle.")
            
            for _ in range(int(cycle_duration)):
                if stop_event.is_set():
                    break
                time.sleep(1)
            st.session_state[f'{state_prefix}waiting_for_next_cycle'] = False



    # ---------------------------------------- SPLIT EXCEL FUNCTIONS
    def split_and_generate_charts_for_slaves(path_file_excel, selected_slave_ids, name_output_folder="single_slaves_with_generated_charts", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_file_excel):
            st.error(f"Error: File '{path_file_excel}' not found.")
            return None

        try:
            if not os.path.exists(name_output_folder):
                os.makedirs(name_output_folder)
            else:
                for file_name in os.listdir(name_output_folder):
                    path_old_file = os.path.join(name_output_folder, file_name)
                    if os.path.isfile(path_old_file):
                        os.remove(path_old_file)
                st.info(f"Cleaning folder '{name_output_folder}'.")

            source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
            generated_files = []

            for i, slave_id in enumerate(selected_slave_ids):
                sheet_name = f"Slave_{slave_id}"
                if status_text:
                    status_text.text(f"Processing Slave_{slave_id}...")
                if sheet_name not in source_workbook.sheetnames:
                    st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                    continue

                sheet = source_workbook[sheet_name]

                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                for row in sheet.iter_rows(values_only=True):
                    new_ws.append(row)

                num_rows = new_ws.max_row
                num_cols = new_ws.max_column

                if num_rows < 2 or num_cols < 2:
                    st.warning(f"Skipping sheet '{sheet_name}': insufficient data to generate charts.")
                    continue

                timestamp_col = 1
                
                config_result = get_device_config_for_slave(slave_id)
                if config_result is None:
                    print(f"Skipping chart generation for Slave_{slave_id} in split function due to missing device configuration.")
                    continue

                device_config, _ = config_result 

                variables_to_plot_for_slave = device_config["plot_variables"]

                for idx, var_name in enumerate(variables_to_plot_for_slave):
                    if var_name in [cell.value for cell in new_ws[1]]:
                        var_col = [cell.value for cell in new_ws[1]].index(var_name) + 1
                        create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)
                    else:
                        st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

                output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
                output_filepath = os.path.join(name_output_folder, output_filename)

                new_wb.save(output_filepath)
                generated_files.append(output_filepath)
                
                if progress_bar and total_slaves > 0:
                    progress_percent = int(((i + 1) / total_slaves) * 100)
                    progress_bar.progress(progress_percent)

            if not generated_files:
                st.warning("No individual files were generated for the selected slaves.")
                return None

            name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
            complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

            with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_to_add in generated_files:
                    zipf.write(file_to_add, os.path.basename(file_to_add))

            st.success(f"All files were zipped into '{complete_zip_path}'")
            return complete_zip_path

        except Exception as e:
            st.error(f"Error splitting and generating charts: {e}")
            return None

    def extract_slave_id(sheet_name):
        match = re.search(r'\d+', sheet_name)
        return int(match.group(0)) if match else None

    def merge_data_and_generate_charts_for_slaves(path_modbus_log, path_backup_data, selected_slave_ids, output_folder="merged_slaves", progress_bar=None, status_text=None, total_slaves=0):
        if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
            st.error("Error: One of the main files (modbus_poll_log.xlsx or backup_poll_data.xlsx) was not found.")
            return None

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        else:
            for file_name in os.listdir(output_folder):
                path_to_remove = os.path.join(output_folder, file_name)
                if os.path.isfile(path_to_remove):
                    os.remove(path_to_remove)
            st.info(f"Folder '{output_folder}' cleaned.")

        wb_modbus = openpyxl.load_workbook(path_modbus_log)
        wb_backup = openpyxl.load_workbook(path_backup_data)
        
        generated_files = []

        for i, slave_id in enumerate(selected_slave_ids):
            sheet_name = f"Slave_{slave_id}"
            if status_text:
                status_text.text(f"Processing Slave_{slave_id}...")
            if sheet_name not in wb_modbus.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the main file. Skipping.")
                continue
            if sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in the backup file. Skipping.")
                continue

            device_config = get_device_config_for_slave(slave_id)
            if device_config is None:
                print(f"Skipping merge/chart generation for Slave_{slave_id} due to missing device configuration.")
                continue
            
            columns_for_df = ["timestamp"] + [var[2] for var in device_config["variables"]]

            modbus_sheet = wb_modbus[sheet_name]
            backup_sheet = wb_backup[sheet_name]

            modbus_data_raw = modbus_sheet.iter_rows(min_row=2, values_only=True)
            backup_data_raw = backup_sheet.iter_rows(min_row=2, values_only=True)

            modbus_df = pd.DataFrame(modbus_data_raw, columns=columns_for_df)
            backup_df = pd.DataFrame(backup_data_raw, columns=columns_for_df)

            merged_data = pd.concat([backup_df, modbus_df], ignore_index=True)
            merged_data.dropna(subset=["timestamp"], inplace=True)
            merged_data.drop_duplicates(subset=["timestamp"], keep='first', inplace=True)
            merged_data.sort_values("timestamp", inplace=True)
            merged_data.reset_index(drop=True, inplace=True)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            new_sheet = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in modbus_sheet[1]]
            new_sheet.append(headers)

            for r_idx, row in merged_data.iterrows():
                new_sheet.append(list(row.values))

            num_rows = new_sheet.max_row
            current_slave_id = extract_slave_id(sheet_name)
            
            variables_to_plot_for_slave = device_config["plot_variables"]

            for chart_index, var_name in enumerate(variables_to_plot_for_slave):
                if var_name in headers:
                    col_idx = headers.index(var_name) + 1
                    create_scatter_chart(
                        new_sheet,
                        timestamp_col=1,
                        var_col=col_idx,
                        num_rows=num_rows,
                        var_name=var_name,
                        slave_id=current_slave_id,
                        chart_index=chart_index
                    )
                else:
                    st.warning(f"Variable '{var_name}' not found in sheet '{sheet_name}'. Chart will not be generated for this variable.")

            safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
            new_wb.save(path_output)
            generated_files.append(path_output)
            
            if progress_bar and total_slaves > 0:
                progress_percent = int(((i + 1) / total_slaves) * 100)
                progress_bar.progress(progress_percent)

        if not generated_files:
            st.warning("No individual files were generated for the selected slaves.")
            return None

        zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slave_files.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        st.success(f"All merged individual files have been zipped into: {zip_path}")
        return zip_path

    def compare_timestamps_and_decide_action(excel_file_final, excel_file_backup, selected_slave_ids):
        if not os.path.exists(excel_file_final) or not os.path.exists(excel_file_backup):
            st.error("Excel files (final or backup) not found for comparison.")
            return False

        wb_final = load_workbook(excel_file_final, data_only=True)
        wb_backup = load_workbook(excel_file_backup, data_only=True)

        needs_merge = False
        for slave_id in selected_slave_ids:
            sheet_name = f"Slave_{slave_id}"
            if sheet_name not in wb_final.sheetnames or sheet_name not in wb_backup.sheetnames:
                st.warning(f"Sheet '{sheet_name}' not found in both files for comparison. Skipping.")
                continue

            ws_final = wb_final[sheet_name]
            ws_backup = wb_backup[sheet_name]

            timestamps_final = set([cell.value for cell in ws_final['A'][1:] if cell.value is not None])
            timestamps_backup = set([cell.value for cell in ws_backup['A'][1:] if cell.value is not None])

            if not timestamps_backup.issubset(timestamps_final):
                needs_merge = True
                break
        
        return needs_merge


    # ----------------------------------------- STREAMLIT DASHBOARD
    with st.container():
        st.title("Long Term Test Setup for SCU200")

        if f'{state_prefix}polling' not in st.session_state:
            st.session_state[f'{state_prefix}last_valid_poll_data'] = {}
            st.session_state[f'{state_prefix}polling'] = False
            st.session_state[f'{state_prefix}stop_event'] = threading.Event()
            st.session_state[f'{state_prefix}thread'] = None
            st.session_state[f'{state_prefix}all_poll_data'] = []
            
            st.session_state[f'{state_prefix}selected_slaves_by_type'] = {
                device_type_key: [] for device_type_key in DEVICE_TYPE.keys()
            }
            st.session_state[f'{state_prefix}realtime_table_placeholders'] = {}
            st.session_state[f'{state_prefix}last_data_length'] = 0
            st.session_state[f'{state_prefix}last_ui_update_time'] = time.time()
            st.session_state[f'{state_prefix}last_table_update_timestamp'] = "N/A"

        if st.button("Start Monitoring", key=f"start_monitoring_{state_prefix}", disabled=st.session_state[f'{state_prefix}polling']):
            if not st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}all_poll_data'] = []
                st.session_state[f'{state_prefix}stop_event'].clear()

                st.session_state[f'{state_prefix}thread'] = threading.Thread(
                    target=periodic_polling_loop,
                    args=(
                        st.session_state[f'{state_prefix}stop_event'],
                        st.session_state[f'{state_prefix}all_poll_data'],
                        state_prefix
                    ),
                    daemon=True
                )

                st.session_state[f'{state_prefix}thread'].start()
                st.session_state[f'{state_prefix}polling'] = True
                st.session_state[f'{state_prefix}last_data_length'] = 0
                st.success("Monitoring started!")

        if st.button("Stop Monitoring", key=f"stop_monitoring_{state_prefix}"):
            if st.session_state[f'{state_prefix}polling']:
                st.session_state[f'{state_prefix}stop_event'].set()
                if st.session_state[f'{state_prefix}thread'] and st.session_state[f'{state_prefix}thread'].is_alive():
                    st.session_state[f'{state_prefix}thread'].join(timeout=10)
                    if st.session_state[f'{state_prefix}thread'].is_alive():
                        st.warning("Polling thread did not terminate gracefully. It might still be running in the background.")
                    else:
                        st.info("Polling thread terminated.")
                st.session_state[f'{state_prefix}polling'] = False

                with st.spinner("Saving data to Excel..."):
                    save_all_data_to_excel(st.session_state[f'{state_prefix}all_poll_data'], filename=EXCEL_FILE_FINAL)

                    try:
                        add_charts_to_excel(filename=EXCEL_FILE_FINAL)
                        st.success("Monitoring stopped, data and charts saved to Excel.")
                    except Exception as e:
                        st.error(f"Error adding charts: {e}")
            else:
                st.warning("Monitoring is not running.")

        st.divider()

        if st.session_state.get(f"{state_prefix}all_poll_data"):
            all_poll_data = st.session_state[f"{state_prefix}all_poll_data"]

            st.markdown(
                f"Total of **{int(len(all_poll_data))}** reading(s) completed so far."
            )

            device_counts = {}

            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                key = assignment["device_type_key"]
                normalized_key = re.match(r'^[A-Za-z]+', key)
                normalized_key = normalized_key.group(0) if normalized_key else key
                if "ips" in assignment:
                    count = len(assignment["ips"])
                elif "slaves" in assignment:
                    count = len(assignment["slaves"])
                else:
                    count = 0
                device_counts[key] = device_counts.get(key, 0) + count

            device_count_lines = [f"- **{k}**: {v} device(s)" for k, v in device_counts.items()]
            st.markdown("### Active devices on the system:")
            st.markdown("\n".join(device_count_lines))

            last_poll_data = all_poll_data[-1]
            if "elapsed" not in last_poll_data and len(all_poll_data) >= 2:
                last_poll_data = all_poll_data[-2]

            last_poll_time = last_poll_data.get("elapsed", None)
            last_sleep_time = last_poll_data.get("sleep_time", None)

            if last_poll_time is not None and last_sleep_time is not None:
                st.markdown(
                    f"Completed one polling in **{last_poll_time:.2f}s**, sleeping for **{last_sleep_time:.2f}s**."
                )

            if last_poll_data.get("cycle_completed"):
                cycle_duration = last_poll_data.get("cycle_duration", 0)
                st.markdown(
                    f"Waiting **{int(cycle_duration / 60)} minutes** before next reading cycle."
                )

        st.subheader("Realtime data monitoring by sensor type")

        st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'] = st.empty()

        assigned_device_types_for_display = sorted(list(DEVICE_TYPE.keys()))

        for device_type_key in assigned_device_types_for_display:
            device_config = DEVICE_TYPE.get(device_type_key)
            if not device_config:
                continue

            has_slaves_assigned = any(
                assignment["device_type_key"] == device_type_key
                for assignment in SLAVE_DEVICE_ASSIGNMENTS
            )
            if not has_slaves_assigned:
                continue

            available_slaves_for_type = []
            for assignment in SLAVE_DEVICE_ASSIGNMENTS:
                if assignment["device_type_key"] == device_type_key:
                    if isinstance(assignment["slaves"], range):
                        available_slaves_for_type.extend(list(assignment["slaves"]))
                    elif isinstance(assignment["slaves"], list):
                        available_slaves_for_type.extend(assignment["slaves"])
            available_slaves_for_type = sorted(list(set(available_slaves_for_type)))

            if not available_slaves_for_type:
                continue

            with st.expander(f"Configuration and Monitoring: **{device_type_key}**", expanded=True):
                selected = st.multiselect(
                    f"Select Slave IDs for {device_type_key}:",
                    options=available_slaves_for_type,
                    default=st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, []),
                    key=f"{state_prefix}multiselect_{device_type_key}"
                )

                if st.button(f"Confirm selection for {device_type_key}", key=f"{state_prefix}confirm_button_{device_type_key}"):
                    st.session_state[f'{state_prefix}selected_slaves_by_type'][device_type_key] = selected

                st.session_state[f'{state_prefix}realtime_table_placeholders'][device_type_key] = st.empty()

        def update_realtime_tables_by_type():
            current_ui_update_timestamp = time.strftime("%d/%m/%Y %H:%M:%S")

            st.session_state[f'{state_prefix}realtime_ui_timestamp_placeholder'].write(
                f"Last widget update at **{current_ui_update_timestamp}**"
            )

            if st.session_state[f'{state_prefix}all_poll_data']:
                all_poll_data = st.session_state[f'{state_prefix}all_poll_data']
                latest_poll_entry = all_poll_data[-1]

                if latest_poll_entry.get("data"):
                    raw_data = latest_poll_entry['data']
                    st.session_state[f'{state_prefix}last_valid_poll_data'] = latest_poll_entry
                else:
                    raw_data = st.session_state.get(f'{state_prefix}last_valid_poll_data', {}).get('data', {})

                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    selected_slaves = st.session_state[f'{state_prefix}selected_slaves_by_type'].get(device_type_key, [])
                    device_config = DEVICE_TYPE.get(device_type_key)

                    placeholder.empty()

                    if not selected_slaves or not device_config:
                        placeholder.info(f"No slaves selected or configuration not found for {device_type_key}.")
                        continue

                    table_data = []
                    variable_names = [var[2] for var in device_config["variables"]]

                    for slave_id in sorted(selected_slaves):
                        if slave_id in raw_data:
                            values = raw_data[slave_id]
                            row_dict = {"Slave ID": f"Slave {slave_id}"}
                            for i, var_name in enumerate(variable_names):
                                if i < len(values) and values[i] is not None:
                                    display_value = values[i]
                                    if isinstance(display_value, float):
                                        display_value = round(display_value, 2)
                                else:
                                    display_value = np.nan
                                row_dict[var_name] = display_value
                            table_data.append(row_dict)
                        else:
                            table_data.append({"Slave ID": f"Slave {slave_id}", "Status": "Data not available"})

                    if table_data:
                        df = pd.DataFrame(table_data)
                        df = df.set_index("Slave ID")

                        placeholder.dataframe(df, use_container_width=True, key=f"realtime_table_{device_type_key}_{time.time()}")
                    else:
                        placeholder.info(f"No real-time data available for the selected slaves of {device_type_key} yet.")
            else:
                for device_type_key, placeholder in st.session_state[f'{state_prefix}realtime_table_placeholders'].items():
                    placeholder.empty()
                    placeholder.info(f"Waiting for first polling data for {device_type_key}...")

        update_realtime_tables_by_type()

        if st.session_state[f'{state_prefix}polling']:
            st_autorefresh(interval=5000, key=f"realtime_autorefresh_{state_prefix}")

        if not os.path.exists(EXCEL_FILE_FINAL):
            if not st.session_state[f'{state_prefix}polling']:
                st.info("The final Excel file (modbus_poll_log.xlsx) does not exist. Start monitoring to create it.")
        else:
            st.divider()
            all_possible_slave_ids = sorted(list(set(SLAVE_IDS)))

            display_options = ["All"] + [f"Slave_{s_id}" for s_id in all_possible_slave_ids]

            if "All" in st.session_state.get('selected_display_options', []):
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=["All"],
                    disabled=True
                )
            else:
                selected_display_options = st.multiselect(
                    "Choose slaves to separate:",
                    display_options,
                    default=[]
                )

            st.session_state.selected_display_options = selected_display_options

            selected_slave_ids_numeric = []
            if "All" in selected_display_options:
                selected_slave_ids_numeric = all_possible_slave_ids
            elif selected_display_options:
                for ds_id in selected_display_options:
                    match = re.search(r'\d+', ds_id)
                    if match:
                        selected_slave_ids_numeric.append(int(match.group(0)))

            if st.button("Generate Individual Files", key=f"generate_individual_files_{state_prefix}"):
                if not selected_slave_ids_numeric:
                    st.warning("Please select at least one slave or 'All'.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    total_slaves_to_process = len(selected_slave_ids_numeric)
                    zip_path = None

                    if not os.path.exists(EXCEL_FILE_BACKUP):
                        st.warning("Backup file not found. Performing only the split of the final Excel.")
                        with st.spinner("Splitting Excel and generating charts..."):
                            zip_path = split_and_generate_charts_for_slaves(
                                EXCEL_FILE_FINAL,
                                selected_slave_ids_numeric,
                                progress_bar=progress_bar,
                                status_text=status_text,
                                total_slaves=total_slaves_to_process
                            )
                    else:
                        st.info("Comparing timestamps between final Excel and backup...")
                        needs_merge = compare_timestamps_and_decide_action(EXCEL_FILE_FINAL, EXCEL_FILE_BACKUP, selected_slave_ids_numeric)

                        if needs_merge:
                            st.warning("Discrepancies found! Merging and splitting Excel for selected Slaves.")
                            with st.spinner("Merging data from backup and final, and generating individual files..."):
                                zip_path = merge_data_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    EXCEL_FILE_BACKUP,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )
                        else:
                            st.info("Final and backup Excel files are identical in the timestamp column. Performing only the split of the final Excel.")
                            with st.spinner("Splitting Excel and generating charts..."):
                                zip_path = split_and_generate_charts_for_slaves(
                                    EXCEL_FILE_FINAL,
                                    selected_slave_ids_numeric,
                                    progress_bar=progress_bar,
                                    status_text=status_text,
                                    total_slaves=total_slaves_to_process
                                )

                    if zip_path:
                        status_text.text("Processing complete!")
                        st.success("Individual files generated and available for download.")



# -------------------------
# MAIN STREAMLIT DASHBOARD
# -------------------------
st.set_page_config(layout="wide")

st.sidebar.title("Select Monitoring Device")
available_scripts = ["Direct TCP Communication", "Direct RTU Communication", "ABB Ability™ Edge Industrial Gateway", "SCU100", "SCU200"]
selected_scripts = st.sidebar.multiselect(
    "Select each device to monitor:",
    options=available_scripts
)

if not selected_scripts:
    st.info("Please select one or more scripts from the sidebar to begin.")
else:
    function_map = {
        "Direct TCP Communication": run_tcp,
        "Direct RTU Communication": run_rtu,
        "ABB Ability™ Edge Industrial Gateway": run_gateway,
        "SCU100": run_SCU100,
        "SCU200": run_SCU200
    }
    
    tabs = st.tabs(selected_scripts)
    
    for i, script_name in enumerate(selected_scripts):
        with tabs[i]:
            function_to_run = function_map.get(script_name)
            if function_to_run:
                function_to_run()
