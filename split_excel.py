# Run on the folder: python split_excel.py
# Edit the line with the correct path and original file name

import os
import zipfile
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
    units = {
        "currentTrms": "(A)",
        "currentAc": "(A)",
        "currentDc": "(A)",
        "activePowerTotal": "(W)",
        "activeEnergyTotal": "(kWh)"
    }

    chart = ScatterChart()
    chart.title = f"{var_name} over time - ID {slave_id}"
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = f"{var_name} {units.get(var_name, '')}".strip()
    chart.legend = None

    chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    title_with_unit = f"{var_name} {units.get(var_name, '')}".strip()
    series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.smooth = True
    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'H'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)

def split_and_generate_charts(path_file_excel, name_output_folder="single_slaves_with_generated_charts"):
    if not os.path.exists(path_file_excel):
        print(f"Error: File '{path_file_excel}' not found.")
        return None

    try:
        if not os.path.exists(name_output_folder):
            os.makedirs(name_output_folder)
        else:
            for file_name in os.listdir(name_output_folder):
                path_old_file = os.path.join(name_output_folder, file_name)
                if os.path.isfile(path_old_file):
                    os.remove(path_old_file)
            print(f"Cleaning folder '{name_output_folder}'.")

        print(f"Loading Excel file '{path_file_excel}'...")

        source_workbook = openpyxl.load_workbook(path_file_excel, data_only=True)
        generated_files = []

        for sheet_name in source_workbook.sheetnames:
            sheet = source_workbook[sheet_name]

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            for row in sheet.iter_rows(values_only=True):
                new_ws.append(row)

            num_rows = new_ws.max_row
            num_cols = new_ws.max_column

            if num_rows < 2 or num_cols < 2:
                print(f"Skipping sheet '{sheet_name}': not enough data.")
                continue

            timestamp_col = 1
            slave_id = sheet_name.replace("Slave_", "")

            # Para cada coluna (variável) a partir da 2 até a última, gera um gráfico empilhado
            for idx, var_col in enumerate(range(2, num_cols + 1)):
                var_name = new_ws.cell(row=1, column=var_col).value or f"Variable_{var_col}"
                create_scatter_chart(new_ws, timestamp_col, var_col, num_rows, var_name, slave_id, idx)

            output_filename = f"{sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')}.xlsx"
            output_filepath = os.path.join(name_output_folder, output_filename)

            new_wb.save(output_filepath)
            generated_files.append(output_filepath)
            print(f"Sheet '{sheet_name}' saved in '{output_filepath}' with new charts.")

        # Cria ZIP
        name_zip_file = f"{os.path.basename(os.path.splitext(path_file_excel)[0])}_with_generated_charts.zip"
        complete_zip_path = os.path.join(os.path.dirname(path_file_excel), name_zip_file)

        with zipfile.ZipFile(complete_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_to_add in generated_files:
                zipf.write(file_to_add, os.path.basename(file_to_add))

        print(f"\nAll the files were zipped in '{complete_zip_path}'")
        return complete_zip_path

    except Exception as e:
        print(f"Error: {e}")
        return None

if __name__ == "__main__":
    my_file_path = r'C:\Users\ITGUSAP\Desktop\Monitoring\modbus_poll_log.xlsx'  # Edite conforme necessário
    try:
        zip_path = split_and_generate_charts(my_file_path)
        if zip_path:
            print(f"\nOperation finished. The ZIP file is in: {zip_path}")
        else:
            print("\nOperation failed.")
    except KeyboardInterrupt:
        print("\nProcess interrupted")