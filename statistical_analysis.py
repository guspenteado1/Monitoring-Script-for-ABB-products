import os
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties

def configure_chart_axes(chart):
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, chart_title, y_axis_label, chart_index, total_slave_files):
    chart = ScatterChart()
    chart.title = chart_title
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = y_axis_label
    chart.legend = None

    configure_chart_axes(chart)

    chart.y_axis.scaling.max = total_slave_files
    chart.y_axis.scaling.min = max(0, total_slave_files - 10)
    chart.y_axis.majorUnit = 1

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    series = Series(values=y_values, xvalues=x_values, title=var_name)
    series.marker.symbol = "circle"
    series.marker.size = 2
    series.smooth = True
    
    series.graphicalProperties.line.width = 12700

    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'H'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)


def calculate_missing_readings(input_folder, variables_to_track):
    readings_found_per_ts_per_var = defaultdict(lambda: defaultdict(int))
    all_timestamps = set()
    total_slave_files = 0

    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(input_folder, filename)
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                total_slave_files += 1

                header = [cell.value for cell in ws[1]]
                col_indices = {name: idx + 1 for idx, name in enumerate(header)}
                
                timestamp_col_name = "timestamp"
                if timestamp_col_name not in col_indices:
                    print(f"Warning: '{timestamp_col_name}' column not found in '{filename}'. "
                          f"Skipping this file for summary calculation.")
                    continue

                for row_idx in range(2, ws.max_row + 1):
                    timestamp_value = ws.cell(row=row_idx, column=col_indices[timestamp_col_name]).value
                    
                    if timestamp_value is None:
                        continue
                    
                    all_timestamps.add(timestamp_value)

                    for variable in variables_to_track:
                        if variable in col_indices:
                            cell_value = ws.cell(row=row_idx, column=col_indices[variable]).value
                            if cell_value is not None and \
                               not (isinstance(cell_value, str) and \
                                    cell_value.strip().lower() in ['', 'nan', 'none', 'null', '#n/a', 'na']):
                                readings_found_per_ts_per_var[timestamp_value][variable] += 1

            except Exception as e:
                print(f"Error processing file '{filename}': {e}")
                continue
    
    data_for_excel = []
    header_row = ['Timestamp'] + variables_to_track
    data_for_excel.append(header_row)

    sorted_timestamps = sorted(list(all_timestamps))

    for ts in sorted_timestamps:
        row_data = [ts]
        for var in variables_to_track:
            missing_count = total_slave_files - readings_found_per_ts_per_var[ts][var]
            row_data.append(missing_count)
        data_for_excel.append(row_data)
    
    return data_for_excel, total_slave_files


def generate_missing_counts_table_data(missing_data_for_excel, variables_to_track, max_missing_count=25):
    missing_counts_aggregation = defaultdict(lambda: defaultdict(int))
    
    if not missing_data_for_excel or len(missing_data_for_excel) < 2:
        return []

    header_row = missing_data_for_excel[0]
    
    var_col_map = {var: header_row.index(var) for var in variables_to_track}

    for row_data in missing_data_for_excel[1:]:
        for var in variables_to_track:
            col_idx = var_col_map[var]
            missing_val = row_data[col_idx]
            if isinstance(missing_val, (int, float)):
                missing_counts_aggregation[var][int(missing_val)] += 1

    table_header = ['Missing data'] + variables_to_track
    table_data = [table_header]

    for i in range(max_missing_count + 1):
        row = [i]
        for var in variables_to_track:
            row.append(missing_counts_aggregation[var][i])
        table_data.append(row)
    
    return table_data


def combine_and_chart_data(input_folder, output_combined_filepath, output_summary_filepath):
    variables_for_summary = ["currentTrms", "currentAc", "currentDc", "activePowerTotal", "activeEnergyTotal"]

    print(f"1. Calculating missing readings summary from files in '{input_folder}'")
    missing_data_for_excel, total_slave_files = calculate_missing_readings(input_folder, variables_for_summary)
    print(f"    Processed data from {total_slave_files} individual slave files.")
    
    print(f"2. Creating combined Excel file at '{output_combined_filepath}'")
    combined_wb = openpyxl.Workbook()
    
    if 'Sheet' in combined_wb.sheetnames:
        combined_wb.remove(combined_wb['Sheet'])

    summary_wb = openpyxl.Workbook()
    if 'Sheet' in summary_wb.sheetnames:
        summary_wb.remove(summary_wb['Sheet'])

    if missing_data_for_excel and len(missing_data_for_excel) > 1:
        summary_ws_combined = combined_wb.create_sheet("Summary_Charts", 0)
        summary_ws_separate = summary_wb.create_sheet("Summary_Charts", 0)
        
        present_data_for_excel = []
        present_header_row = ['Timestamp'] + [f"{var}" for var in variables_for_summary]
        present_data_for_excel.append(present_header_row)

        for row_idx in range(1, len(missing_data_for_excel)):
            original_row = missing_data_for_excel[row_idx]
            timestamp = original_row[0]
            new_row = [timestamp]
            for val in original_row[1:]:
                if isinstance(val, (int, float)):
                    present_count = total_slave_files - int(val)
                    new_row.append(present_count)
                else:
                    new_row.append(val)
            present_data_for_excel.append(new_row)

        for row_data in present_data_for_excel:
            summary_ws_combined.append(row_data)
            summary_ws_separate.append(row_data)

        num_summary_rows = summary_ws_combined.max_row
        if num_summary_rows > 1:
            for idx, var_name in enumerate(variables_for_summary):
                var_col_for_summary = idx + 2
                create_scatter_chart(
                    summary_ws_combined,
                    timestamp_col=1,
                    var_col=var_col_for_summary,
                    num_rows=num_summary_rows,
                    var_name=f"{var_name}",
                    chart_title=f"Missing data reading for {var_name}",
                    y_axis_label="Expected - missing readings",
                    chart_index=idx,
                    total_slave_files=total_slave_files
                )

                create_scatter_chart(
                    summary_ws_separate,
                    timestamp_col=1,
                    var_col=var_col_for_summary,
                    num_rows=num_summary_rows,
                    var_name=f"{var_name}",
                    chart_title=f"Missing data reading for {var_name}",
                    y_axis_label="Expected - missing readings",
                    chart_index=idx,
                    total_slave_files=total_slave_files
                )
            print("    'Summary_Charts' sheet created with charts.")
            
            missing_counts_table_data = generate_missing_counts_table_data(missing_data_for_excel, variables_for_summary, max_missing_count=25)
            if missing_counts_table_data:
                next_empty_row_combined = summary_ws_combined.max_row + 2
                next_empty_row_separate = summary_ws_separate.max_row + 2

                summary_ws_combined.cell(row=next_empty_row_combined, column=1, value="Summary of frequency of missing data")
                summary_ws_separate.cell(row=next_empty_row_separate, column=1, value="Summary of frequency of missing data")
                next_empty_row_combined += 2
                next_empty_row_separate += 2

                for row_data in missing_counts_table_data:
                    summary_ws_combined.append(row_data)
                    summary_ws_separate.append(row_data)
                print("    Summary of frequency of missing data added to 'Summary_Charts' sheet.")

        else:
            print("    Not enough data to create summary charts on 'Summary_Charts' sheet.")
    else:
        print("    No valid data found or processed to create 'Summary_Charts' sheet.")

    units = {
        "currentTrms": "(A)",
        "currentAc": "(A)",
        "currentDc": "(A)",
        "activePowerTotal": "(W)",
        "activeEnergyTotal": "(kWh)"
    }

    processed_individual_files_count = 0
    print(f"3. Adding individual slave sheets...")
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(input_folder, filename)
            try:
                slave_wb = openpyxl.load_workbook(file_path, data_only=True)
                slave_ws = slave_wb.active
                
                sheet_name = os.path.splitext(filename)[0] 
                
                new_ws = combined_wb.create_sheet(sheet_name)
                for row in slave_ws.iter_rows(values_only=True):
                    new_ws.append(row)
                
                processed_individual_files_count += 1
                print(f"    Sheet '{sheet_name}' added to combined file.")

            except Exception as e:
                print(f"    Error processing individual file '{filename}': {e}")
                continue
    
    if processed_individual_files_count == 0:
        print("    No individual slave files were successfully processed and added to the combined file.")

    try:
        combined_wb.save(output_combined_filepath)
        print(f"\nOperation complete. The combined Excel file has been saved to: '{output_combined_filepath}'")
    except Exception as e:
        print(f"\nError saving the combined Excel file: {e}")
        return None
    
    try:
        summary_wb.save(output_summary_filepath)
        print(f"Summary charts saved to: '{output_summary_filepath}'")
    except Exception as e:
        print(f"Error saving summary charts to separate file: {e}")

    return output_combined_filepath


if __name__ == "__main__":
    input_folder_path = r'C:\Users\ITGUSAP\Desktop\Monitoring\HUB and CurrentSensor\2 INS-HUB\modbus_poll_log_individual_2hub' 
    output_combined_excel_path = r'C:\Users\ITGUSAP\Desktop\Monitoring\HUB and CurrentSensor\2 INS-HUB\combined_slaves_data.xlsx'
    output_summary_excel_path = r'C:\Users\ITGUSAP\Desktop\Monitoring\HUB and CurrentSensor\2 INS-HUB\summary_charts_data.xlsx'

    output_dir = os.path.dirname(output_combined_excel_path)
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"Created output directory: {output_dir}")
        except OSError as e:
            print(f"Error: Could not create output directory '{output_dir}': {e}")
            exit()
            
    summary_output_dir = os.path.dirname(output_summary_excel_path)
    if not os.path.exists(summary_output_dir):
        try:
            os.makedirs(summary_output_dir)
            print(f"Created summary output directory: {summary_output_dir}")
        except OSError as e:
            print(f"Error: Could not create summary output directory '{summary_output_dir}': {e}")
            exit()


    try:
        print("Starting the process of combining individual Excel files and generating charts...")
        result_file_path = combine_and_chart_data(input_folder_path, output_combined_excel_path, output_summary_excel_path)
        
        if result_file_path:
            print(f"\nSuccess: Combined Excel file generated and saved at: {result_file_path}")
        else:
            print("\nOperation failed: The combined Excel file could not be generated.")
            
    except KeyboardInterrupt:
        print("\nProcess interrupted by the user.")
    except Exception as e:
        print(f"\nAn unexpected error occurred during the main process: {e}")