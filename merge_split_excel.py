# Run on the folder: python merge_split_excel.py
# Edit the lines with the correct paths and original file names

import os
import zipfile
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from copy import copy
import re

def copy_sheet_without_charts(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col_dim in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_dim] = copy(source_sheet.column_dimensions[col_dim])
    for row_dim in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row_dim] = copy(source_sheet.row_dimensions[row_dim])

def extract_slave_id(sheet_name):
    match = re.search(r'\d+', sheet_name)
    return match.group(0) if match else "?"

def create_scatter_chart(sheet, timestamp_col, var_col, num_rows, var_name, slave_id, chart_index):
    units = {
        "currentTrms": "(A)",
        "currentAc": "(A)",
        "currentDc": "(A)",
        "activePowerTotal": "(W)",
        "activeEnergyTotal": "(kWh)"
    }

    chart = ScatterChart()
    chart.title = f"{var_name} over time - ID {slave_id}"
    chart.style = 13
    chart.x_axis.title = "Timestamp"
    chart.y_axis.title = f"{var_name} {units.get(var_name, '')}".strip()
    chart.legend = None

    chart.x_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines()
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 500

    x_values = Reference(sheet, min_col=timestamp_col, min_row=2, max_row=num_rows)
    y_values = Reference(sheet, min_col=var_col, min_row=2, max_row=num_rows)

    title_with_unit = f"{var_name} {units.get(var_name, '')}".strip()
    series = Series(values=y_values, xvalues=x_values, title=title_with_unit)
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.smooth = True
    chart.series.append(series)

    chart.x_axis.axId = 10 + chart_index * 2
    chart.y_axis.axId = 20 + chart_index * 2
    chart.y_axis.crossAx = 10 + chart_index * 2
    chart.x_axis.crossAx = 20 + chart_index * 2

    chart.height = 9.5
    chart.width = 20.44

    anchor_col = 'H'
    anchor_row = 1 + 20 * chart_index
    chart.anchor = f"{anchor_col}{anchor_row}"

    sheet.add_chart(chart)

def merge_data_and_generate_charts(path_modbus_log, path_backup_data, output_folder="merged_sheets"):
    if not os.path.exists(path_modbus_log) or not os.path.exists(path_backup_data):
        print("Error: One of the files was not found.")
        return

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    else:
        for file_name in os.listdir(output_folder):
            os.remove(os.path.join(output_folder, file_name))
        print(f"Folder '{output_folder}' cleaned.")

    print("Loading Excel files...")
    wb_modbus = openpyxl.load_workbook(path_modbus_log)
    wb_backup = openpyxl.load_workbook(path_backup_data)

    for sheet_name in wb_modbus.sheetnames:
        if sheet_name not in wb_backup.sheetnames:
            print(f"Sheet '{sheet_name}' not found in backup. Skipping...")
            continue

        print(f"Processing sheet '{sheet_name}'...")

        modbus_sheet = wb_modbus[sheet_name]
        backup_sheet = wb_backup[sheet_name]

        columns = ["timestamp", "currentTrms", "currentAc", "currentDc", "activePowerTotal", "activeEnergyTotal"]

        modbus_data = pd.DataFrame(modbus_sheet.iter_rows(min_row=2, max_col=6, values_only=True), columns=columns)
        backup_data = pd.DataFrame(backup_sheet.iter_rows(min_row=2, max_col=6, values_only=True), columns=columns)

        merged_data = pd.concat([backup_data, modbus_data], ignore_index=True)
        merged_data.dropna(subset=["timestamp"], inplace=True)
        merged_data.sort_values("timestamp", inplace=True)
        merged_data.reset_index(drop=True, inplace=True)

        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        new_sheet = new_wb.create_sheet(title=sheet_name)

        copy_sheet_without_charts(modbus_sheet, new_sheet)

        if new_sheet.max_row > 1:
            new_sheet.delete_rows(2, new_sheet.max_row)

        for r_idx, row in merged_data.iterrows():
            for c_idx, value in enumerate(row, start=1):
                new_sheet.cell(row=r_idx + 2, column=c_idx, value=value)

        num_rows = merged_data.shape[0] + 1
        slave_id = extract_slave_id(sheet_name)

        columns_to_chart = {
            2: "currentTrms",
            3: "currentAc",
            4: "currentDc",
            5: "activePowerTotal",
            6: "activeEnergyTotal"
        }

        for chart_index, (col_idx, var_name) in enumerate(columns_to_chart.items()):
            create_scatter_chart(
                new_sheet, 
                timestamp_col=1, 
                var_col=col_idx, 
                num_rows=num_rows, 
                var_name=var_name, 
                slave_id=slave_id,
                chart_index=chart_index
            )

        safe_name = sheet_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        path_output = os.path.join(output_folder, f"{safe_name}.xlsx")
        new_wb.save(path_output)
        print(f"Sheet '{sheet_name}' saved as '{path_output}'.")

    zip_path = os.path.join(os.path.dirname(path_modbus_log), "merged_slaves.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(output_folder):
            zipf.write(os.path.join(output_folder, file), file)

    print(f"\nAll individual files have been compressed into: {zip_path}")
    return zip_path

if __name__ == "__main__":
    path_modbus = r'C:\Users\ITGUSAP\Desktop\Monitoring\modbus_poll_log.xlsx'   # Edit this line
    path_backup = r'C:\Users\ITGUSAP\Desktop\Monitoring\backup_poll_data.xlsx'  # Edit this line
    try:
        merge_data_and_generate_charts(path_modbus, path_backup)
    except KeyboardInterrupt:
        print("\nProcess interrupted")